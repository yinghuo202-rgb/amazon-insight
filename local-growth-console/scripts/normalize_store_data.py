from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
PRODUCT_DETAIL_FILE = ROOT / "产品明细表-一店.xlsx"
COST_FILE = ROOT / "产品总成本计算表8.21.xlsx"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_sku(value: Any) -> str:
    return clean_text(value).upper()


def to_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    if isinstance(value, Decimal):
        number = float(value)
        return int(number) if number.is_integer() else number

    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def to_bool_flag(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = clean_text(value).lower()
    if text in {"1", "true", "yes", "y", "是", "有"}:
        return True
    if text in {"0", "false", "no", "n", "否", "无"}:
        return False
    return bool(text)


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def get_headers(ws) -> list[str]:
    values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [clean_text(value) for value in values]


def infer_product_profile(title_cn: str, category_cn: str) -> tuple[str, str, list[str]]:
    text = f"{title_cn} {category_cn}"
    rules = [
        (["压力表", "测压表", "工程表"], "pressure_gauge", "pressure_measurement"),
        (["减压阀", "调压阀"], "pressure_regulator", "water_pressure_control"),
        (["园艺水管", "花园水管", "水管"], "garden_hose", "garden_watering"),
        (["园艺接头", "水管接头", "接头"], "hose_connector", "garden_watering"),
        (["防冻配件", "防冻", "冬季"], "freeze_protection", "winterization"),
        (["罩子", "保护罩", "盖"], "cover", "protection"),
        (["阀门", "球阀", "阀"], "valve", "flow_control"),
        (["过滤器", "滤芯", "滤网"], "filter", "water_filtration"),
        (["烧烤炉配件", "烧烤", "烤炉"], "grill_accessory", "outdoor_cooking"),
    ]

    matched_keywords: list[str] = []
    for keywords, product_type, sub_scenario in rules:
        if any(keyword in text for keyword in keywords):
            matched_keywords.extend([keyword for keyword in keywords if keyword in text])
            return product_type, sub_scenario, sorted(set(matched_keywords + [category_cn, product_type, sub_scenario]))

    fallback_type = re.sub(r"\W+", "_", category_cn.lower()).strip("_") or "general_product"
    return fallback_type, "general_store_product", sorted(set(filter(None, [category_cn, fallback_type, title_cn[:12]])))


def price_band(price: int | float | None) -> str:
    if price is None:
        return "unknown"
    if price < 20:
        return "under_20"
    if price <= 80:
        return "20_80"
    return "over_80"


def profit_band(profit: int | float | None, margin: int | float | None) -> str:
    if profit is None and margin is None:
        return "unknown"
    if profit is not None and profit < 0:
        return "negative"
    if margin is not None:
        if margin < 0:
            return "negative"
        if margin < 0.15:
            return "low"
        if margin < 0.35:
            return "medium"
        return "high"
    if profit is not None:
        if profit < 2:
            return "low"
        if profit < 8:
            return "medium"
        return "high"
    return "unknown"


def size_risk(product_weight_g: int | float | None, gross_weight_kg: int | float | None, carton_volume_cbm: int | float | None) -> str:
    if (gross_weight_kg is not None and gross_weight_kg >= 15) or (carton_volume_cbm is not None and carton_volume_cbm >= 0.08):
        return "high"
    if (gross_weight_kg is not None and gross_weight_kg >= 5) or (product_weight_g is not None and product_weight_g >= 1000) or (carton_volume_cbm is not None and carton_volume_cbm >= 0.035):
        return "medium"
    return "low"


def load_product_details() -> list[dict[str, Any]]:
    wb = load_workbook(PRODUCT_DETAIL_FILE, read_only=True, data_only=True)
    ws = wb["一店"]
    records: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = normalize_sku(row[5] if len(row) > 5 else None)
        if not sku:
            continue

        title_cn = clean_text(row[2] if len(row) > 2 else None)
        category_cn = clean_text(row[3] if len(row) > 3 else None)
        product_type, sub_scenario, keywords = infer_product_profile(title_cn, category_cn)

        records.append({
            "sku": sku,
            "asin": "",
            "parent_asin": "",
            "title_cn": title_cn,
            "amazon_title": "",
            "brand": "",
            "category_cn": category_cn,
            "product_type": product_type,
            "sub_scenario": sub_scenario,
            "packaging": clean_text(row[4] if len(row) > 4 else None),
            "carton_qty": to_number(row[6] if len(row) > 6 else None),
            "net_weight_kg": to_number(row[7] if len(row) > 7 else None),
            "gross_weight_kg": to_number(row[8] if len(row) > 8 else None),
            "carton_length_cm": to_number(row[9] if len(row) > 9 else None),
            "carton_width_cm": to_number(row[10] if len(row) > 10 else None),
            "carton_height_cm": to_number(row[11] if len(row) > 11 else None),
            "carton_volume_cbm": to_number(row[12] if len(row) > 12 else None),
            "product_weight_g": to_number(row[13] if len(row) > 13 else None),
            "product_size_text": clean_text(row[14] if len(row) > 14 else None),
            "purchase_cost_rmb_tax_included": to_number(row[15] if len(row) > 15 else None),
            "purchase_cost_rmb_ex_tax": to_number(row[16] if len(row) > 16 else None),
            "exchange_rate": to_number(row[18] if len(row) > 18 else None),
            "cost_usd_ex_tax": to_number(row[19] if len(row) > 19 else None),
            "keywords": keywords,
            "status": "active",
            "source": "产品明细表-一店.xlsx#一店",
        })

    wb.close()
    return records


def load_sheet_records(sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(COST_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = get_headers(ws)
    records = [row_dict(headers, row) for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return records


def load_us_cost_profile() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in load_sheet_records("美国"):
        sku = normalize_sku(row.get("SKU"))
        if not sku:
            continue
        records.append({
            "sku": sku,
            "marketplace": "US",
            "current_price": to_number(row.get("现价")),
            "product_cost_usd": to_number(row.get("美金价")),
            "cost_usd_ex_tax": to_number(row.get("未税价格")),
            "import_tax_rate": to_number(row.get("进口税率")),
            "import_tax": to_number(row.get("进口税")),
            "product_length_cm": to_number(row.get("尺寸cm：长")),
            "product_width_cm": to_number(row.get("尺寸cm：宽")),
            "product_height_cm": to_number(row.get("尺寸cm：高")),
            "gross_weight_kg": to_number(row.get("毛重(kg)")),
            "product_volume_cbm": to_number(row.get("产品体积（m³）")),
            "carton_qty": to_number(row.get("装箱数")),
            "carton_volume_cbm": to_number(row.get("外箱体积（m³）")),
            "inbound_fba_shipping_cost": to_number(row.get("海运入FBA仓成本")),
            "fba_storage_fee": to_number(row.get("FBA仓储费")),
            "fulfillment_fee": to_number(row.get("订单处理费")),
            "referral_fee_rate": to_number(row.get("佣金比例")),
            "referral_fee": to_number(row.get("佣金")),
            "estimated_profit": to_number(row.get("利润")),
            "estimated_profit_margin": to_number(row.get("利润率")),
            "cost_note": clean_text(row.get("备注")),
            "competitor_asin": clean_text(row.get("竞品ASIN")),
        })
    return records


def load_ca_cost_profile() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in load_sheet_records("加拿大"):
        sku = normalize_sku(row.get("SKU"))
        if not sku:
            continue
        records.append({
            "sku": sku,
            "marketplace": "CA",
            "current_price": to_number(row.get("现价")),
            "us_price_reference": to_number(row.get("美国售价")),
            "fob_cost_usd": to_number(row.get("FOB价")),
            "carton_qty": to_number(row.get("装箱数")),
            "carton_volume_cbm": to_number(row.get("外箱体积（m³）")),
            "inbound_fba_shipping_cost": to_number(row.get("海运入FBA仓成本")),
            "landed_cost": to_number(row.get("FOB+头程")),
            "fulfillment_fee": to_number(row.get("订单处理费")),
            "referral_fee_rate": to_number(row.get("佣金比例")),
            "referral_fee": to_number(row.get("佣金")),
            "estimated_profit": to_number(row.get("利润")),
            "estimated_profit_margin": to_number(row.get("利润率")),
            "cost_note": clean_text(row.get("备注")),
            "competitor_asin": clean_text(row.get("竞品ASIN")),
            "slow_moving_flag": to_bool_flag(row.get("是否是滞销品")),
        })
    return records


def load_sales_snapshot() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in load_sheet_records("Sheet1"):
        sku = normalize_sku(row.get("MSKU"))
        if not sku:
            continue
        records.append({
            "sku": sku,
            "current_price": to_number(row.get("售价")),
            "prime_discount_price": to_number(row.get("Prime折扣价")),
            "promo_profit": to_number(row.get("利润")),
            "promo_margin": to_number(row.get("毛利率")),
            "inventory_units": to_number(row.get("库存")),
            "monthly_sales_units": to_number(row.get("月销")),
            "prime_flag": to_bool_flag(row.get("prime")),
        })
    return records


def by_sku(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result = {}
    for record in records:
        sku = record.get("sku")
        if sku and sku not in result:
            result[sku] = record
    return result


def merge_profiles(
    products: list[dict[str, Any]],
    us_costs: list[dict[str, Any]],
    sales: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    us_by_sku = by_sku(us_costs)
    sales_by_sku = by_sku(sales)
    product_skus = {record["sku"] for record in products}

    merged: list[dict[str, Any]] = []
    for product in products:
        sku = product["sku"]
        us = us_by_sku.get(sku, {})
        sale = sales_by_sku.get(sku, {})
        current_price = us.get("current_price") if us.get("current_price") is not None else sale.get("current_price")
        profit = us.get("estimated_profit")
        margin = us.get("estimated_profit_margin")

        merged.append({
            "sku": sku,
            "asin": product.get("asin", ""),
            "parent_asin": product.get("parent_asin", ""),
            "title_cn": product.get("title_cn", ""),
            "amazon_title": product.get("amazon_title", ""),
            "brand": product.get("brand", ""),
            "category_cn": product.get("category_cn", ""),
            "product_type": product.get("product_type", ""),
            "sub_scenario": product.get("sub_scenario", ""),
            "current_price_usd": current_price,
            "cost_usd": us.get("product_cost_usd") if us.get("product_cost_usd") is not None else product.get("cost_usd_ex_tax"),
            "estimated_profit_usd": profit,
            "estimated_profit_margin": margin,
            "monthly_sales_units": sale.get("monthly_sales_units"),
            "inventory_units": sale.get("inventory_units"),
            "product_weight_g": product.get("product_weight_g"),
            "gross_weight_kg": us.get("gross_weight_kg") if us.get("gross_weight_kg") is not None else product.get("gross_weight_kg"),
            "carton_volume_cbm": us.get("carton_volume_cbm") if us.get("carton_volume_cbm") is not None else product.get("carton_volume_cbm"),
            "packaging": product.get("packaging", ""),
            "competitor_asin": us.get("competitor_asin", ""),
            "keywords": product.get("keywords", []),
            "price_band": price_band(current_price),
            "profit_band": profit_band(profit, margin),
            "size_risk": size_risk(product.get("product_weight_g"), us.get("gross_weight_kg") if us.get("gross_weight_kg") is not None else product.get("gross_weight_kg"), us.get("carton_volume_cbm") if us.get("carton_volume_cbm") is not None else product.get("carton_volume_cbm")),
            "store_existing_product": True,
        })

    unmatched = {
        "products_missing_us_cost": sorted(product_skus - set(us_by_sku)),
        "products_missing_sales_snapshot": sorted(product_skus - set(sales_by_sku)),
        "us_cost_without_product_detail": sorted(set(us_by_sku) - product_skus),
        "sales_snapshot_without_product_detail": sorted(set(sales_by_sku) - product_skus),
    }
    return merged, unmatched


def write_json(path: Path, records: Any) -> None:
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def log_counts(name: str, records: list[dict[str, Any]]) -> None:
    sku_counts = Counter(record.get("sku") for record in records if record.get("sku"))
    duplicate_skus = sorted(sku for sku, count in sku_counts.items() if count > 1)
    print(f"{name}: {len(records)} rows, {len(sku_counts)} unique skus")
    if duplicate_skus:
        print(f"  duplicate skus ({len(duplicate_skus)}): {', '.join(duplicate_skus[:20])}")


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    products = load_product_details()
    us_costs = load_us_cost_profile()
    ca_costs = load_ca_cost_profile()
    sales = load_sales_snapshot()
    merged, unmatched = merge_profiles(products, us_costs, sales)

    write_json(DATA_DIR / "store_products.json", products)
    write_json(DATA_DIR / "store_cost_profile_us.json", us_costs)
    write_json(DATA_DIR / "store_cost_profile_ca.json", ca_costs)
    write_json(DATA_DIR / "store_sales_snapshot.json", sales)
    write_json(DATA_DIR / "store_product_profile_merged.json", merged)

    print(f"normalized_at: {datetime.now().isoformat(timespec='seconds')}")
    log_counts("store_products", products)
    log_counts("store_cost_profile_us", us_costs)
    log_counts("store_cost_profile_ca", ca_costs)
    log_counts("store_sales_snapshot", sales)
    log_counts("store_product_profile_merged", merged)
    print("unmatched_skus:")
    for key, values in unmatched.items():
        preview = ", ".join(values[:30])
        suffix = " ..." if len(values) > 30 else ""
        print(f"  {key}: {len(values)}" + (f" [{preview}{suffix}]" if values else ""))


if __name__ == "__main__":
    main()
