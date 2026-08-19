import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from store_ops.uploaded_data import inspect_batch, publish_batch, restore_version


def dashboard(market: str):
    return {
        "schemaVersion": 2,
        "generatedAt": "2026-08-01T00:00:00+00:00",
        "market": market,
        "summary": {"fbaSellable": 5, "localInventory": 10},
        "rows": [{"sku": "MA001", "productName": "旧名称", "fbaSellable": 5, "localInventory": 10, "pendingOrderQty": 2, "domesticSupplyTotal": 12, "dailySales": 1, "salesHistoryByMonth": [], "salesByMonth": []}],
    }


class UploadedDataTests(unittest.TestCase):
    def test_shipment_upload_adds_sku_history_to_document_master(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            batch = root / "batch-ship123456"
            source = batch / "source"
            reports = root / "reports"
            snapshots = root / "snapshots"
            source.mkdir(parents=True)
            reports.mkdir()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Measureman"
            sheet["B2"] = "产品编号"
            sheet["B3"] = "MA001"
            sheet["D3"] = 10
            sheet["K3"] = 120
            workbook.save(source / "CM320-US2026.7.10发货清单.xlsx")
            (reports / "document_master.json").write_text(json.dumps({
                "generatedAt": "2026-08-01T00:00:00+00:00",
                "logistics": {"US": {}, "CA": {}},
                "shipmentHistory": [],
                "coverage": {"shipmentHistoryEvents": 0},
                "sources": {"shipments": []},
            }), encoding="utf-8")

            preview = inspect_batch(batch)

            self.assertEqual(preview["files"][0]["type"], "shipment")
            self.assertTrue(preview["files"][0]["publishable"])
            self.assertEqual(preview["files"][0]["preview"]["shipmentEventCount"], 1)
            published = publish_batch(batch, reports, snapshots)
            self.assertEqual(published["updatedReports"], ["document_master.json"])
            report = json.loads((reports / "document_master.json").read_text())
            self.assertEqual(report["coverage"]["shipmentHistoryEvents"], 1)
            self.assertEqual(report["shipmentHistory"][0]["sku"], "MA001")
            self.assertEqual(report["shipmentHistory"][0]["quantity"], 120)

    def test_inventory_upload_previews_publishes_and_restores(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            batch = root / "batch-test123456"
            source = batch / "source"
            reports = root / "reports"
            snapshots = root / "snapshots"
            source.mkdir(parents=True)
            reports.mkdir()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "库存规划"
            sheet.append(["SKU", "品名", "工厂库存及已下订单", "FBA库存", "FBA+在途库存", "最近月销售"])
            sheet.append(["MA001", "新名称", 30, 20, 50, 60])
            ca = workbook.create_sheet("加拿大库存计划 ")
            ca.append(["SKU", "品名", "工厂库存及已下订单", "FBA库存", "FBA+在途库存", "最近月销"])
            ca.append(["MA001", "新名称", 25, 15, 45, 30])
            workbook.save(source / "库存规划20260813.xlsx")
            (reports / "inventory_dashboard.json").write_text(json.dumps(dashboard("US")), encoding="utf-8")
            (reports / "inventory_dashboard.ca.json").write_text(json.dumps(dashboard("CA")), encoding="utf-8")

            preview = inspect_batch(batch)
            self.assertEqual(preview["summary"]["recognizedCount"], 1)
            self.assertEqual(preview["files"][0]["preview"]["markets"]["US"]["skuCount"], 1)

            published = publish_batch(batch, reports, snapshots)
            self.assertEqual(set(published["updatedReports"]), {"inventory_dashboard.json", "inventory_dashboard.ca.json"})
            current = json.loads((reports / "inventory_dashboard.json").read_text())
            self.assertEqual(current["rows"][0]["fbaSellable"], 20)
            self.assertEqual(current["rows"][0]["inTransitInventory"], 30)
            self.assertEqual(current["rows"][0]["localInventory"], 30)
            self.assertEqual(current["rows"][0]["salesByMonth"][0]["units"], 60)
            self.assertEqual(current["snapshots"]["fbaDate"], "2026-08-13")
            self.assertEqual(current["parameters"]["targetCoverDays"], 90)

            restored = restore_version(published["dataVersion"], reports, snapshots)
            self.assertIn("inventory_dashboard.json", restored["restoredReports"])
            previous = json.loads((reports / "inventory_dashboard.json").read_text())
            self.assertEqual(previous["rows"][0]["fbaSellable"], 5)

    def test_advertising_upload_replaces_month_and_publishes_latest_campaigns(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            batch = root / "batch-ads123456"
            source = batch / "source"
            reports = root / "reports"
            snapshots = root / "snapshots"
            source.mkdir(parents=True)
            reports.mkdir()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "按月"
            sheet.append(["数据开始时间", "数据结束时间", "国家", "广告活动", "状态", "预算", "曝光量", "点击量", "花费", "广告总销售额", "广告总订单量"])
            sheet.append(["2026-08-01", "2026-08-11", "US", "MA001 核心词", "正在投放", 20, 1000, 50, 25, 100, 5])
            workbook.save(source / "广告活动报表-按月-20260813.xlsx")
            current = dashboard("US")
            current["snapshots"] = {"fbaDate": "2026-08-13"}
            current["rows"][0]["riskLevel"] = "healthy"
            current["rows"][0]["daysCoverNetwork"] = 90
            current["advertising"] = {
                "parameters": {"targetAcosPercent": 30},
                "latestMonth": "2026-07",
                "monthlySeries": [{"month": "2026-07", "spend": 10, "advertisingSales": 50, "orders": 2, "clicks": 20, "impressions": 500, "acos": 20, "roas": 5}],
                "campaigns": [],
            }
            current["sources"] = []
            (reports / "inventory_dashboard.json").write_text(json.dumps(current), encoding="utf-8")

            preview = inspect_batch(batch)
            self.assertTrue(preview["files"][0]["publishable"])
            published = publish_batch(batch, reports, snapshots)
            self.assertEqual(published["updatedReports"], ["inventory_dashboard.json"])
            updated = json.loads((reports / "inventory_dashboard.json").read_text())
            self.assertEqual(updated["advertising"]["latestMonth"], "2026-08")
            self.assertEqual(updated["advertising"]["monthlySeries"][-1]["spend"], 25)
            self.assertEqual(updated["advertising"]["campaigns"][0]["sku"], "MA001")


if __name__ == "__main__":
    unittest.main()
