import unittest

import openpyxl

from store_ops.jobs.new_product_research import extract_research_rows


class NewProductResearchExtractionTests(unittest.TestCase):
    def test_reads_all_cost_table_sheets_and_name_only_month_rows(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["NP001", None, "亚马逊售价", "头程", "FBA仓储费", "佣金", "订单处理费", "HS编码", "进口关税", "采购价"])
        sheet.append([None, None, 39.99, 2, 0.5, 6, 8, None, 1, 72])
        cost = workbook.create_sheet("汽车")
        cost.append([None, "现价", "美金成本价", "未税价格", "进口税率", "进口税", "装箱数", "外箱体积（m³）", "海运入FBA仓成本", "FBA仓储费", "订单处理费", "佣金比例", "佣金", "利润", "利润率", "ASIN", "竞品链接"])
        cost.append([None, 49.99, 10, 8.85, 0.1, 1, 20, 0.02, 1, 0.2, 5, 0.15, 7.5, 25.29, 0.5, "B0123456789", "https://www.amazon.com/dp/B0123456789"])
        monthly = workbook.create_sheet("7月新品")
        monthly.append(["7月下单新品"])
        monthly.append(["序号", "SKU", "图片", "产品名称", "下单数量", "含税价"])
        monthly.append([1, None, None, "无 SKU 新品", 20, 12])

        candidates, orders = extract_research_rows(workbook)

        self.assertEqual(len(candidates), 2)
        self.assertEqual(candidates[1]["sku"], "B0123456789")
        self.assertEqual(candidates[1]["purchaseCostRmb"], 72)
        self.assertAlmostEqual(candidates[1]["totalCostUsd"], 33.55)
        self.assertAlmostEqual(candidates[1]["grossProfit"], 16.44)
        self.assertAlmostEqual(candidates[1]["grossMargin"], 16.44 / 49.99)
        self.assertEqual(len(orders), 1)
        self.assertEqual(orders[0]["sku"], "2026-07-ROW-3")


if __name__ == "__main__":
    unittest.main()
