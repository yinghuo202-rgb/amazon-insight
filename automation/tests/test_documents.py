import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from store_ops.config import ProjectConfig
from store_ops.documents import PurchaseOrderLot, PurchaseOrderShortage, allocate_purchase_orders_fifo
from store_ops.jobs.document_master import _market, _read_payment_ledger, _read_shipments
from store_ops.sku import SkuNormalizer
from store_ops.jobs.export_documents import _chunk_rows, _declaration_layout, _declaration_rows, _validate_request


def lot(number: str, day: str, ordered: int, shipped: int = 0) -> PurchaseOrderLot:
    return PurchaseOrderLot(number, date.fromisoformat(day), "MA001", "可名", ordered, shipped, 31.0, "压力表")


class PurchaseOrderAllocationTests(unittest.TestCase):
    def test_market_recognizes_ca_immediately_followed_by_date(self):
        self.assertEqual(_market(Path("CM321/CM321-CA2026.7.14发货清单.xlsx")), "CA")

    def test_market_recognizes_ca_immediately_followed_by_chinese_text(self):
        self.assertEqual(_market(Path("CM321/CM321-CA亚马逊报运单.xlsx")), "CA")

    def test_market_keeps_us_shipment_as_us(self):
        self.assertEqual(_market(Path("CM320/CM320-US-20260713发货清单.xlsx")), "US")

    def test_document_rows_are_paginated_without_limiting_business_batch(self):
        rows = [{"sku": f"MA{index:03d}"} for index in range(1, 88)]
        parts = _chunk_rows(rows, 43)
        self.assertEqual([len(part) for part in parts], [43, 43, 1])
        self.assertEqual([item["sku"] for part in parts for item in part], [item["sku"] for item in rows])

    def test_export_entries_sort_by_sku_in_requested_direction(self):
        base = {"market": "US", "documentTypes": ["shipment"], "entries": [
            {"sku": "MC003", "quantity": 100},
            {"sku": "MA049", "quantity": 100},
            {"sku": "MA007", "quantity": 100},
        ]}
        ascending = _validate_request({**base, "skuSort": "asc"})[2]
        descending = _validate_request({**base, "skuSort": "desc"})[2]
        self.assertEqual([item["sku"] for item in ascending], ["MA007", "MA049", "MC003"])
        self.assertEqual([item["sku"] for item in descending], ["MC003", "MA049", "MA007"])

    def test_allocates_oldest_available_order_first(self):
        result = allocate_purchase_orders_fifo("MA001", 120, [
            lot("AM260403-6", "2026-04-03", 100),
            lot("AM260320-9", "2026-03-20", 100),
        ])
        self.assertEqual([(item.po_number, item.quantity) for item in result], [
            ("AM260320-9", 100),
            ("AM260403-6", 20),
        ])

    def test_uses_only_remaining_quantity(self):
        result = allocate_purchase_orders_fifo("MA001", 40, [
            lot("AM260320-9", "2026-03-20", 100, 75),
            lot("AM260403-6", "2026-04-03", 100),
        ])
        self.assertEqual(result[0].quantity, 25)
        self.assertEqual(result[1].quantity, 15) if len(result) > 1 else None

    def test_is_deterministic_for_same_date(self):
        result = allocate_purchase_orders_fifo("MA001", 20, [
            lot("AM260320-10", "2026-03-20", 20),
            lot("AM260320-2", "2026-03-20", 20),
        ])
        self.assertEqual(result[0].po_number, "AM260320-10")

    def test_shortage_blocks_export(self):
        with self.assertRaises(PurchaseOrderShortage) as context:
            allocate_purchase_orders_fifo("MA001", 101, [lot("AM260320-9", "2026-03-20", 100)])
        self.assertEqual(context.exception.available, 100)

    def test_declaration_logistics_totals_follow_shipment_parent_quantity(self):
        rows = _declaration_rows(
            [{"sku": "MA001", "quantity": 100}],
            {"MA001": {"cartonQty": 10, "grossWeightKg": 5, "cartonVolumeM3": 0.1}},
            {"MA001": {"components": [{"declarationSku": "MA001-A", "factory": "可名", "productName": "压力表", "quantityFactor": 2}]}},
            [PurchaseOrderLot("AM260320-9", date(2026, 3, 20), "MA001", "可名", 200, 0, 31.0, "压力表")],
            {"AM260320-9": {"paymentPayers": ["联合"], "paymentMethods": ["含税"]}},
        )
        self.assertEqual(sum(row["quantity"] for row in rows), 200)
        self.assertAlmostEqual(sum(row["cartons"] for row in rows), 10)
        self.assertAlmostEqual(sum(row["weightKg"] for row in rows), 50)
        self.assertAlmostEqual(sum(row["volumeM3"] for row in rows), 1)
        self.assertEqual(rows[0]["note"], "联合已付款\n付款方式：含税\n货出完了")

    def test_declaration_layout_uses_first_visible_duplicate_header(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "template.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "报关空表"
            sheet["E2"] = "出运日期"
            headers = ["合同号", "工厂", "品名", "数量", "箱数", "重量", "体积", "销售金额", "采购金额", "备注", "退税率", "", "品名", "单价", "采购金额"]
            for column, value in enumerate(headers, 1):
                sheet.cell(9, column, value)
            sheet["A20"] = "合计"
            workbook.save(path)
            workbook.close()
            layout = _declaration_layout(path, "US")
        self.assertEqual(layout["columns"]["productName"], 3)
        self.assertEqual(layout["columns"]["purchaseAmountRmb"], 9)
        self.assertEqual(layout["shipmentDateCell"], "E3")

    def test_payment_ledger_continuation_rows_inherit_payment_group(self):
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            ledger = root / "ledger.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["公司名称", "订单号", "订单金额", "付款金额", "付款方式", "付款时间", "付款方", "备注"])
            sheet.append(["可名", "AM260101-1", 100, 50, "含税", date(2026, 1, 2), "联合", "预付款"])
            sheet.append([None, "AM260101-2", 200, 100, None, None, None, None])
            workbook.save(ledger)
            workbook.close()
            config = ProjectConfig(
                config_path=root / "config.json",
                project_root=root,
                data_root=root,
                runtime_root=root / "runtime",
                sku_pattern="",
                ignore_values=frozenset(),
                sources=(),
                inventory_dashboard={"document_master_sources": {"payment_ledger_workbook": "ledger.xlsx"}},
            )
            payments, exceptions, source = _read_payment_ledger(config)
        self.assertEqual(exceptions, [])
        self.assertEqual(source, ledger.resolve())
        self.assertEqual(payments["AM260101-2"]["paymentPayers"], ["联合"])
        self.assertEqual(payments["AM260101-2"]["paymentMethods"], ["含税"])

    def test_shipment_reader_emits_sku_history_with_cartons_and_source(self):
        with TemporaryDirectory() as temporary:
            root = Path(temporary).resolve()
            folder = root / "MEASUREMAN" / "发货清单" / "美国" / "2026" / "CM320"
            folder.mkdir(parents=True)
            path = folder / "CM320-US2026.7.2发货清单.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Measureman"
            sheet["B2"] = "产品编号"
            sheet["B3"] = "MA001"
            sheet["D3"] = 10
            sheet["K3"] = 100
            workbook.save(path)
            workbook.close()
            config = ProjectConfig(
                config_path=root / "config.json",
                project_root=root,
                data_root=root,
                runtime_root=root / "runtime",
                sku_pattern=r"(?<![A-Z0-9])([A-Z]{2}\s*[-_]?\s*\d{3})(?!\d)",
                ignore_values=frozenset(),
                sources=(),
                inventory_dashboard={"document_master_sources": {"shipment_root": "MEASUREMAN/发货清单"}},
            )
            _, _, events, sources, _ = _read_shipments(config, SkuNormalizer(config.sku_pattern, config.ignore_values))
        self.assertEqual(events, [{
            "market": "US",
            "batch": 320,
            "shipmentDate": "2026-07-02",
            "sku": "MA001",
            "quantity": 100,
            "cartonCount": 10,
            "sourcePath": "MEASUREMAN/发货清单/美国/2026/CM320/CM320-US2026.7.2发货清单.xlsx",
        }])
        self.assertEqual(sources, ["MEASUREMAN/发货清单/美国/2026/CM320/CM320-US2026.7.2发货清单.xlsx"])

    def test_incremental_shipment_reader_also_emits_sku_history(self):
        with TemporaryDirectory() as temporary:
            root = Path(temporary).resolve()
            incremental = root / "weekly" / "CM321"
            incremental.mkdir(parents=True)
            path = incremental / "CM321-CA2026.7.14发货清单.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Measureman"
            sheet["B2"] = "产品编号"
            sheet["B3"] = "MA002"
            sheet["D3"] = 8
            sheet["K3"] = 80
            workbook.save(path)
            workbook.close()
            config = ProjectConfig(
                config_path=root / "config.json",
                project_root=root,
                data_root=root,
                runtime_root=root / "runtime",
                sku_pattern=r"(?<![A-Z0-9])([A-Z]{2}\s*[-_]?\s*\d{3})(?!\d)",
                ignore_values=frozenset(),
                sources=(),
                inventory_dashboard={"document_master_sources": {"shipment_root": "missing", "shipment_incremental_root": "weekly"}},
            )

            _, _, events, sources, _ = _read_shipments(config, SkuNormalizer(config.sku_pattern, config.ignore_values))

        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["market"], "CA")
        self.assertEqual(events[0]["batch"], 321)
        self.assertEqual(events[0]["sku"], "MA002")
        self.assertEqual(events[0]["quantity"], 80)
        self.assertEqual(sources, ["weekly/CM321/CM321-CA2026.7.14发货清单.xlsx"])


if __name__ == "__main__":
    unittest.main()
