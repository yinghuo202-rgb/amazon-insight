from __future__ import annotations

import hashlib
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from ..config import ProjectConfig
from ..db import StateDb
from ..sku import SkuNormalizer


REPORT_NAME_RE = re.compile(r"(20\d{2})[._-]?(0?[1-9]|1[0-2])月.*销售和毛利报告-(US|CA|MX)\.xlsx$", re.IGNORECASE)


def _number(value: object) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _latest_reports(root: Path) -> dict[str, tuple[str, Path]]:
    latest: dict[str, tuple[str, float, Path]] = {}
    for path in root.rglob("*销售和毛利报告-*.xlsx"):
        match = REPORT_NAME_RE.search(path.name)
        if not match:
            continue
        month = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
        market = match.group(3).upper()
        marker = path.stat().st_mtime
        current = latest.get(market)
        if current is None or (month, marker) > (current[0], current[1]):
            latest[market] = (month, marker, path)
    return {market: (month, path) for market, (month, _, path) in latest.items()}


def _read_report(path: Path, market: str, report_month: str, normalizer: SkuNormalizer) -> list[dict]:
    aggregate: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "SKU销售汇总" not in workbook.sheetnames:
            raise ValueError(f"Profitability report is missing SKU销售汇总: {path}")
        worksheet = workbook["SKU销售汇总"]
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            match = normalizer.extract(row[0] if row else None)
            if not match or len(match.bases) != 1:
                continue
            sku = match.bases[0]
            item = aggregate[sku]
            item["units"] += max(0, _number(row[1] if len(row) > 1 else None))
            item["productSales"] += _number(row[2] if len(row) > 2 else None)
            item["settlementPayout"] += _number(row[3] if len(row) > 3 else None)
            item["returns"] += max(0, _number(row[5] if len(row) > 5 else None))
            item["landedCost"] += max(0, _number(row[8] if len(row) > 8 else None))
            item["grossProfit"] += _number(row[9] if len(row) > 9 else None)
            item["advertisingCost"] += max(0, -_number(row[11] if len(row) > 11 else None))
            item["storageCost"] += max(0, -_number(row[12] if len(row) > 12 else None))
            item["actualProfit"] += _number(row[13] if len(row) > 13 else None)
    finally:
        workbook.close()

    rows = []
    for sku, item in sorted(aggregate.items()):
        units = int(round(item["units"]))
        returns = min(units, int(round(item["returns"])))
        net_units = max(0, units - returns)
        product_sales = round(item["productSales"], 2)
        actual_profit = round(item["actualProfit"], 2)
        storage_cost = round(item["storageCost"], 2)
        rows.append({
            "market": market,
            "currency": {"CA": "CAD", "MX": "MXN"}.get(market, "USD"),
            "reportMonth": report_month,
            "sku": sku,
            "units": units,
            "returns": returns,
            "netUnits": net_units,
            "productSales": product_sales,
            "settlementPayout": round(item["settlementPayout"], 2),
            "landedCost": round(item["landedCost"], 2),
            "grossProfit": round(item["grossProfit"], 2),
            "advertisingCost": round(item["advertisingCost"], 2),
            "storageCost": storage_cost,
            "actualProfit": actual_profit,
            "currentPrice": round(product_sales / net_units, 2) if net_units > 0 else None,
            "grossMargin": round(item["grossProfit"] / product_sales, 4) if product_sales > 0 else None,
            "actualMargin": round(actual_profit / product_sales, 4) if product_sales > 0 else None,
            "conservativeMargin": round((actual_profit - storage_cost) / product_sales, 4) if product_sales > 0 else None,
        })
    return rows


def run(config: ProjectConfig, db: StateDb) -> dict:
    run_id = db.start_run("build-profitability-data")
    try:
        configured_root = os.environ.get("STORE_OPS_PROFITABILITY_ROOT") or config.inventory_dashboard.get("profitability_report_root")
        if not configured_root:
            raise ValueError("profitability_report_root is not configured")
        report_root = Path(str(configured_root)).expanduser().resolve()
        reports = _latest_reports(report_root)
        if not {"US", "CA"}.issubset(reports):
            raise FileNotFoundError(f"Latest US/CA profitability reports were not both found in {report_root}")
        normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
        sources = []
        rows = []
        markets = [market for market in ("US", "CA", "MX") if market in reports]
        for market in markets:
            report_month, path = reports[market]
            rows.extend(_read_report(path, market, report_month, normalizer))
            sources.append({
                "market": market,
                "reportMonth": report_month,
                "path": str(path),
                "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds"),
                "sha256": _sha256(path),
                "sheet": "SKU销售汇总",
            })
        payload = {
            "schemaVersion": 1,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "method": "按 SKU 汇总产品销售额、结算净额、FOB+头程、广告与仓储；清货期暂停广告，定价按每件平台扣费、FOB+头程和仓储保留 5% 毛利缓冲。",
            "sources": sources,
            "rows": rows,
        }
        report_path = config.runtime_root / "reports" / "profitability.json"
        report_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = report_path.with_suffix(".json.tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(report_path)
        summary = {
            "run_id": run_id,
            "report_path": str(report_path),
            "row_count": len(rows),
            "markets": {market: reports[market][0] for market in markets},
        }
        db.finish_run(run_id, "completed", summary=summary)
        return summary
    except Exception as exc:
        db.finish_run(run_id, "failed", error=repr(exc))
        raise
