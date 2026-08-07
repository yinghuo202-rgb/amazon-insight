from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import defaultdict
from dataclasses import replace
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

from ..advertising import AdvertisingParameters, recommend_campaign
from ..config import ProjectConfig
from ..db import StateDb
from ..replenishment import ReplenishmentParameters, calculate_replenishment
from ..sku import SkuNormalizer
from .document_master import _batch, _market, _read_shipments
from .inventory_dashboard import (
    _read_fba_from_master,
    _read_master,
    _read_sales_from_master,
    _reconcile_domestic_supply,
)
from .purchase_plan import run as run_purchase_plan


INVENTORY_RE = re.compile(r"库存规划(20\d{2})[._-]?(\d{1,2})[._-]?(\d{1,2}).*\.xlsx$", re.IGNORECASE)


def _atomic_json(path: Path, payload: dict) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _inventory_date(path: Path) -> date | None:
    match = INVENTORY_RE.search(path.name)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _latest_inventory_workbook(local_root: Path) -> tuple[Path, date]:
    candidates = []
    for path in local_root.glob("库存规划*.xlsx"):
        business_date = _inventory_date(path)
        if business_date and not path.name.startswith(("~$", ".~")):
            candidates.append((business_date, path.stat().st_mtime, path.stat().st_size, path))
    if not candidates:
        raise FileNotFoundError(f"未在 {local_root} 找到带日期的库存规划工作簿")
    business_date, _, _, path = max(candidates)
    return path, business_date


def _planning_shipment_totals(path: Path) -> dict[tuple[str, int], int]:
    totals: dict[tuple[str, int], int] = {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            market = "CA" if "加拿大" in sheet.title else "US"
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns: dict[int, tuple[str, int]] = {}
            column_totals: dict[int, int] = {}
            for index, value in enumerate(header):
                match = re.search(r"发货[（(]CM\s*0*(\d+)[）)]", str(value or ""), re.IGNORECASE)
                if match:
                    key = (market, int(match.group(1)))
                    columns[index] = key
                    column_totals[index] = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                for index, key in columns.items():
                    cell = row[index] if index < len(row) else None
                    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                        column_totals[index] += int(round(cell))
            for index, key in columns.items():
                totals[key] = max(totals.get(key, 0), column_totals[index])
    finally:
        workbook.close()
    return totals


def _source_meta(config: ProjectConfig, path: Path, kind: str) -> dict:
    return {
        "kind": kind,
        "path": str(path.resolve().relative_to(config.data_root.resolve())).replace("\\", "/"),
        "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds"),
        "sha256": _sha256(path),
    }


def _settings_by_market(config: ProjectConfig) -> dict[str, dict]:
    base = dict(config.inventory_dashboard)
    overrides = dict(base.pop("markets", {}))
    result = {str(base.get("market", "US")).upper(): base}
    for market, values in overrides.items():
        result[str(market).upper()] = {**base, **dict(values)}
    return result


def _apply_shipment_events_to_lots(lots: list[dict], events: list[dict]) -> int:
    by_sku: dict[str, list[dict]] = defaultdict(list)
    for lot in lots:
        by_sku[str(lot.get("sku", ""))].append(lot)
    for items in by_sku.values():
        items.sort(key=lambda item: (str(item.get("poDate", "")), str(item.get("poNumber", ""))))
    unmatched = 0
    for event in sorted(events, key=lambda item: (item["shipmentDate"], item["batch"], item["sku"])):
        remaining = int(event["quantity"])
        for lot in by_sku.get(str(event["sku"]), []):
            if str(lot.get("poDate", "")) > str(event["shipmentDate"]):
                break
            available = int(lot.get("availableQuantity", 0) or 0)
            if available <= 0:
                continue
            consumed = min(remaining, available)
            lot["previouslyShippedQuantity"] = int(lot.get("previouslyShippedQuantity", 0) or 0) + consumed
            lot["availableQuantity"] = available - consumed
            remaining -= consumed
            if remaining == 0:
                break
        unmatched += remaining
    return unmatched


def _merge_local_shipments(config: ProjectConfig, document: dict, local_root: Path, db: StateDb, run_id: int) -> dict:
    local_config = replace(
        config,
        data_root=local_root,
        inventory_dashboard={
            **config.inventory_dashboard,
            "document_master_sources": {"shipment_root": "."},
        },
    )
    normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
    catalog, _, events, sources, _ = _read_shipments(local_config, normalizer)
    usable_sources = [
        source for source in sources
        if "measureman-ops-mac-extracted/" not in source
        and "automation/runtime/output/" not in source
        and not Path(source).name.startswith(("~$", ".~"))
    ]
    source_by_key = {(_market(Path(source)), _batch(Path(source))): source for source in usable_sources if _batch(Path(source)) is not None}
    existing_keys = {
        (_market(Path(source)), _batch(Path(source)))
        for source in document.get("sources", {}).get("shipments", [])
        if _batch(Path(source)) is not None
    }
    latest_existing_batch: dict[str, int] = defaultdict(int)
    for market, batch in existing_keys:
        latest_existing_batch[market] = max(latest_existing_batch[market], int(batch))
    register = {(str(item.get("market", "")), int(item.get("batch", 0) or 0)): item for item in document.get("shipmentRegister", [])}
    events_by_key: dict[tuple[str, int], list[dict]] = defaultdict(list)
    for event in events:
        key = (str(event["market"]), int(event["batch"]))
        if key in source_by_key:
            events_by_key[key].append(event)

    applied = []
    rejected = []
    lots = document.get("purchaseOrderLots", [])
    logistics = document.setdefault("logistics", {"US": {}, "CA": {}})
    new_keys = {
        key for key in set(source_by_key) - existing_keys
        if int(key[1]) > latest_existing_batch[str(key[0])]
    }
    for key in sorted(new_keys, key=lambda item: (item[1], item[0])):
        market, batch = key
        source = source_by_key[key]
        batch_events = events_by_key.get(key, [])
        detail_quantity = sum(int(item["quantity"]) for item in batch_events)
        register_row = register.get(key)
        register_quantity = int(register_row.get("productQuantity", 0) or 0) if register_row else None
        if register_quantity is None or detail_quantity != register_quantity:
            details = {
                "market": market,
                "batch": batch,
                "source": source,
                "detailQuantity": detail_quantity,
                "registerQuantity": register_quantity,
            }
            rejected.append(details)
            document.setdefault("exceptions", []).append({
                "category": "shipment_register_quantity_mismatch",
                "severity": "high",
                **details,
            })
            db.add_exception(
                run_id,
                category="shipment_register_quantity_mismatch",
                severity="high",
                source="local-source-refresh",
                raw=f"{market}-CM{batch}",
                base=None,
                cell=None,
                details=details,
            )
            continue

        source_label = f"local://Downloads/{source}"
        for sku, item in catalog.get(market, {}).items():
            if int(item.get("sourceBatchNumber", 0) or 0) != batch:
                continue
            current = logistics.setdefault(market, {}).get(sku)
            if current is None or batch >= int(current.get("sourceBatchNumber", 0) or 0):
                logistics[market][sku] = {**item, "sourcePath": source_label}
        unmatched = _apply_shipment_events_to_lots(lots, batch_events)
        if unmatched:
            unmatched_details = {
                "market": market,
                "batch": batch,
                "source": source_label,
                "unmatchedPurchaseOrderQuantity": unmatched,
            }
            document.setdefault("exceptions", []).append({
                "category": "shipment_quantity_without_open_po",
                "severity": "medium",
                **unmatched_details,
            })
            db.add_exception(
                run_id,
                category="shipment_quantity_without_open_po",
                severity="medium",
                source="local-source-refresh",
                raw=f"{market}-CM{batch}",
                base=None,
                cell=None,
                details=unmatched_details,
            )
        document.setdefault("sources", {}).setdefault("shipments", []).append(source_label)
        applied.append({
            "market": market,
            "batch": batch,
            "source": source_label,
            "skuCount": len(batch_events),
            "quantity": detail_quantity,
            "unmatchedPurchaseOrderQuantity": unmatched,
        })

    document["purchaseOrderLots"] = lots
    document.setdefault("coverage", {})["logistics"] = {market: len(items) for market, items in logistics.items()}
    document["coverage"]["purchaseOrderLotsAvailable"] = sum(1 for item in lots if int(item.get("availableQuantity", 0) or 0) > 0)
    document["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    document["localRefresh"] = {"appliedShipments": applied, "rejectedShipments": rejected}
    return {"applied": applied, "rejected": rejected}


def _refresh_dashboard(
    config: ProjectConfig,
    payload: dict,
    settings: dict,
    inventory_path: Path,
    inventory_date: date,
    purchase_lots: list[dict],
    latest_archive_source: str,
) -> dict:
    normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
    market = str(payload.get("market", settings.get("market", "US"))).upper()
    fba_sheet = settings.get("fba_master_sheet", settings.get("master_sheet", "库存规划"))
    fba_header = settings.get("fba_master_header", "FBA库存")
    domestic_sheet = settings.get("domestic_master_sheet", settings.get("master_sheet", "库存规划"))
    sales_sheet = settings.get("sales_master_sheet", fba_sheet)
    fba, invalid_fba = _read_fba_from_master(inventory_path, fba_sheet, fba_header, normalizer)
    master = _read_master(inventory_path, domestic_sheet, normalizer, settings.get("local_inventory_header", "工厂库存及已下订单"))
    sales, sales_months, _ = _read_sales_from_master(
        inventory_path,
        sales_sheet,
        str(settings["sales_baseline_month"]),
        settings.get("sales_recent_month_header", "最近月销售"),
        normalizer,
    )
    rows_by_sku = {str(row["sku"]): dict(row) for row in payload.get("rows", [])}
    prior_core = {
        sku: (row.get("fbaSellable"), row.get("sourceDomesticSupplyTotal"), row.get("dailySales"), row.get("cartonQty"))
        for sku, row in rows_by_sku.items()
    }
    lots_by_sku: dict[str, list[dict]] = defaultdict(list)
    for lot in purchase_lots:
        lots_by_sku[str(lot.get("sku", ""))].append(lot)
    params = payload.get("parameters", {})
    replenishment = ReplenishmentParameters(
        lead_time_days=int(params.get("leadTimeDays", 75)),
        review_cycle_days=int(params.get("reviewCycleDays", 7)),
        target_cover_days=int(params.get("targetCoverDays", 45)),
        safety_stock_days=int(params.get("safetyStockDays", 21)),
        excess_cover_days=int(params.get("excessCoverDays", 240)),
        fba_transfer_trigger_days=int(params.get("fbaTransferTriggerDays", 30)),
    )
    overdue_days = int(params.get("purchaseOrderOverdueDays", 45))
    all_skus = sorted(set(rows_by_sku) | set(fba) | set(sales) | {sku for sku, item in master.items() if int(item.get("localInventory", 0) or 0) > 0})
    rows = []
    for sku in all_skus:
        row = rows_by_sku.get(sku, {"sku": sku})
        fba_item = fba.get(sku, {})
        master_item = master.get(sku, {})
        sales_item = sales.get(sku, {"dailySales": 0.0, "salesByMonth": []})
        row.update({
            "productName": master_item.get("productName") or row.get("productName") or sku,
            "factory": master_item.get("factory") or row.get("factory") or "",
            "cartonQty": master_item.get("cartonQty") or row.get("cartonQty"),
            "sourceDomesticSupplyTotal": int(master_item.get("localInventory", 0) or 0),
            "fbaSellable": int(fba_item.get("fbaSellable", 0) or 0),
            "fbaReservedTransfer": int(fba_item.get("fbaReservedTransfer", 0) or 0),
            "fbaReservedProcessing": int(fba_item.get("fbaReservedProcessing", 0) or 0),
            "awdOnHand": int(row.get("awdOnHand", 0) or 0),
            "awdAvailable": int(row.get("awdAvailable", 0) or 0),
            "awdPendingShipment": int(row.get("awdPendingShipment", 0) or 0),
            "awdOutboundToFba": int(row.get("awdOutboundToFba", 0) or 0),
            "awdInbound": int(row.get("awdInbound", 0) or 0),
            "awdFbaReconciliation": int(row.get("awdFbaReconciliation", 0) or 0),
            "dailySales": float(sales_item.get("dailySales", 0.0) or 0.0),
            "salesByMonth": sales_item.get("salesByMonth", []),
        })
        history = {str(item.get("month")): float(item.get("units", 0) or 0) for item in row.get("salesHistoryByMonth", [])}
        for item in row["salesByMonth"]:
            history[str(item["month"])] = float(item.get("units", 0) or 0)
        row["salesHistoryByMonth"] = [{"month": month, "units": history[month]} for month in sorted(history)]
        domestic = _reconcile_domestic_supply(
            row["sourceDomesticSupplyTotal"],
            lots_by_sku.get(sku, []),
            overdue_days,
            latest_archive_source=latest_archive_source,
        )
        row.update(domestic)
        row["domesticSupplyTotal"] = int(row["localInventory"]) + int(row["pendingOrderQty"])
        row.update(calculate_replenishment(
            daily_sales=row["dailySales"],
            fba_sellable=row["fbaSellable"],
            awd_available=int(row.get("awdAvailable", 0) or 0),
            awd_outbound_to_fba=int(row.get("awdOutboundToFba", 0) or 0),
            carton_quantity=row.get("cartonQty"),
            parameters=replenishment,
        ))
        quality = []
        if row["dailySales"] <= 0:
            quality.append("missing_sales_baseline")
        if not row.get("cartonQty"):
            quality.append("missing_carton_quantity")
        if sku not in master:
            quality.append("missing_product_master")
        row["dataQuality"] = quality
        row["readyToShipQty"] = min(int(row["localInventory"]), int(row["suggestedShipmentQty"]))
        row["suggestedProductionQty"] = max(0, int(row["suggestedShipmentQty"]) - int(row["domesticSupplyTotal"]))
        if any((row["fbaSellable"], row.get("awdAvailable", 0), row.get("awdOutboundToFba", 0), row.get("awdInbound", 0), row["domesticSupplyTotal"], row["dailySales"])):
            rows.append(row)

    rows.sort(key=lambda row: ({"critical": 0, "watch": 1, "data": 2, "healthy": 3, "excess": 4}.get(row["riskLevel"], 5), -int(row["suggestedShipmentQty"]), row["sku"]))
    payload["rows"] = rows
    payload["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    awd_date = date.fromisoformat(payload.get("snapshots", {}).get("awdDate", inventory_date.isoformat()))
    if not payload.get("snapshots", {}).get("awdSourceAvailable", False):
        awd_date = inventory_date
    snapshot_date = max(inventory_date, awd_date)
    tolerance = int(settings.get("snapshot_alignment_tolerance_days", 2))
    stale_after = int(settings.get("stale_after_days", 14))
    gap = abs((inventory_date - awd_date).days)
    payload["snapshots"] = {
        **payload.get("snapshots", {}),
        "fbaDate": inventory_date.isoformat(),
        "awdDate": awd_date.isoformat(),
        "alignmentGapDays": gap,
        "aligned": gap <= tolerance,
        "ageDays": (date.today() - snapshot_date).days,
        "staleAfterDays": stale_after,
        "isStale": (date.today() - snapshot_date).days > stale_after,
    }
    sales_year, sales_month = (int(value) for value in sales_months[-1].split("-"))
    if sales_month == 12:
        next_month = date(sales_year + 1, 1, 1)
    else:
        next_month = date(sales_year, sales_month + 1, 1)
    sales_end = next_month.fromordinal(next_month.toordinal() - 1)
    payload.setdefault("sales", {})["windowMonths"] = sales_months
    payload["sales"]["ageDaysAtSnapshot"] = (snapshot_date - sales_end).days
    payload["sales"]["historyMonths"] = sorted({item["month"] for row in rows for item in row.get("salesHistoryByMonth", [])})

    new_meta = _source_meta(config, inventory_path, "fba_inventory")
    retained = [
        item for item in payload.get("sources", [])
        if item.get("kind") not in {"fba_inventory", "product_master"}
        and not (item.get("kind") == "sales_month" and "库存规划" in str(item.get("path", "")))
    ]
    payload["sources"] = [
        new_meta,
        {**new_meta, "kind": "sales_month"},
        {**new_meta, "kind": "product_master"},
        *retained,
    ]
    payload["summary"] = {
        "skuCount": len(rows),
        "fbaSellable": sum(int(row.get("fbaSellable", 0) or 0) for row in rows),
        "awdAvailable": sum(int(row.get("awdAvailable", 0) or 0) for row in rows),
        "awdOutboundToFba": sum(int(row.get("awdOutboundToFba", 0) or 0) for row in rows),
        "awdInboundNotCounted": sum(int(row.get("awdInbound", 0) or 0) for row in rows),
        "localInventory": sum(int(row.get("localInventory", 0) or 0) for row in rows),
        "pendingOrderQty": sum(int(row.get("pendingOrderQty", 0) or 0) for row in rows),
        "overdueOrderCount": sum(1 for row in rows for order in row.get("pendingOrders", []) if order.get("overdue")),
        "overduePurchaseOrderCount": len({order.get("poNumber") for row in rows for order in row.get("pendingOrders", []) if order.get("overdue")}),
        "overdueOrderSkuCount": sum(1 for row in rows if any(order.get("overdue") for order in row.get("pendingOrders", []))),
        "readyToShipQty": sum(int(row.get("readyToShipQty", 0) or 0) for row in rows),
        "suggestedProductionQty": sum(int(row.get("suggestedProductionQty", 0) or 0) for row in rows),
        "criticalSkuCount": sum(1 for row in rows if row.get("riskLevel") == "critical"),
        "reviewSkuCount": sum(1 for row in rows if row.get("action") == "REVIEW_DATA"),
        "suggestedShipmentQty": sum(int(row.get("suggestedShipmentQty", 0) or 0) for row in rows),
    }
    payload["dataQuality"] = {
        "invalidFbaSkuRows": len(invalid_fba),
        "invalidAwdSkuRows": int(payload.get("dataQuality", {}).get("invalidAwdSkuRows", 0) or 0),
        "missingSalesSkuCount": sum(1 for row in rows if "missing_sales_baseline" in row.get("dataQuality", [])),
        "missingCartonSkuCount": sum(1 for row in rows if "missing_carton_quantity" in row.get("dataQuality", [])),
    }

    advertising = payload.get("advertising", {})
    ad_params = advertising.get("parameters", {})
    ad_model = AdvertisingParameters(
        target_acos_percent=float(ad_params.get("targetAcosPercent", 30)),
        minimum_evidence_spend=float(ad_params.get("minimumEvidenceSpend", 20)),
        no_order_spend=float(ad_params.get("noOrderSpend", 20)),
        winner_min_orders=int(ad_params.get("winnerMinOrders", 5)),
        scale_min_orders=int(ad_params.get("scaleMinOrders", 1)),
        low_volume_max_clicks=int(ad_params.get("lowVolumeMaxClicks", 30)),
        budget_utilization_threshold_percent=float(ad_params.get("budgetUtilizationThresholdPercent", 80)),
        scale_max_acos_ratio=float(ad_params.get("scaleMaxAcosRatio", 0.9)),
    )
    row_index = {row["sku"]: row for row in rows}
    for campaign in advertising.get("campaigns", []):
        inventory_row = row_index.get(campaign.get("sku"))
        risk = inventory_row.get("riskLevel") if inventory_row else None
        campaign["inventoryRisk"] = risk
        campaign["inventoryDaysCover"] = inventory_row.get("daysCoverNetwork") if inventory_row else None
        campaign.update(recommend_campaign(
            spend=float(campaign.get("spend", 0) or 0),
            advertising_sales=float(campaign.get("advertisingSales", 0) or 0),
            orders=int(campaign.get("orders", 0) or 0),
            clicks=int(campaign.get("clicks", 0) or 0),
            impressions=int(campaign.get("impressions", 0) or 0),
            budget=float(campaign.get("budget", 0) or 0),
            period_days=int(campaign.get("periodDays", 30) or 30),
            inventory_risk=risk,
            parameters=ad_model,
        ))
    new_core = {
        row["sku"]: (row.get("fbaSellable"), row.get("sourceDomesticSupplyTotal"), row.get("dailySales"), row.get("cartonQty"))
        for row in rows
    }
    changed = sorted(sku for sku in set(prior_core) | set(new_core) if prior_core.get(sku) != new_core.get(sku))
    payload["localRefresh"] = {"inventoryWorkbook": payload["sources"][0], "changedSkuCount": len(changed), "changedSkus": changed}
    return {"market": market, "changedSkuCount": len(changed), "changedSkus": changed}


def run(config: ProjectConfig, db: StateDb, local_root: Path) -> dict:
    run_id = db.start_run("refresh-local-sources")
    try:
        local_root = local_root.resolve()
        inventory_source, inventory_date = _latest_inventory_workbook(local_root)
        archive_dir = config.runtime_root / "archive" / "local-sources"
        archive_dir.mkdir(parents=True, exist_ok=True)
        digest = _sha256(inventory_source)
        archived_inventory = archive_dir / f"inventory-planning-{inventory_date:%Y%m%d}-{digest[:12]}.xlsx"
        if not archived_inventory.exists():
            shutil.copy2(inventory_source, archived_inventory)

        reports = config.runtime_root / "reports"
        backup_dir = config.runtime_root / "archive" / "pre-local-refresh" / datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        backup_dir.mkdir(parents=True, exist_ok=True)
        for filename in ("document_master.json", "inventory_dashboard.json", "inventory_dashboard.ca.json", "purchase_plan.json"):
            source = reports / filename
            if source.exists():
                shutil.copy2(source, backup_dir / filename)
        document_path = reports / "document_master.json"
        document = json.loads(document_path.read_text(encoding="utf-8"))
        shipment_result = _merge_local_shipments(config, document, local_root, db, run_id)
        planning_totals = _planning_shipment_totals(archived_inventory)
        shipment_planning_mismatches = []
        for item in [*shipment_result["applied"], *shipment_result["rejected"]]:
            key = (str(item["market"]), int(item["batch"]))
            detail_quantity = int(item.get("quantity", item.get("detailQuantity", 0)) or 0)
            planning_quantity = planning_totals.get(key)
            if planning_quantity is None or planning_quantity == detail_quantity:
                continue
            details = {
                "market": key[0],
                "batch": key[1],
                "detailQuantity": detail_quantity,
                "planningQuantity": planning_quantity,
                "inventoryWorkbook": str(inventory_source),
            }
            shipment_planning_mismatches.append(details)
            document.setdefault("exceptions", []).append({
                "category": "planning_shipment_quantity_mismatch",
                "severity": "high",
                **details,
            })
            db.add_exception(
                run_id,
                category="planning_shipment_quantity_mismatch",
                severity="high",
                source="local-source-refresh",
                raw=f"{key[0]}-CM{key[1]}",
                base=None,
                cell=None,
                details=details,
            )
        document.setdefault("localRefresh", {})["planningShipmentMismatches"] = shipment_planning_mismatches
        latest_archive_source = str(document.get("sources", {}).get("purchaseOrderArchive", ""))
        settings = _settings_by_market(config)
        dashboards = {}
        dashboard_payloads = {}
        for market, market_settings in settings.items():
            filename = "inventory_dashboard.json" if market == "US" else f"inventory_dashboard.{market.lower()}.json"
            path = reports / filename
            payload = json.loads(path.read_text(encoding="utf-8"))
            dashboards[market] = _refresh_dashboard(
                config,
                payload,
                market_settings,
                archived_inventory,
                inventory_date,
                document.get("purchaseOrderLots", []),
                latest_archive_source,
            )
            dashboard_payloads[path] = payload

        _atomic_json(document_path, document)
        for path, payload in dashboard_payloads.items():
            _atomic_json(path, payload)

        relative_inventory = str(archived_inventory.relative_to(config.data_root)).replace("\\", "/")
        purchase_config = replace(config, inventory_dashboard={**config.inventory_dashboard, "master_workbook": relative_inventory})
        purchase_summary = run_purchase_plan(purchase_config, db)
        summary = {
            "run_id": run_id,
            "local_root": str(local_root),
            "inventory_source": str(inventory_source),
            "inventory_date": inventory_date.isoformat(),
            "archived_inventory": str(archived_inventory),
            "backup_dir": str(backup_dir),
            "shipments": shipment_result,
            "shipment_planning_mismatches": shipment_planning_mismatches,
            "dashboards": dashboards,
            "purchase_plan": purchase_summary,
        }
        _atomic_json(reports / "local_sources_refresh.json", {
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            **summary,
        })
        db.commit()
        db.finish_run(run_id, "completed", summary=summary)
        return summary
    except Exception as exc:
        db.finish_run(run_id, "failed", error=repr(exc))
        raise
