from __future__ import annotations

import json
import hashlib
import re
import subprocess
import zipfile
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel

from ..config import ProjectConfig
from ..db import StateDb
from ..sku import SkuNormalizer


PO_RE = re.compile(r"\b([A-Z]{2}\d{6}-\d+)\b", re.IGNORECASE)
BATCH_RE = re.compile(r"\bCM\s*0*(\d+)\b", re.IGNORECASE)
REGISTER_BATCH_RE = re.compile(r"(?:CM|AM)[- ]?0*(\d+)", re.IGNORECASE)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float | None:
    if value in (None, "", "-") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _configured_path(config: ProjectConfig, key: str, default: str) -> Path:
    settings = config.inventory_dashboard.get("document_master_sources", {})
    return (config.data_root / str(settings.get(key, default))).resolve()


def _source_label(config: ProjectConfig, path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(config.data_root.resolve())).replace("\\", "/")
    except ValueError:
        downloads = (Path.home() / "Downloads").resolve()
        try:
            return f"local://Downloads/{resolved.relative_to(downloads)}"
        except ValueError:
            return str(resolved)


def _date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    text = _text(value)
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            continue
    return None


def _compact_header(value: object) -> str:
    return re.sub(r"\s+", "", _text(value))


def _base_sku(normalizer: SkuNormalizer, value: object) -> str | None:
    match = normalizer.extract(value)
    return match.bases[0] if match and len(match.bases) == 1 else None


def _market(path: Path) -> str:
    value = str(path).upper()
    return "CA" if "加拿大" in value or re.search(r"(?:^|[-_ ])CA(?=\d|[^A-Z0-9]|$)", value) else "US"


def _batch(path: Path) -> int | None:
    matches = BATCH_RE.findall(str(path))
    return int(matches[-1]) if matches else None


def _business_date(path: Path) -> date:
    values = [path.name, *[part for part in reversed(path.parts[-4:-1])]]
    patterns = (
        re.compile(r"(20\d{2})[._-]?(\d{1,2})[._-]?(\d{1,2})"),
        re.compile(r"(?<!\d)(\d{2})[._-](\d{1,2})[._-](\d{1,2})(?!\d)"),
    )
    for value in values:
        for pattern in patterns:
            for match in pattern.finditer(value):
                year = int(match.group(1))
                if year < 100:
                    year += 2000
                try:
                    return date(year, int(match.group(2)), int(match.group(3)))
                except ValueError:
                    continue
    return datetime.fromtimestamp(path.stat().st_mtime).date()


def _file_score(path: Path) -> tuple[int, int, float, int]:
    name = path.name
    clean = not (name.startswith("~$") or name.startswith(".~"))
    preferred = not any(token in name for token in ("副本", "范例", "空表"))
    return (int(clean), int(preferred), path.stat().st_mtime, path.stat().st_size)


def _choose_by_batch(paths: list[Path]) -> dict[tuple[str, int], Path]:
    grouped: dict[tuple[str, int], list[Path]] = defaultdict(list)
    for path in paths:
        batch = _batch(path)
        if batch is not None and not path.name.startswith(("~$", ".~")):
            grouped[(_market(path), batch)].append(path)
    return {key: max(items, key=_file_score) for key, items in grouped.items()}


def _shipment_roots(config: ProjectConfig) -> tuple[Path, list[Path]]:
    primary = _configured_path(config, "shipment_root", "MEASUREMAN/发货清单")
    incremental = _configured_path(config, "shipment_incremental_root", "")
    extras = [incremental] if incremental != config.data_root.resolve() and incremental.exists() else []
    return primary, extras


def _declaration_columns(sheet) -> dict[str, int]:
    header_row = next((row for row in range(1, min(sheet.max_row, 40) + 1) if _text(sheet.cell(row, 1).value) == "合同号"), 9)
    headers: dict[str, int] = {}
    for column in range(1, min(sheet.max_column, 32) + 1):
        label = _compact_header(sheet.cell(header_row, column).value).lower()
        if label:
            headers.setdefault(label, column)
    sku_column = next((headers[label] for label in ("sku", "产品编号", "采购产品编号", "货号") if label in headers), None)
    return {
        "header": header_row,
        "sku": sku_column or 13,
        "factory": headers.get("工厂", 2),
        "productName": headers.get("品名", 3),
        "quantity": headers.get("数量", 4),
        "purchaseAmount": headers.get("采购金额", 9),
        "unitPrice": headers.get("单价", 14),
        "taxRate": headers.get("退税率", 11),
    }


def _read_shipments(config: ProjectConfig, normalizer: SkuNormalizer):
    root, incremental_roots = _shipment_roots(config)
    candidates = [
        path for candidate_root in [root, *incremental_roots] for path in candidate_root.rglob("*.xlsx")
        if "发货清单" in path.name and "报运单" not in path.name and path.name != "发货清单.xlsx"
    ]
    chosen = _choose_by_batch(candidates)
    catalog: dict[str, dict[str, dict]] = {"US": {}, "CA": {}}
    batch_quantities: dict[tuple[str, int, str], int] = {}
    shipment_events = []
    sources = []
    for (market, batch), path in sorted(chosen.items(), key=lambda item: item[0]):
        try:
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=False)
        except Exception:
            continue
        try:
            if "Measureman" not in workbook.sheetnames:
                continue
            sheet = workbook["Measureman"]
            if "产品编号" not in _text(sheet["B2"].value).replace("　", ""):
                continue
            sources.append(_source_label(config, path))
            is_incremental = any(path.resolve().is_relative_to(candidate.resolve()) for candidate in incremental_roots)
            for row in range(3, min(sheet.max_row, 500) + 1):
                sku = _base_sku(normalizer, sheet.cell(row, 2).value)
                quantity = _number(sheet.cell(row, 11).value)
                carton = _number(sheet.cell(row, 4).value)
                if not sku or quantity is None or quantity <= 0 or carton is None or carton <= 0:
                    continue
                batch_quantities[(market, batch, sku)] = int(round(quantity))
                if not is_incremental:
                    shipment_events.append({
                        "market": market,
                        "batch": batch,
                        "shipmentDate": _business_date(path).isoformat(),
                        "sku": sku,
                        "quantity": int(round(quantity)),
                    })
                item = {
                    "sku": sku,
                    "cartonQty": int(round(carton)),
                    "netWeightKg": _number(sheet.cell(row, 5).value),
                    "grossWeightKg": _number(sheet.cell(row, 6).value),
                    "lengthCm": _number(sheet.cell(row, 7).value),
                    "widthCm": _number(sheet.cell(row, 8).value),
                    "heightCm": _number(sheet.cell(row, 9).value),
                    "cartonVolumeM3": _number(sheet.cell(row, 10).value),
                    "unitPriceRmb": _number(sheet.cell(row, 16).value),
                    "imageFormula": _text(sheet.cell(row, 3).value),
                    "sourceBatch": f"CM{batch}",
                    "sourcePath": _source_label(config, path),
                    "sourceBatchNumber": batch,
                }
                prior = catalog[market].get(sku)
                if prior is None or batch >= prior["sourceBatchNumber"]:
                    catalog[market][sku] = item
        finally:
            workbook.close()
    templates = {}
    primary_chosen = _choose_by_batch([
        path for path in root.rglob("*.xlsx")
        if "发货清单" in path.name and "报运单" not in path.name and path.name != "发货清单.xlsx"
    ])
    for market in ("US", "CA"):
        market_items = [(batch, path) for (item_market, batch), path in primary_chosen.items() if item_market == market]
        if market_items:
            _, path = max(market_items, key=lambda item: item[0])
            templates[market] = str(path.relative_to(config.data_root)).replace("\\", "/")
    return catalog, batch_quantities, shipment_events, sources, templates


def _read_declarations(config: ProjectConfig, normalizer: SkuNormalizer, shipment_quantities: dict):
    root, incremental_roots = _shipment_roots(config)
    candidates = [path for candidate_root in [root, *incremental_roots] for path in candidate_root.rglob("*.xlsx") if "报运单" in path.name]
    chosen = _choose_by_batch(candidates)
    lines: list[dict] = []
    profile_candidates: dict[tuple[str, str], tuple[int, list[dict]]] = {}
    sources = []
    for (market, batch), path in sorted(chosen.items(), key=lambda item: item[0]):
        try:
            workbook = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=False)
        except Exception:
            continue
        try:
            if "报关空表" not in workbook.sheetnames:
                continue
            sheet = workbook["报关空表"]
            columns = _declaration_columns(sheet)
            if _text(sheet.cell(columns["header"], 1).value) != "合同号":
                continue
            sources.append(_source_label(config, path))
            grouped: dict[str, list[dict]] = defaultdict(list)
            for row in range(columns["header"] + 1, min(sheet.max_row, 300) + 1):
                po_match = PO_RE.search(_text(sheet.cell(row, 1).value))
                raw_sku = _text(sheet.cell(row, columns["sku"]).value)
                sku = _base_sku(normalizer, raw_sku)
                quantity = _number(sheet.cell(row, columns["quantity"]).value)
                if not sku or quantity is None or quantity <= 0:
                    continue
                factory = _text(sheet.cell(row, columns["factory"]).value)
                product_name = _text(sheet.cell(row, columns["productName"]).value)
                amount = _number(sheet.cell(row, columns["purchaseAmount"]).value) or 0.0
                unit_price = _number(sheet.cell(row, columns["unitPrice"]).value) or (amount / quantity if quantity else 0.0)
                entry = {
                    "market": market,
                    "batch": batch,
                    "poNumber": po_match.group(1).upper() if po_match else "",
                    "sku": sku,
                    "declarationSku": raw_sku or sku,
                    "factory": factory,
                    "productName": product_name,
                    "quantity": int(round(quantity)),
                    "unitPrice": round(unit_price, 4),
                    "taxRate": _number(sheet.cell(row, columns["taxRate"]).value) or 0.13,
                    "sourcePath": _source_label(config, path),
                }
                lines.append(entry)
                grouped[sku].append(entry)
            for sku, sku_lines in grouped.items():
                shipped = shipment_quantities.get((market, batch, sku))
                if not shipped:
                    continue
                components = []
                component_groups: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
                for entry in sku_lines:
                    component_groups[(entry["declarationSku"], entry["factory"], entry["productName"])].append(entry)
                for (declaration_sku, factory, product_name), entries in component_groups.items():
                    component_quantity = sum(item["quantity"] for item in entries)
                    factor = component_quantity / shipped
                    if 0 < factor <= 10:
                        components.append({
                            "sku": sku,
                            "declarationSku": declaration_sku,
                            "factory": factory,
                            "productName": product_name,
                            "quantityFactor": round(factor, 6),
                            "taxRate": entries[-1]["taxRate"],
                        })
                if components:
                    key = (market, sku)
                    if key not in profile_candidates or batch >= profile_candidates[key][0]:
                        profile_candidates[key] = (batch, components)
        finally:
            workbook.close()
    profiles = {"US": {}, "CA": {}}
    for (market, sku), (batch, components) in profile_candidates.items():
        profiles[market][sku] = {"sourceBatch": f"CM{batch}", "components": components}
    templates = {}
    primary_chosen = _choose_by_batch([path for path in root.rglob("*.xlsx") if "报运单" in path.name])
    for market in ("US", "CA"):
        market_items = [(batch, path) for (item_market, batch), path in primary_chosen.items() if item_market == market]
        if market_items:
            _, path = max(market_items, key=lambda item: item[0])
            templates[market] = str(path.relative_to(config.data_root)).replace("\\", "/")
    return lines, profiles, sources, templates


def _po_date(po_number: str) -> date:
    digits = re.search(r"(\d{6})", po_number).group(1)
    year = 2000 + int(digits[:2])
    return date(year, int(digits[2:4]), int(digits[4:6]))


def _read_purchase_orders(
    config: ProjectConfig,
    normalizer: SkuNormalizer,
    archive_paths: dict[Path, str] | None = None,
):
    root = _configured_path(config, "purchase_order_root", "一店/采购订单")
    summary_path = _configured_path(config, "purchase_order_summary_workbook", "一店/采购订单/采购订单总表.xlsx")
    lots = []
    exceptions = []
    candidates = {
        path.resolve(): str(path.relative_to(config.data_root)).replace("\\", "/")
        for path in root.rglob("*.xlsx")
    }
    candidates.update({path.resolve(): label for path, label in (archive_paths or {}).items()})
    for path, source_label in sorted(candidates.items(), key=lambda item: str(item[0])):
        if path.name.startswith(("~$", ".~")) or path.resolve() == summary_path:
            continue
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
        except Exception as error:
            exceptions.append({"category": "purchase_order_unreadable", "path": str(path), "details": str(error)})
            continue
        try:
            for sheet in workbook.worksheets:
                if any(token in sheet.title for token in ("采购订单支出", "汇总", "总表")):
                    continue
                rows = [tuple(row) for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 300), max_col=min(sheet.max_column, 24), values_only=True)]
                current_po = ""
                current_factory = ""
                row_contexts = []
                for row in rows:
                    for index, value in enumerate(row):
                        match = PO_RE.search(_text(value))
                        if match:
                            current_po = match.group(1).upper()
                        if _text(value).upper() == "TO:" and index + 1 < len(row):
                            current_factory = _text(row[index + 1])
                    row_contexts.append((current_po, current_factory))
                sku_rows = []
                for index, row in enumerate(rows):
                    found = set()
                    for value in row[:7]:
                        match = normalizer.extract(value)
                        if match:
                            found.update(match.bases)
                    if found:
                        po_number, factory = row_contexts[index]
                        if po_number:
                            sku_rows.append((index, sorted(found), po_number, factory))
                for position, (row_index, skus, po_number, factory) in enumerate(sku_rows):
                    stop = sku_rows[position + 1][0] if position + 1 < len(sku_rows) else min(len(rows), row_index + 14)
                    quantity = unit_price = None
                    quantity_row = None
                    for candidate_index in range(row_index, min(stop, row_index + 14)):
                        if row_contexts[candidate_index][0] != po_number:
                            break
                        row = rows[candidate_index]
                        candidate_quantity = _number(row[7] if len(row) > 7 else None)
                        candidate_price = _number(row[8] if len(row) > 8 else None)
                        if candidate_quantity is not None and candidate_quantity > 0:
                            quantity, unit_price, quantity_row = candidate_quantity, candidate_price, candidate_index
                            break
                    if quantity is None:
                        continue
                    description = " ".join(_text(value) for value in rows[row_index][:7] if _text(value))
                    for sku in skus:
                        lots.append({
                            "poNumber": po_number,
                            "poDate": _po_date(po_number).isoformat(),
                            "sku": sku,
                            "factory": factory,
                            "orderedQuantity": int(round(quantity)),
                            "unitPrice": round(unit_price or 0.0, 4),
                            "productName": description,
                            "sourcePath": source_label,
                            "sourceSheet": sheet.title,
                            "sourceRow": (quantity_row or row_index) + 1,
                        })
        finally:
            workbook.close()
    deduped = {}
    for lot in lots:
        key = (lot["poNumber"], lot["sku"])
        prior = deduped.get(key)
        if prior is None or (lot["orderedQuantity"], bool(lot["unitPrice"])) > (prior["orderedQuantity"], bool(prior["unitPrice"])):
            deduped[key] = lot
    return list(deduped.values()), exceptions


def _read_latest_purchase_order_archive(config: ProjectConfig) -> tuple[dict[Path, str], list[dict], str | None]:
    root = _configured_path(config, "purchase_order_root", "一店/采购订单")
    candidates = [path for path in root.glob("*.zip") if "采购订单" in path.name]
    if not candidates:
        return {}, [], None
    source = max(candidates, key=lambda path: (_business_date(path), path.stat().st_mtime, path.stat().st_size))
    digest = hashlib.sha256(source.read_bytes()).hexdigest()[:12]
    cache_root = config.runtime_root / "source-cache" / "purchase-orders" / f"{source.stem}-{digest}"
    extracted = cache_root / "source"
    converted = cache_root / "xlsx"
    marker = cache_root / "complete.json"
    exceptions: list[dict] = []
    if not marker.exists():
        extracted.mkdir(parents=True, exist_ok=True)
        converted.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(source) as archive:
            for member in archive.infolist():
                target = (extracted / member.filename).resolve()
                if extracted.resolve() not in target.parents and target != extracted.resolve():
                    raise ValueError(f"采购订单压缩包包含不安全路径: {member.filename}")
                archive.extract(member, extracted)
        converter = config.project_root / "src" / "store_ops" / "convert_purchase_orders.ps1"
        result = subprocess.run(
            [
                "powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(converter),
                "-SourceDirectory", str(extracted), "-DestinationDirectory", str(converted),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
            check=False,
        )
        if result.returncode != 0:
            exceptions.append({"category": "purchase_order_archive_conversion", "path": str(source), "details": result.stderr.strip() or result.stdout.strip()})
            return {}, exceptions, str(source.relative_to(config.data_root)).replace("\\", "/")
        try:
            conversion = json.loads(result.stdout.strip() or "{}")
        except json.JSONDecodeError:
            conversion = {}
        for error in conversion.get("errors", []):
            exceptions.append({"category": "purchase_order_archive_file", "path": str(source), "details": json.dumps(error, ensure_ascii=False)})
        marker.write_text(json.dumps({"source": str(source), "sha256": digest}, ensure_ascii=False), encoding="utf-8")
    source_label = str(source.relative_to(config.data_root)).replace("\\", "/")
    archive_paths = {
        path.resolve(): f"{source_label}::{path.name}"
        for path in converted.glob("*.xlsx")
    }
    return archive_paths, exceptions, source_label


def _read_purchase_order_summary(config: ProjectConfig, normalizer: SkuNormalizer):
    """Read the consolidated purchase ledger and preserve its arrival status.

    The ledger is authoritative for whether a purchase line is still in
    production. Blank dates, PO numbers and factories are inherited from the
    preceding line, matching the visual grouping used in the workbook.
    """
    path = _configured_path(config, "purchase_order_summary_workbook", "一店/采购订单/采购订单总表.xlsx")
    if not path.exists():
        return [], [], None
    exceptions = []
    lots = []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as error:
        return [], [{"category": "purchase_order_summary_unreadable", "path": str(path), "details": str(error)}], path
    try:
        for sheet in workbook.worksheets:
            header_row = None
            header_map: dict[str, int] = {}
            for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5), values_only=True), 1):
                candidate = {_compact_header(value): index for index, value in enumerate(row) if _compact_header(value)}
                if {"订单号", "采购产品编号", "采购数量"}.issubset(candidate):
                    header_row = row_number
                    header_map = candidate
                    break
            if header_row is None:
                continue
            current_date: date | None = None
            current_po = ""
            current_factory = ""
            for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                date_index = header_map.get("时间", header_map.get("日期"))
                row_date = _date_value(row[date_index] if date_index is not None and date_index < len(row) else None)
                if row_date:
                    current_date = row_date
                po_value = row[header_map["订单号"]] if header_map["订单号"] < len(row) else None
                po_match = PO_RE.search(_text(po_value))
                if po_match:
                    current_po = po_match.group(1).upper()
                factory_index = header_map.get("供应商")
                factory_value = _text(row[factory_index] if factory_index is not None and factory_index < len(row) else None)
                if factory_value:
                    current_factory = factory_value
                sku_value = row[header_map["采购产品编号"]] if header_map["采购产品编号"] < len(row) else None
                sku = _base_sku(normalizer, sku_value)
                quantity = _number(row[header_map["采购数量"]] if header_map["采购数量"] < len(row) else None)
                if not sku or not current_po or quantity is None or quantity <= 0:
                    continue
                po_date = current_date
                if po_date is None:
                    try:
                        po_date = _po_date(current_po)
                    except (AttributeError, ValueError):
                        exceptions.append({
                            "category": "purchase_order_summary_missing_date",
                            "path": str(path),
                            "details": f"{sheet.title}!{row_number}: {current_po}",
                        })
                        continue
                arrival_index = header_map.get("到货时间")
                arrival_value = row[arrival_index] if arrival_index is not None and arrival_index < len(row) else None
                arrival_text = _text(arrival_value)
                arrival_date = _date_value(arrival_value)
                received = bool(arrival_text)
                price_index = header_map.get("采购单价")
                name_index = header_map.get("品名")
                lots.append({
                    "poNumber": current_po,
                    "poDate": po_date.isoformat(),
                    "sku": sku,
                    "factory": current_factory,
                    "orderedQuantity": int(round(quantity)),
                    "unitPrice": round(_number(row[price_index] if price_index is not None and price_index < len(row) else None) or 0.0, 4),
                    "productName": _text(row[name_index] if name_index is not None and name_index < len(row) else None),
                    "received": received,
                    "receivedAt": arrival_date.isoformat() if arrival_date else arrival_text,
                    "sourcePath": _source_label(config, path),
                    "sourceSheet": sheet.title,
                    "sourceRow": row_number,
                })
    finally:
        workbook.close()
    deduped = {}
    for lot in lots:
        deduped[(lot["poNumber"], lot["sku"])] = lot
    return list(deduped.values()), exceptions, path


def _read_payment_ledger(config: ProjectConfig):
    """Read payment metadata grouped visually in the Amazon payment ledger."""
    path = _configured_path(config, "payment_ledger_workbook", "一店/采购订单/亚马逊付款台账.xlsx")
    if not path.exists():
        return {}, [], None
    exceptions = []
    payments: dict[str, dict[str, set[str]]] = defaultdict(lambda: {
        "companyNames": set(),
        "paymentMethods": set(),
        "paymentPayers": set(),
        "paymentDates": set(),
        "notes": set(),
    })
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as error:
        return {}, [{"category": "payment_ledger_unreadable", "path": str(path), "details": str(error)}], path
    try:
        for sheet in workbook.worksheets:
            header_row = None
            columns: dict[str, int] = {}
            for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5), values_only=True), 1):
                candidate = {_compact_header(value): index for index, value in enumerate(row) if _compact_header(value)}
                if {"订单号", "付款方式"}.issubset(candidate):
                    header_row, columns = row_number, candidate
                    break
            if header_row is None:
                continue
            context = {"companyName": "", "paymentMethod": "", "paymentPayer": "", "paymentDate": "", "note": ""}
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                def value(header: str) -> object:
                    index = columns.get(header)
                    return row[index] if index is not None and index < len(row) else None

                company = _text(value("公司名称"))
                payment_method = _text(value("付款方式"))
                payment_payer = _text(value("付款方"))
                payment_date_value = value("付款时间")
                payment_date = _date_value(payment_date_value)
                payment_date_text = payment_date.isoformat() if payment_date else _text(payment_date_value)
                note = _text(value("备注"))
                if company or payment_method or payment_payer or payment_date_text:
                    context = {
                        "companyName": company,
                        "paymentMethod": payment_method,
                        "paymentPayer": payment_payer,
                        "paymentDate": payment_date_text,
                        "note": note,
                    }
                order_text = _text(value("订单号"))
                po_numbers = {match.group(1).upper() for match in PO_RE.finditer(order_text)}
                for po_number in po_numbers:
                    target = payments[po_number]
                    for source_key, target_key in (
                        ("companyName", "companyNames"),
                        ("paymentMethod", "paymentMethods"),
                        ("paymentPayer", "paymentPayers"),
                        ("paymentDate", "paymentDates"),
                        ("note", "notes"),
                    ):
                        if context[source_key]:
                            target[target_key].add(context[source_key])
    finally:
        workbook.close()
    return {
        po_number: {key: sorted(values) for key, values in fields.items()}
        for po_number, fields in payments.items()
    }, exceptions, path


def _read_shipment_register(config: ProjectConfig):
    path = _configured_path(config, "shipment_register_workbook", "一店/出货记录/MEASUREMAN出货记录.xlsx")
    if not path.exists():
        return [], None
    settings = config.inventory_dashboard.get("document_master_sources", {})
    ca_prefixes = tuple(str(value).upper() for value in settings.get("shipment_register_ca_company_prefixes", ["NBST"]))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    events = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, ())
            columns = {_compact_header(value): index for index, value in enumerate(header) if _compact_header(value)}
            if not {"发货时间", "货号", "产品数量"}.issubset(columns):
                continue
            for row_number, row in enumerate(rows, 2):
                shipment_date = _date_value(row[columns["发货时间"]] if columns["发货时间"] < len(row) else None)
                batch_text = _text(row[columns["货号"]] if columns["货号"] < len(row) else None)
                batch_match = REGISTER_BATCH_RE.search(batch_text)
                quantity = _number(row[columns["产品数量"]] if columns["产品数量"] < len(row) else None)
                if not shipment_date or not batch_match or quantity is None or quantity <= 0:
                    continue
                company_index = columns.get("公司编号")
                company = _text(row[company_index] if company_index is not None and company_index < len(row) else None)
                events.append({
                    "market": "CA" if company.upper().startswith(ca_prefixes) else "US",
                    "batch": int(batch_match.group(1)),
                    "shipmentDate": shipment_date.isoformat(),
                    "productQuantity": int(round(quantity)),
                    "cartonCount": int(round(_number(row[columns.get("件数", -1)]) or 0)) if columns.get("件数") is not None else 0,
                    "companyNumber": company,
                    "fbaShipmentNumber": _text(row[columns.get("FBA货件编号", -1)]) if columns.get("FBA货件编号") is not None else "",
                    "sourceSheet": sheet.title,
                    "sourceRow": row_number,
                })
    finally:
        workbook.close()
    return events, path


def _atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _load_master_seed(config: ProjectConfig) -> dict:
    """Keep packaged historical coverage when the local folder is only incremental."""
    candidates = [config.runtime_root / "reports" / "document_master.json"]
    candidates.extend((config.runtime_root / "archive" / "pre-local-refresh").glob("*/document_master.json"))
    best: dict = {}
    best_score = -1
    for path in sorted(candidates):
        if not path.exists():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError, json.JSONDecodeError):
            continue
        coverage = payload.get("coverage", {})
        score = (
            sum(int(value) for value in coverage.get("logistics", {}).values())
            + sum(int(value) for value in coverage.get("declarationProfiles", {}).values())
            + int(coverage.get("purchaseOrderLots", 0))
            + int(coverage.get("shipmentRegisterEvents", 0))
        )
        if score > best_score:
            best, best_score = payload, score
    return best


def _profile_batch(item: dict) -> int:
    match = BATCH_RE.search(str(item.get("sourceBatch", item.get("sourceBatchNumber", ""))))
    return int(match.group(1)) if match else 0


def _merge_catalog(current: dict, seed: dict) -> dict:
    merged = {market: dict(seed.get(market, {})) for market in ("US", "CA")}
    for market in ("US", "CA"):
        for sku, item in current.get(market, {}).items():
            prior = merged[market].get(sku)
            if prior is None or _profile_batch(item) >= _profile_batch(prior):
                merged[market][sku] = item
    return merged


def run(config: ProjectConfig, db: StateDb) -> dict:
    run_id = db.start_run("build-document-master")
    try:
        seed_master = _load_master_seed(config)
        normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
        logistics, shipment_quantities, shipment_events, shipment_sources, shipment_templates = _read_shipments(config, normalizer)
        declaration_lines, profiles, declaration_sources, declaration_templates = _read_declarations(config, normalizer, shipment_quantities)
        archive_paths, archive_exceptions, purchase_archive_source = _read_latest_purchase_order_archive(config)
        po_lots, exceptions = _read_purchase_orders(config, normalizer, archive_paths)
        exceptions.extend(archive_exceptions)
        summary_lots, summary_exceptions, purchase_summary_path = _read_purchase_order_summary(config, normalizer)
        exceptions.extend(summary_exceptions)
        merged_lots = {(lot["poNumber"], lot["sku"]): lot for lot in po_lots}
        for lot in summary_lots:
            merged_lots[(lot["poNumber"], lot["sku"])] = lot
        po_lots = list(merged_lots.values())
        payments_by_po, payment_exceptions, payment_ledger_path = _read_payment_ledger(config)
        exceptions.extend(payment_exceptions)
        shipment_register, shipment_register_path = _read_shipment_register(config)
        lots_by_sku: dict[str, list[dict]] = defaultdict(list)
        for lot in po_lots:
            lot["previouslyShippedQuantity"] = lot["orderedQuantity"] if lot.get("received") else 0
            lots_by_sku[lot["sku"]].append(lot)
        for items in lots_by_sku.values():
            items.sort(key=lambda item: (item["poDate"], item["poNumber"]))
        for event in sorted(shipment_events, key=lambda item: (item["shipmentDate"], item["batch"], item["market"])):
            remaining = event["quantity"]
            for lot in lots_by_sku.get(event["sku"], []):
                if lot["poDate"] > event["shipmentDate"]:
                    break
                available = lot["orderedQuantity"] - lot["previouslyShippedQuantity"]
                if available <= 0:
                    continue
                consumed = min(remaining, available)
                lot["previouslyShippedQuantity"] += consumed
                remaining -= consumed
                if remaining == 0:
                    break
        seed_lots = {
            (item.get("poNumber", ""), item.get("sku", "")): item
            for item in seed_master.get("purchaseOrderLots", [])
            if item.get("poNumber") and item.get("sku")
        }
        current_lots = {(item["poNumber"], item["sku"]): item for item in po_lots}
        for key, lot in current_lots.items():
            prior = seed_lots.get(key)
            if prior:
                lot["previouslyShippedQuantity"] = min(
                    lot["orderedQuantity"],
                    max(int(lot.get("previouslyShippedQuantity", 0)), int(prior.get("previouslyShippedQuantity", 0))),
                )
        for key, prior in seed_lots.items():
            if key not in current_lots:
                po_lots.append(dict(prior))
        for lot in po_lots:
            lot["availableQuantity"] = max(0, lot["orderedQuantity"] - lot["previouslyShippedQuantity"])
            payment = payments_by_po.get(lot["poNumber"], {})
            lot["paymentMethods"] = payment.get("paymentMethods", [])
            lot["paymentPayers"] = payment.get("paymentPayers", [])
            lot["paymentDates"] = payment.get("paymentDates", [])
            lot["paymentNotes"] = payment.get("notes", [])

        logistics = _merge_catalog(logistics, seed_master.get("logistics", {}))
        profiles = _merge_catalog(profiles, seed_master.get("declarationProfiles", {}))
        seed_sources = seed_master.get("sources", {})
        shipment_sources = sorted(set(seed_sources.get("shipments", [])) | set(shipment_sources))
        declaration_sources = sorted(set(seed_sources.get("declarations", [])) | set(declaration_sources))
        seed_register = seed_master.get("shipmentRegister", [])
        register_by_key = {
            (item.get("market"), item.get("batch"), item.get("shipmentDate"), item.get("companyNumber")): item
            for item in seed_register
        }
        for item in shipment_register:
            register_by_key[(item.get("market"), item.get("batch"), item.get("shipmentDate"), item.get("companyNumber"))] = item
        shipment_register = list(register_by_key.values())

        payload = {
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "logistics": logistics,
            "declarationProfiles": profiles,
            "purchaseOrderLots": sorted(po_lots, key=lambda item: (item["poDate"], item["poNumber"], item["sku"])),
            "shipmentRegister": sorted(shipment_register, key=lambda item: (item["shipmentDate"], item["batch"])),
            "coverage": {
                "logistics": {market: len(items) for market, items in logistics.items()},
                "declarationProfiles": {market: len(items) for market, items in profiles.items()},
                "purchaseOrderLots": len(po_lots),
                "purchaseOrderLotsAvailable": sum(1 for item in po_lots if item["availableQuantity"] > 0),
                "shipmentRegisterEvents": len(shipment_register),
            },
            "sources": {
                "shipments": shipment_sources,
                "declarations": declaration_sources,
                "purchaseOrderSummary": _source_label(config, purchase_summary_path) if purchase_summary_path else None,
                "purchaseOrderArchive": purchase_archive_source or seed_sources.get("purchaseOrderArchive"),
                "paymentLedger": _source_label(config, payment_ledger_path) if payment_ledger_path else None,
                "shipmentRegister": _source_label(config, shipment_register_path) if shipment_register_path else seed_sources.get("shipmentRegister"),
            },
            "templates": {
                "shipment": shipment_templates,
                "declaration": declaration_templates,
            },
            "exceptions": exceptions,
        }
        output = config.runtime_root / "reports" / "document_master.json"
        _atomic_json(output, payload)
        summary = {"status": "completed", "output": str(output), **payload["coverage"], "exceptions": len(exceptions)}
        db.finish_run(run_id, "completed", summary)
        return summary
    except Exception as error:
        db.finish_run(run_id, "failed", error=str(error))
        raise
