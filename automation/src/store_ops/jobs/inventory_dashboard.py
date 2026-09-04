from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

import openpyxl

from ..config import ProjectConfig
from ..db import StateDb
from ..advertising import AdvertisingParameters, recommend_campaign
from ..replenishment import ReplenishmentParameters, calculate_replenishment
from ..sku import SkuNormalizer
from .document_master import run as run_document_master
from .product_catalog import run as run_product_catalog
from .content_workflow import run as run_content_workflow


def _atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _number(value: object) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _integer(value: object) -> int:
    return int(round(_number(value)))


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _base_sku(normalizer: SkuNormalizer, value: object) -> str | None:
    match = normalizer.extract(value)
    return match.bases[0] if match and len(match.bases) == 1 else None


def _file_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_meta(config: ProjectConfig, path: Path, kind: str) -> dict:
    return {
        "kind": kind,
        "path": str(path.relative_to(config.data_root)).replace("\\", "/"),
        "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds"),
        "sha256": _file_fingerprint(path),
    }


def _latest_inventory_file(folder: Path, suffix: str) -> tuple[Path, date]:
    pattern = re.compile(rf"^(\d{{8}}){re.escape(suffix)}$")
    candidates: list[tuple[date, Path]] = []
    for path in folder.glob("*.xlsx"):
        match = pattern.match(path.name)
        if match:
            candidates.append((datetime.strptime(match.group(1), "%Y%m%d").date(), path))
    if not candidates:
        raise FileNotFoundError(f"No inventory source matched *{suffix} in {folder}")
    snapshot_date, path = max(candidates, key=lambda item: item[0])
    return path, snapshot_date


def _snapshot_date_from_path(path: Path) -> date:
    match = re.search(r"(20\d{6})", path.name)
    if not match:
        raise ValueError(f"Inventory source has no snapshot date: {path}")
    return datetime.strptime(match.group(1), "%Y%m%d").date()


def _read_fba(path: Path, normalizer: SkuNormalizer, warehouse_filter: str) -> tuple[dict, list[str]]:
    aggregate: dict[str, dict] = defaultdict(lambda: {
        "fbaSellable": 0,
        "fbaReservedTransfer": 0,
        "fbaReservedProcessing": 0,
    })
    invalid: list[str] = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook["Amazon.FBA库存"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if warehouse_filter not in _text(row[0] if len(row) > 0 else None):
                continue
            raw_sku = row[1] if len(row) > 1 else None
            sku = _base_sku(normalizer, raw_sku)
            if not sku:
                if _text(raw_sku):
                    invalid.append(_text(raw_sku))
                continue
            item = aggregate[sku]
            item["fbaReservedTransfer"] += _integer(row[2] if len(row) > 2 else None)
            item["fbaReservedProcessing"] += _integer(row[3] if len(row) > 3 else None)
            item["fbaSellable"] += _integer(row[4] if len(row) > 4 else None)
    finally:
        workbook.close()
    return dict(aggregate), invalid


def _read_fba_from_master(
    path: Path,
    sheet: str,
    header_name: str,
    normalizer: SkuNormalizer,
    network_inventory_header: str | None = None,
) -> tuple[dict, list[str]]:
    """Read the current FBA balance embedded in the dated planning workbook."""
    aggregate: dict[str, dict] = {}
    invalid: list[str] = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        inventory_index = next((index for index, value in enumerate(header) if _text(value) == header_name), None)
        if inventory_index is None:
            raise ValueError(f"Planning workbook is missing FBA inventory column: {header_name}")
        network_inventory_index = next(
            (index for index, value in enumerate(header) if network_inventory_header and _text(value) == network_inventory_header),
            None,
        )
        for row in rows:
            raw_sku = row[0] if row else None
            sku = _base_sku(normalizer, raw_sku)
            if not sku:
                if _text(raw_sku):
                    invalid.append(_text(raw_sku))
                continue
            fba_sellable = max(0, _integer(row[inventory_index] if len(row) > inventory_index else None))
            network_inventory = max(
                fba_sellable,
                _integer(row[network_inventory_index] if network_inventory_index is not None and len(row) > network_inventory_index else None),
            )
            aggregate[sku] = {
                "fbaSellable": fba_sellable,
                "fbaReservedTransfer": 0,
                "fbaReservedProcessing": 0,
                "inTransitInventory": max(0, network_inventory - fba_sellable),
            }
    finally:
        workbook.close()
    return aggregate, invalid


def _read_awd(path: Path, normalizer: SkuNormalizer, warehouse_filter: str) -> tuple[dict, list[str]]:
    aggregate: dict[str, dict] = defaultdict(lambda: {
        "productName": "",
        "awdOnHand": 0,
        "awdAvailable": 0,
        "awdPendingShipment": 0,
        "awdOutboundToFba": 0,
        "awdInbound": 0,
        "awdFbaReconciliation": 0,
    })
    invalid: list[str] = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook["Amazon.AWD库存"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if warehouse_filter not in _text(row[0] if len(row) > 0 else None):
                continue
            raw_sku = row[2] if len(row) > 2 else None
            sku = _base_sku(normalizer, raw_sku)
            if not sku:
                if _text(raw_sku):
                    invalid.append(_text(raw_sku))
                continue
            item = aggregate[sku]
            item["productName"] = item["productName"] or _text(row[3] if len(row) > 3 else None)
            item["awdOnHand"] += _integer(row[5] if len(row) > 5 else None)
            item["awdAvailable"] += _integer(row[6] if len(row) > 6 else None)
            item["awdPendingShipment"] += _integer(row[7] if len(row) > 7 else None)
            item["awdOutboundToFba"] += _integer(row[8] if len(row) > 8 else None)
            item["awdInbound"] += _integer(row[9] if len(row) > 9 else None)
            item["awdFbaReconciliation"] += _integer(row[10] if len(row) > 10 else None)
    finally:
        workbook.close()
    return dict(aggregate), invalid


def _read_master(path: Path, sheet: str, normalizer: SkuNormalizer, local_inventory_header: str) -> dict:
    result: dict[str, dict] = {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        local_inventory_index = next((index for index, value in enumerate(header) if _text(value) == local_inventory_header), None)
        if local_inventory_index is None:
            raise ValueError(f"Product master is missing local inventory column: {local_inventory_header}")
        for row in rows:
            sku = _base_sku(normalizer, row[0] if row else None)
            if not sku:
                continue
            carton = _integer(row[4] if len(row) > 4 else None)
            result.setdefault(sku, {
                "productName": _text(row[1] if len(row) > 1 else None).replace("\n", " "),
                "cartonQty": carton if carton > 0 else None,
                "factory": _text(row[5] if len(row) > 5 else None),
                "localInventory": max(0, _integer(row[local_inventory_index] if local_inventory_index is not None and len(row) > local_inventory_index else None)),
            })
    finally:
        workbook.close()
    return result


def _month_days(year: int, month: int) -> int:
    if month == 12:
        return (date(year + 1, 1, 1) - date(year, month, 1)).days
    return (date(year, month + 1, 1) - date(year, month, 1)).days


def _read_sales_forecast(
    path: Path,
    sheet: str,
    year: int,
    window_months: int,
    normalizer: SkuNormalizer,
) -> tuple[dict, list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        month_columns: list[tuple[int, int]] = []
        for index, value in enumerate(header):
            try:
                month = int(value)
            except (TypeError, ValueError):
                continue
            if 1 <= month <= 12:
                month_columns.append((index, month))
        selected = month_columns[-window_months:]
        labels = [f"{year:04d}-{month:02d}" for _, month in selected]
        result: dict[str, dict] = {}
        for row in rows:
            sku = _base_sku(normalizer, row[0] if row else None)
            if not sku:
                continue
            month_values = []
            for index, month in selected:
                units = _number(row[index] if len(row) > index else None)
                month_values.append({"month": f"{year:04d}-{month:02d}", "units": round(units, 2)})
            daily_rates = [entry["units"] / _month_days(year, int(entry["month"][-2:])) for entry in month_values]
            weights = list(range(1, len(daily_rates) + 1))
            daily_sales = sum(rate * weight for rate, weight in zip(daily_rates, weights)) / sum(weights) if weights else 0.0
            if sku in result:
                prior = result[sku]
                prior["dailySales"] += daily_sales
                for index, entry in enumerate(month_values):
                    prior["salesByMonth"][index]["units"] += entry["units"]
            else:
                result[sku] = {"dailySales": daily_sales, "salesByMonth": month_values}
        for item in result.values():
            item["dailySales"] = round(item["dailySales"], 3)
            for entry in item["salesByMonth"]:
                entry["units"] = round(entry["units"], 2)
        return result, labels
    finally:
        workbook.close()


def _read_sales_from_master(
    path: Path,
    sheet: str,
    month: str,
    recent_month_header: str,
    normalizer: SkuNormalizer,
) -> tuple[dict, list[str], dict[str, dict[str, float]]]:
    """Read the latest completed-month SKU sales embedded in the planning workbook."""
    match = re.fullmatch(r"(20\d{2})-(0[1-9]|1[0-2])", month)
    if not match:
        raise ValueError(f"Invalid sales baseline month: {month}")
    year, month_number = (int(value) for value in match.groups())
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        sales_index = next((index for index, value in enumerate(header) if _text(value) == recent_month_header), None)
        if sales_index is None:
            raise ValueError(f"Planning workbook is missing recent sales column: {recent_month_header}")
        monthly_units: dict[str, float] = defaultdict(float)
        for row in rows:
            sku = _base_sku(normalizer, row[0] if row else None)
            if not sku:
                continue
            monthly_units[sku] += max(0.0, _number(row[sales_index] if len(row) > sales_index else None))
    finally:
        workbook.close()

    days_in_month = _month_days(year, month_number)
    sales = {
        sku: {
            "dailySales": round(units / days_in_month, 3),
            "salesByMonth": [{"month": month, "units": round(units, 2)}],
        }
        for sku, units in monthly_units.items()
    }
    history = {sku: {month: round(units, 2)} for sku, units in monthly_units.items()}
    return sales, [month], history


def _read_actual_sales(path: Path, sheet: str) -> dict[str, dict]:
    aggregate: dict[str, dict] = defaultdict(lambda: {
        "units": 0,
        "revenue": 0.0,
        "promotionRevenue": 0.0,
        "advertisingRevenue": 0.0,
    })
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_period = _text(row[0] if row else None)
            match = re.match(r"^(\d{4}-\d{2})", raw_period)
            if not match:
                continue
            month = match.group(1)
            item = aggregate[month]
            item["units"] += _integer(row[3] if len(row) > 3 else None)
            item["revenue"] += _number(row[4] if len(row) > 4 else None)
            item["promotionRevenue"] += _number(row[5] if len(row) > 5 else None)
            item["advertisingRevenue"] += _number(row[6] if len(row) > 6 else None)
    finally:
        workbook.close()
    return {
        month: {
            "units": values["units"],
            "revenue": round(values["revenue"], 2),
            "promotionRevenue": round(values["promotionRevenue"], 2),
            "advertisingRevenue": round(values["advertisingRevenue"], 2),
        }
        for month, values in aggregate.items()
    }


def _read_sku_sales_history(path: Path, sheet: str, normalizer: SkuNormalizer) -> dict[str, dict[str, float]]:
    history: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw_period = _text(row[0] if row else None)
            match = re.match(r"^(\d{4}-\d{2})", raw_period)
            sku = _base_sku(normalizer, row[1] if len(row) > 1 else None)
            if not match or not sku:
                continue
            history[sku][match.group(1)] += _number(row[3] if len(row) > 3 else None)
    finally:
        workbook.close()
    return {
        sku: {month: round(units, 2) for month, units in sorted(months.items())}
        for sku, months in history.items()
    }


def _read_monthly_forecast_totals(
    path: Path,
    sheet: str,
    year: int,
    normalizer: SkuNormalizer,
) -> dict[str, int]:
    totals: dict[str, float] = defaultdict(float)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        month_columns: list[tuple[int, int]] = []
        for index, value in enumerate(header):
            try:
                month = int(value)
            except (TypeError, ValueError):
                continue
            if 1 <= month <= 12:
                month_columns.append((index, month))
        for row in rows:
            if not _base_sku(normalizer, row[0] if row else None):
                continue
            for index, month in month_columns:
                totals[f"{year:04d}-{month:02d}"] += _number(row[index] if len(row) > index else None)
    finally:
        workbook.close()
    return {month: int(round(value)) for month, value in totals.items()}


def _month_from_report_name(path: Path) -> str | None:
    match = re.search(r"(20\d{2})[._-]?(0?[1-9]|1[0-2])(?:月|(?=\D|$))", path.name)
    if not match:
        return None
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def _read_monthly_sales_reports(
    folder: Path | list[Path] | tuple[Path, ...],
    market: str,
    window_months: int,
    normalizer: SkuNormalizer,
) -> tuple[dict, list[str], dict[str, dict], list[Path], dict[str, dict[str, float]]]:
    by_month: dict[str, tuple[float, dict[str, dict], Path]] = {}
    folders = [folder] if isinstance(folder, Path) else list(folder)
    paths = sorted({path for root in folders if root.exists() for path in root.rglob("*.xlsx")})
    for path in paths:
        upper_name = path.name.upper()
        if market == "CA":
            if "CA" not in upper_name and "加拿大" not in str(path):
                continue
        elif any(token in upper_name for token in ("-CA", "CA.", "MX")) or "加拿大" in str(path):
            continue
        month = _month_from_report_name(path)
        if not month:
            continue
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception:
            continue
        try:
            if "SKU销售汇总" not in workbook.sheetnames:
                continue
            worksheet = workbook["SKU销售汇总"]
            monthly_rows: dict[str, dict] = defaultdict(lambda: {"units": 0, "revenue": 0.0})
            for row in worksheet.iter_rows(min_row=4, values_only=True):
                sku = _base_sku(normalizer, row[0] if row else None)
                if not sku:
                    continue
                monthly_rows[sku]["units"] += _integer(row[1] if len(row) > 1 else None)
                monthly_rows[sku]["revenue"] += _number(row[2] if len(row) > 2 else None)
        finally:
            workbook.close()
        if not monthly_rows:
            continue
        marker = path.stat().st_mtime
        current = by_month.get(month)
        if current is None or marker > current[0]:
            by_month[month] = (marker, dict(monthly_rows), path)

    months = sorted(by_month)
    if not months:
        return {}, [], {}, [], {}
    selected_months = months[-window_months:]
    all_skus = sorted({sku for month in selected_months for sku in by_month[month][1]})
    sales: dict[str, dict] = {}
    for sku in all_skus:
        sales_by_month = [
            {"month": month, "units": float(by_month[month][1].get(sku, {}).get("units", 0))}
            for month in selected_months
        ]
        daily_rates = [
            item["units"] / _month_days(int(item["month"][:4]), int(item["month"][-2:]))
            for item in sales_by_month
        ]
        weights = list(range(1, len(daily_rates) + 1))
        daily_sales = sum(rate * weight for rate, weight in zip(daily_rates, weights)) / sum(weights)
        sales[sku] = {"dailySales": round(daily_sales, 3), "salesByMonth": sales_by_month}

    actual_sales: dict[str, dict] = {}
    for month in months:
        rows = by_month[month][1]
        actual_sales[month] = {
            "units": sum(int(item["units"]) for item in rows.values()),
            "revenue": round(sum(float(item["revenue"]) for item in rows.values()), 2),
            "promotionRevenue": 0.0,
            "advertisingRevenue": 0.0,
        }
    source_paths = [by_month[month][2] for month in months]
    history_by_sku: dict[str, dict[str, float]] = defaultdict(dict)
    for month in months:
        for sku, item in by_month[month][1].items():
            history_by_sku[sku][month] = float(item["units"])
    return sales, selected_months, actual_sales, source_paths, dict(history_by_sku)


def _header_index(header: tuple, label: str) -> int:
    for index, value in enumerate(header):
        if _text(value) == label:
            return index
    raise ValueError(f"Advertising report is missing column: {label}")


def _read_advertising_reports(folder: Path, normalizer: SkuNormalizer) -> tuple[list[dict], list[dict], list[Path]]:
    by_month: dict[str, tuple[float, dict, list[dict], Path]] = {}
    for path in sorted(folder.glob("*.xlsx")):
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_links=False)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if len(rows) < 2:
            continue
        header = rows[0]
        indexes = {
            "start": _header_index(header, "数据开始时间"),
            "campaign": _header_index(header, "广告活动"),
            "status": _header_index(header, "状态"),
            "budget": _header_index(header, "预算"),
            "impressions": _header_index(header, "曝光量"),
            "clicks": _header_index(header, "点击量"),
            "spend": _header_index(header, "花费"),
            "sales": _header_index(header, "广告总销售额"),
            "orders": _header_index(header, "广告总订单量"),
        }
        data_rows = [row for row in rows[1:] if _text(row[indexes["start"]] if len(row) > indexes["start"] else None)]
        if not data_rows:
            continue
        month = _text(data_rows[0][indexes["start"]])[:7]
        period_year, period_month = (int(value) for value in month.split("-"))
        period_days = _month_days(period_year, period_month)
        totals = {"month": month, "spend": 0.0, "advertisingSales": 0.0, "orders": 0, "clicks": 0, "impressions": 0}
        campaigns: dict[str, dict] = {}
        for row in data_rows:
            name = _text(row[indexes["campaign"]] if len(row) > indexes["campaign"] else None)
            if not name:
                continue
            spend = _number(row[indexes["spend"]] if len(row) > indexes["spend"] else None)
            sales = _number(row[indexes["sales"]] if len(row) > indexes["sales"] else None)
            orders = _integer(row[indexes["orders"]] if len(row) > indexes["orders"] else None)
            clicks = _integer(row[indexes["clicks"]] if len(row) > indexes["clicks"] else None)
            impressions = _integer(row[indexes["impressions"]] if len(row) > indexes["impressions"] else None)
            totals["spend"] += spend
            totals["advertisingSales"] += sales
            totals["orders"] += orders
            totals["clicks"] += clicks
            totals["impressions"] += impressions
            item = campaigns.setdefault(name, {
                "campaign": name,
                "sku": _base_sku(normalizer, name),
                "status": _text(row[indexes["status"]] if len(row) > indexes["status"] else None),
                "budget": _number(row[indexes["budget"]] if len(row) > indexes["budget"] else None),
                "spend": 0.0,
                "advertisingSales": 0.0,
                "orders": 0,
                "clicks": 0,
                "impressions": 0,
                "periodDays": period_days,
            })
            item["spend"] += spend
            item["advertisingSales"] += sales
            item["orders"] += orders
            item["clicks"] += clicks
            item["impressions"] += impressions
        totals["spend"] = round(totals["spend"], 2)
        totals["advertisingSales"] = round(totals["advertisingSales"], 2)
        totals["acos"] = round(totals["spend"] / totals["advertisingSales"] * 100, 2) if totals["advertisingSales"] > 0 else None
        totals["roas"] = round(totals["advertisingSales"] / totals["spend"], 2) if totals["spend"] > 0 else None
        campaign_rows = []
        for item in campaigns.values():
            item["spend"] = round(item["spend"], 2)
            item["advertisingSales"] = round(item["advertisingSales"], 2)
            item["budget"] = round(item["budget"], 2)
            campaign_rows.append(item)
        marker = path.stat().st_mtime
        current = by_month.get(month)
        if current is None or marker > current[0]:
            by_month[month] = (marker, totals, campaign_rows, path)

    monthly = [by_month[month][1] for month in sorted(by_month)]
    if not monthly:
        return [], [], []
    latest_month = monthly[-1]["month"]
    latest_campaigns = by_month[latest_month][2]
    source_paths = [by_month[month][3] for month in sorted(by_month)]
    return monthly, latest_campaigns, source_paths


def _sum(rows: Iterable[dict], key: str) -> int:
    return sum(int(row.get(key, 0) or 0) for row in rows)


def _reconcile_domestic_supply(
    combined_quantity: int,
    purchase_order_lots: list[dict],
    overdue_days: int,
    as_of: date | None = None,
    latest_archive_source: str = "",
) -> dict:
    """Split the combined domestic figure using newest open PO lots first.

    The source workbook only stores one combined value. Capping open purchase
    orders to that value prevents stale historical lots from inflating supply.
    """
    current_date = as_of or date.today()
    remaining_combined = max(0, int(combined_quantity))
    unreflected_latest = 0
    pending_orders: list[dict] = []
    eligible = [lot for lot in purchase_order_lots if int(lot.get("availableQuantity", 0) or 0) > 0]
    eligible.sort(key=lambda lot: (lot.get("poDate", ""), lot.get("poNumber", "")), reverse=True)
    for lot in eligible:
        source_path = str(lot.get("sourcePath", ""))
        is_latest_archive = bool(latest_archive_source and source_path.startswith(f"{latest_archive_source}::"))
        if remaining_combined <= 0 and not is_latest_archive:
            continue
        source_remaining = int(lot.get("availableQuantity", 0) or 0)
        allocated = min(source_remaining, remaining_combined)
        unreflected = max(0, source_remaining - allocated) if is_latest_archive else 0
        effective_remaining = allocated + unreflected
        if effective_remaining <= 0:
            continue
        po_date = date.fromisoformat(str(lot["poDate"]))
        age_days = max(0, (current_date - po_date).days)
        expected_delivery = po_date + timedelta(days=overdue_days)
        pending_orders.append({
            "poNumber": str(lot.get("poNumber", "")),
            "poDate": po_date.isoformat(),
            "factory": str(lot.get("factory", "")),
            "orderedQuantity": int(lot.get("orderedQuantity", 0) or 0),
            "remainingQuantity": effective_remaining,
            "sourceRemainingQuantity": source_remaining,
            "unreflectedLatestQuantity": unreflected,
            "expectedDeliveryDate": expected_delivery.isoformat(),
            "ageDays": age_days,
            "overdueDays": max(0, age_days - overdue_days),
            "overdue": age_days > overdue_days,
        })
        remaining_combined -= allocated
        unreflected_latest += unreflected
    return {
        "localInventory": remaining_combined,
        "pendingOrderQty": max(0, int(combined_quantity)) - remaining_combined + unreflected_latest,
        "unreflectedLatestOrderQty": unreflected_latest,
        "pendingOrders": pending_orders,
    }


def _run_market(config: ProjectConfig, db: StateDb, settings: dict, purchase_order_lots: list[dict], latest_archive_source: str = "") -> dict:
    market = str(settings.get("market", "US")).upper()
    run_id = db.start_run(f"build-inventory-dashboard-data-{market}")
    normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
    try:
        folder = (config.data_root / settings["inventory_folder"]).resolve()
        master_path = (config.data_root / settings["master_workbook"]).resolve()
        fba_mode = settings.get("fba_mode", "warehouse_export")
        if fba_mode == "planning_workbook":
            fba_path = master_path
            fba_date = _snapshot_date_from_path(fba_path)
        elif settings.get("fba_file"):
            fba_path = (folder / settings["fba_file"]).resolve()
            fba_date = _snapshot_date_from_path(fba_path)
        else:
            fba_path, fba_date = _latest_inventory_file(folder, settings.get("fba_suffix", "库存.xlsx"))
        awd_enabled = bool(settings.get("awd_enabled", True))
        if awd_enabled:
            awd_path, awd_date = _latest_inventory_file(folder, settings.get("awd_suffix", "AWD库存.xlsx"))
        else:
            awd_path, awd_date = None, fba_date
        sales_path = (config.data_root / settings["sales_workbook"]).resolve()
        advertising_folder = (config.data_root / settings["advertising_folder"]).resolve()
        runtime_monthly_folder = config.runtime_root / "incoming" / "monthly-sales-reports"
        monthly_sales_folder = (config.data_root / settings["sales_monthly_folder"]).resolve() if settings.get("sales_monthly_folder") else None
        if monthly_sales_folder is not None and not monthly_sales_folder.exists() and runtime_monthly_folder.exists():
            monthly_sales_folder = runtime_monthly_folder
        history_monthly_folder = (config.data_root / settings["sales_history_monthly_folder"]).resolve() if settings.get("sales_history_monthly_folder") else monthly_sales_folder
        if history_monthly_folder is not None and not history_monthly_folder.exists() and runtime_monthly_folder.exists():
            history_monthly_folder = runtime_monthly_folder
        additional_history_folders = [
            (config.data_root / str(value)).resolve()
            for value in settings.get("sales_history_additional_folders", [])
        ]
        for required in (sales_path, master_path, advertising_folder, monthly_sales_folder, history_monthly_folder, *additional_history_folders):
            if required is None:
                continue
            if not required.exists():
                raise FileNotFoundError(required)

        if fba_mode == "planning_workbook":
            fba, invalid_fba = _read_fba_from_master(
                fba_path,
                settings.get("fba_master_sheet", settings.get("master_sheet", "库存规划")),
                settings.get("fba_master_header", "FBA库存"),
                normalizer,
                settings.get("network_inventory_header", "FBA+在途库存"),
            )
        else:
            fba, invalid_fba = _read_fba(fba_path, normalizer, settings.get("fba_warehouse_filter", "US_FBA"))
        if awd_path:
            awd, invalid_awd = _read_awd(awd_path, normalizer, settings.get("awd_warehouse_filter", "US_AWD"))
        else:
            awd, invalid_awd = {}, []
        domestic_master_sheet = settings.get("domestic_master_sheet", settings.get("master_sheet", "库存规划"))
        master = _read_master(
            master_path,
            domestic_master_sheet,
            normalizer,
            settings.get("local_inventory_header", "工厂库存及已下订单"),
        )
        workbook_history = _read_sku_sales_history(
            sales_path,
            settings.get("sales_history_sheet", settings.get("actual_sales_sheet", "US-按月导出")),
            normalizer,
        )
        sales_mode = settings.get("sales_mode", "forecast")
        sales_source_kind = "forecast"
        sales_source_sheet = settings.get("sales_sheet", "US2025预估")
        sales_method = "最近三个月按 1:2:3 加权后换算日销"
        if sales_mode == "monthly_reports":
            sales, sales_months, actual_sales, sales_source_paths, report_history = _read_monthly_sales_reports(
                [monthly_sales_folder, *additional_history_folders],
                market,
                int(settings.get("sales_window_months", 3)),
                normalizer,
            )
            forecast_totals = {}
            sales_source_kind = "monthly_reports"
            sales_source_sheet = settings.get("sales_sheet", "SKU销售汇总（月报）")
            sales_method = "最近月度销售报表按月汇总并换算日销"
        elif sales_mode == "planning_workbook":
            sales_source_sheet = settings.get(
                "sales_master_sheet",
                settings.get("fba_master_sheet", settings.get("master_sheet", "库存规划")),
            )
            sales, sales_months, current_history = _read_sales_from_master(
                master_path,
                sales_source_sheet,
                str(settings["sales_baseline_month"]),
                settings.get("sales_recent_month_header", "最近月销售"),
                normalizer,
            )
            sales_source_paths = [master_path]
            actual_sales = _read_actual_sales(sales_path, settings.get("actual_sales_sheet", "US-按月导出"))
            forecast_totals = (
                _read_monthly_forecast_totals(
                    sales_path,
                    settings.get("sales_sheet", "US2025预估"),
                    int(settings.get("sales_year", 2025)),
                    normalizer,
                )
                if settings.get("forecast_totals_enabled", market == "US")
                else {}
            )
            report_history: dict[str, dict[str, float]] = {}
            if history_monthly_folder:
                _, _, history_actual_sales, history_source_paths, report_history = _read_monthly_sales_reports(
                    history_monthly_folder,
                    market,
                    int(settings.get("sales_window_months", 3)),
                    normalizer,
                )
                if history_actual_sales:
                    actual_sales = history_actual_sales
                sales_source_paths.extend(history_source_paths)
            for sku, months in current_history.items():
                report_history.setdefault(sku, {}).update(months)
            sales_source_kind = "planning_workbook"
            sales_method = "库存规划表最近完整月销量换算日销"
        else:
            sales, sales_months = _read_sales_forecast(
                sales_path,
                settings.get("sales_sheet", "US2025预估"),
                int(settings.get("sales_year", 2025)),
                int(settings.get("sales_window_months", 3)),
                normalizer,
            )
            actual_sales = _read_actual_sales(sales_path, settings.get("actual_sales_sheet", "US-按月导出"))
            forecast_totals = _read_monthly_forecast_totals(
                sales_path,
                settings.get("sales_sheet", "US2025预估"),
                int(settings.get("sales_year", 2025)),
                normalizer,
            )
            sales_source_paths = [sales_path]
            report_history = {}
            if history_monthly_folder:
                _, _, _, history_source_paths, report_history = _read_monthly_sales_reports(
                    history_monthly_folder,
                    market,
                    int(settings.get("sales_window_months", 3)),
                    normalizer,
                )
                sales_source_paths.extend(history_source_paths)
        sales_history: dict[str, dict[str, float]] = defaultdict(dict)
        for sku, months in workbook_history.items():
            sales_history[sku].update(months)
        for sku, months in report_history.items():
            sales_history[sku].update(months)
        history_months = sorted({month for months in sales_history.values() for month in months})
        advertising_monthly, advertising_campaigns, advertising_paths = _read_advertising_reports(advertising_folder, normalizer)

        defaults = settings.get("parameters", {})
        parameters = ReplenishmentParameters(
            lead_time_days=int(defaults.get("lead_time_days", 75)),
            review_cycle_days=int(defaults.get("review_cycle_days", 7)),
            target_cover_days=int(defaults.get("target_cover_days", 90)),
            safety_stock_days=int(defaults.get("safety_stock_days", 21)),
            excess_cover_days=int(defaults.get("excess_cover_days", 240)),
            fba_transfer_trigger_days=int(defaults.get("fba_transfer_trigger_days", 30)),
        )
        purchase_order_overdue_days = int(settings.get("purchase_order_overdue_days", 45))
        purchase_orders_by_sku: dict[str, list[dict]] = defaultdict(list)
        for lot in purchase_order_lots:
            purchase_orders_by_sku[str(lot.get("sku", ""))].append(lot)

        local_inventory_skus = {sku for sku, item in master.items() if int(item.get("localInventory", 0)) > 0}
        latest_order_skus = {
            str(lot.get("sku", "")) for lot in purchase_order_lots
            if latest_archive_source and str(lot.get("sourcePath", "")).startswith(f"{latest_archive_source}::")
        }
        all_skus = sorted(set(fba) | set(awd) | set(sales) | local_inventory_skus | latest_order_skus)
        rows: list[dict] = []
        for sku in all_skus:
            fba_item = fba.get(sku, {})
            awd_item = awd.get(sku, {})
            master_item = master.get(sku, {})
            sales_item = sales.get(sku, {"dailySales": 0.0, "salesByMonth": []})
            carton = master_item.get("cartonQty")
            daily_sales = float(sales_item.get("dailySales", 0.0))
            source_domestic_supply_total = int(master_item.get("localInventory", 0))
            domestic_supply = _reconcile_domestic_supply(
                source_domestic_supply_total,
                purchase_orders_by_sku.get(sku, []),
                purchase_order_overdue_days,
                latest_archive_source=latest_archive_source,
            )
            domestic_supply_total = int(domestic_supply["localInventory"]) + int(domestic_supply["pendingOrderQty"])
            decision = calculate_replenishment(
                daily_sales=daily_sales,
                fba_sellable=int(fba_item.get("fbaSellable", 0)),
                awd_available=int(awd_item.get("awdAvailable", 0)),
                awd_outbound_to_fba=int(awd_item.get("awdOutboundToFba", 0)),
                carton_quantity=carton,
                parameters=parameters,
                in_transit_inventory=int(fba_item.get("inTransitInventory", awd_item.get("awdInbound", 0))),
            )
            quality: list[str] = []
            if daily_sales <= 0:
                quality.append("missing_sales_baseline")
            if not carton:
                quality.append("missing_carton_quantity")
            if sku not in master:
                quality.append("missing_product_master")
            row = {
                "sku": sku,
                "productName": awd_item.get("productName") or master_item.get("productName") or sku,
                "factory": master_item.get("factory") or "",
                "cartonQty": carton,
                "sourceDomesticSupplyTotal": source_domestic_supply_total,
                "domesticSupplyTotal": domestic_supply_total,
                **domestic_supply,
                "fbaSellable": int(fba_item.get("fbaSellable", 0)),
                "fbaReservedTransfer": int(fba_item.get("fbaReservedTransfer", 0)),
                "fbaReservedProcessing": int(fba_item.get("fbaReservedProcessing", 0)),
                "awdOnHand": int(awd_item.get("awdOnHand", 0)),
                "awdAvailable": int(awd_item.get("awdAvailable", 0)),
                "awdPendingShipment": int(awd_item.get("awdPendingShipment", 0)),
                "awdOutboundToFba": int(awd_item.get("awdOutboundToFba", 0)),
                "awdInbound": int(awd_item.get("awdInbound", 0)),
                "awdFbaReconciliation": int(awd_item.get("awdFbaReconciliation", 0)),
                "dailySales": daily_sales,
                "salesByMonth": sales_item.get("salesByMonth", []),
                "salesHistoryByMonth": [
                    {"month": month, "units": float(sales_history.get(sku, {}).get(month, 0))}
                    for month in history_months
                ],
                "dataQuality": quality,
                **decision,
            }
            row["readyToShipQty"] = min(row["localInventory"], row["suggestedShipmentQty"])
            row["suggestedProductionQty"] = max(0, row["suggestedShipmentQty"] - row["domesticSupplyTotal"])
            if any((row["fbaSellable"], row["awdAvailable"], row["awdOutboundToFba"], row["awdInbound"], row["domesticSupplyTotal"], daily_sales)):
                rows.append(row)

        rows.sort(key=lambda row: (
            {"critical": 0, "watch": 1, "data": 2, "healthy": 3, "excess": 4}.get(row["riskLevel"], 5),
            -row["suggestedShipmentQty"],
            row["sku"],
        ))
        age_days = (date.today() - max(fba_date, awd_date)).days
        tolerance = int(settings.get("snapshot_alignment_tolerance_days", 2))
        alignment_gap = abs((fba_date - awd_date).days)
        stale_after = int(settings.get("stale_after_days", 14))
        sales_year_value, sales_month_value = (int(value) for value in sales_months[-1].split("-"))
        sales_end = date(sales_year_value, sales_month_value, _month_days(sales_year_value, sales_month_value))
        sales_age_days = (max(fba_date, awd_date) - sales_end).days

        performance_months = [f"{month:02d}" for month in range(1, 13)]
        actual_year = int(settings.get("actual_sales_year", 2024))
        forecast_year = int(settings.get("sales_year", 2025))
        performance_series = []
        for month in performance_months:
            actual = actual_sales.get(f"{actual_year:04d}-{month}", {})
            performance_series.append({
                "month": month,
                "actualUnits": int(actual.get("units", 0)),
                "forecastUnits": int(forecast_totals.get(f"{forecast_year:04d}-{month}", 0)),
                "actualRevenue": round(float(actual.get("revenue", 0.0)), 2),
                "promotionRevenue": round(float(actual.get("promotionRevenue", 0.0)), 2),
                "advertisingRevenue": round(float(actual.get("advertisingRevenue", 0.0)), 2),
            })
        annual_actual_units = sum(item["actualUnits"] for item in performance_series)
        annual_actual_revenue = round(sum(item["actualRevenue"] for item in performance_series), 2)
        annual_advertising_revenue = round(sum(item["advertisingRevenue"] for item in performance_series), 2)
        latest_actual = performance_series[-1]
        previous_actual = performance_series[-2]
        latest_unit_change = (
            round((latest_actual["actualUnits"] - previous_actual["actualUnits"]) / previous_actual["actualUnits"] * 100, 1)
            if previous_actual["actualUnits"] > 0 else None
        )

        advertising_defaults = settings.get("advertising_parameters", {})
        advertising_parameters = AdvertisingParameters(
            target_acos_percent=float(advertising_defaults.get("target_acos_percent", 30)),
            minimum_evidence_spend=float(advertising_defaults.get("minimum_evidence_spend", 20)),
            no_order_spend=float(advertising_defaults.get("no_order_spend", 20)),
            winner_min_orders=int(advertising_defaults.get("winner_min_orders", 5)),
            scale_min_orders=int(advertising_defaults.get("scale_min_orders", 1)),
            low_volume_max_clicks=int(advertising_defaults.get("low_volume_max_clicks", 30)),
            budget_utilization_threshold_percent=float(advertising_defaults.get("budget_utilization_threshold_percent", 80)),
            scale_max_acos_ratio=float(advertising_defaults.get("scale_max_acos_ratio", 0.9)),
        )
        row_by_sku = {row["sku"]: row for row in rows}
        for campaign in advertising_campaigns:
            inventory_row = row_by_sku.get(campaign["sku"]) if campaign.get("sku") else None
            inventory_risk = inventory_row.get("riskLevel") if inventory_row else None
            campaign["inventoryRisk"] = inventory_risk
            campaign["inventoryDaysCover"] = inventory_row.get("daysCoverNetwork") if inventory_row else None
            campaign.update(recommend_campaign(
                spend=float(campaign["spend"]),
                advertising_sales=float(campaign["advertisingSales"]),
                orders=int(campaign["orders"]),
                clicks=int(campaign["clicks"]),
                impressions=int(campaign["impressions"]),
                budget=float(campaign["budget"]),
                period_days=int(campaign["periodDays"]),
                inventory_risk=inventory_risk,
                parameters=advertising_parameters,
            ))
        advertising_campaigns.sort(key=lambda item: (
            {"PAUSE_STOCK_RISK": 0, "NO_ORDER_REVIEW": 1, "REDUCE_BID_OR_BUDGET": 2, "INCREASE_BUDGET": 3, "INCREASE_BID": 4, "EXPAND_WINNER": 5, "MONITOR": 6, "NO_CHANGE_LOW_DATA": 7}.get(item["action"], 8),
            -float(item["spend"]),
        ))
        latest_advertising_month = advertising_monthly[-1]["month"] if advertising_monthly else None
        latest_advertising_end = date.fromisoformat(f"{latest_advertising_month}-01") if latest_advertising_month else None
        advertising_age_days = (max(fba_date, awd_date) - latest_advertising_end).days if latest_advertising_end else None

        if alignment_gap > tolerance:
            db.add_exception(run_id, category="inventory_snapshot_misaligned", severity="high", source="inventory-dashboard", raw=None, base=None, cell=None, details={"fba_date": fba_date.isoformat(), "awd_date": awd_date.isoformat(), "gap_days": alignment_gap})
        if age_days > stale_after:
            db.add_exception(run_id, category="inventory_snapshot_stale", severity="high", source="inventory-dashboard", raw=None, base=None, cell=None, details={"snapshot_date": max(fba_date, awd_date).isoformat(), "age_days": age_days})

        payload = {
            "schemaVersion": 2,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "market": market,
            "currency": settings.get("currency", "USD"),
            "domesticPool": {
                "id": settings.get("domestic_pool_id", "CN-SHARED"),
                "sharedAcrossMarkets": bool(settings.get("domestic_pool_shared", True)),
                "markets": [str(item).upper() for item in settings.get("domestic_pool_markets", ["US", "CA"])],
                "sourceSheet": domestic_master_sheet,
            },
            "snapshots": {
                "fbaDate": fba_date.isoformat(),
                "awdDate": awd_date.isoformat(),
                "alignmentGapDays": alignment_gap,
                "aligned": alignment_gap <= tolerance,
                "ageDays": age_days,
                "staleAfterDays": stale_after,
                "isStale": age_days > stale_after,
                "awdSourceAvailable": awd_enabled,
            },
            "sales": {
                "sourceKind": sales_source_kind,
                "sourceSheet": sales_source_sheet,
                "windowMonths": sales_months,
                "historyMonths": history_months,
                "ageDaysAtSnapshot": sales_age_days,
                "confidence": "low" if sales_age_days > 62 else "medium" if sales_age_days > 31 else "high",
                "method": sales_method,
                "historyMethod": "按月导出、月度销售报告与最新库存规划月销合并后的 SKU 实际销量",
            },
            "businessPerformance": {
                "actualYear": actual_year,
                "forecastYear": forecast_year,
                "series": performance_series,
                "summary": {
                    "annualActualUnits": annual_actual_units,
                    "annualActualRevenue": annual_actual_revenue,
                    "annualAdvertisingRevenue": annual_advertising_revenue,
                    "latestMonthUnits": latest_actual["actualUnits"],
                    "latestMonthRevenue": latest_actual["actualRevenue"],
                    "latestMonthUnitChangePercent": latest_unit_change,
                },
            },
            "advertising": {
                "latestMonth": latest_advertising_month,
                "ageDaysAtSnapshot": advertising_age_days,
                "confidence": "low" if advertising_age_days is None or advertising_age_days > 62 else "medium",
                "parameters": {
                    "targetAcosPercent": advertising_parameters.target_acos_percent,
                    "minimumEvidenceSpend": advertising_parameters.minimum_evidence_spend,
                    "noOrderSpend": advertising_parameters.no_order_spend,
                    "winnerMinOrders": advertising_parameters.winner_min_orders,
                    "scaleMinOrders": advertising_parameters.scale_min_orders,
                    "lowVolumeMaxClicks": advertising_parameters.low_volume_max_clicks,
                    "budgetUtilizationThresholdPercent": advertising_parameters.budget_utilization_threshold_percent,
                    "scaleMaxAcosRatio": advertising_parameters.scale_max_acos_ratio,
                },
                "monthlySeries": advertising_monthly,
                "campaigns": advertising_campaigns,
            },
            "parameters": {
                "leadTimeDays": parameters.lead_time_days,
                "reviewCycleDays": parameters.review_cycle_days,
                "targetCoverDays": parameters.target_cover_days,
                "safetyStockDays": parameters.safety_stock_days,
                "excessCoverDays": parameters.excess_cover_days,
                "fbaTransferTriggerDays": parameters.fba_transfer_trigger_days,
                "purchaseOrderOverdueDays": purchase_order_overdue_days,
            },
            "sources": [
                _source_meta(config, fba_path, "fba_inventory"),
                *([_source_meta(config, awd_path, "awd_inventory")] if awd_path else []),
                *[_source_meta(config, path, "sales_month") for path in sales_source_paths],
                _source_meta(config, master_path, "product_master"),
                *[_source_meta(config, path, "advertising_campaign_month") for path in advertising_paths],
            ],
            "summary": {
                "skuCount": len(rows),
                "fbaSellable": _sum(rows, "fbaSellable"),
                "awdAvailable": _sum(rows, "awdAvailable"),
                "awdOutboundToFba": _sum(rows, "awdOutboundToFba"),
                "awdInboundNotCounted": _sum(rows, "awdInbound"),
                "inTransitInventory": _sum(rows, "inTransitInventory"),
                "localInventory": _sum(rows, "localInventory"),
                "pendingOrderQty": _sum(rows, "pendingOrderQty"),
                "overdueOrderCount": sum(1 for row in rows for order in row["pendingOrders"] if order["overdue"]),
                "overduePurchaseOrderCount": len({order["poNumber"] for row in rows for order in row["pendingOrders"] if order["overdue"]}),
                "overdueOrderSkuCount": sum(1 for row in rows if any(order["overdue"] for order in row["pendingOrders"])),
                "readyToShipQty": _sum(rows, "readyToShipQty"),
                "suggestedProductionQty": _sum(rows, "suggestedProductionQty"),
                "criticalSkuCount": sum(1 for row in rows if row["riskLevel"] == "critical"),
                "reviewSkuCount": sum(1 for row in rows if row["action"] == "REVIEW_DATA"),
                "suggestedShipmentQty": _sum(rows, "suggestedShipmentQty"),
            },
            "dataQuality": {
                "invalidFbaSkuRows": len(invalid_fba),
                "invalidAwdSkuRows": len(invalid_awd),
                "missingSalesSkuCount": sum(1 for row in rows if "missing_sales_baseline" in row["dataQuality"]),
                "missingCartonSkuCount": sum(1 for row in rows if "missing_carton_quantity" in row["dataQuality"]),
            },
            "rows": rows,
        }
        report_name = "inventory_dashboard.json" if market == "US" else f"inventory_dashboard.{market.lower()}.json"
        report = config.runtime_root / "reports" / report_name
        _atomic_json(report, payload)
        summary = {
            "run_id": run_id,
            "report_path": str(report),
            "snapshot_date": max(fba_date, awd_date).isoformat(),
            **payload["summary"],
            "data_quality": payload["dataQuality"],
        }
        db.finish_run(run_id, "completed", summary=summary)
        return summary
    except Exception as exc:
        db.finish_run(run_id, "failed", error=repr(exc))
        raise


def run(config: ProjectConfig, db: StateDb) -> dict:
    base = dict(config.inventory_dashboard)
    if not base:
        raise ValueError("inventory_dashboard is not configured")
    market_overrides = dict(base.pop("markets", {}))
    settings_by_market = {str(base.get("market", "US")).upper(): base}
    for market, overrides in market_overrides.items():
        settings_by_market[str(market).upper()] = {**base, **dict(overrides)}
    product_catalog = run_product_catalog(config, db)
    content_workflow = run_content_workflow(config, db)
    document_master = run_document_master(config, db)
    document_master_path = config.runtime_root / "reports" / "document_master.json"
    document_master_data = json.loads(document_master_path.read_text(encoding="utf-8"))
    purchase_order_lots = document_master_data.get("purchaseOrderLots", [])
    latest_archive_source = str(document_master_data.get("sources", {}).get("purchaseOrderArchive", ""))
    summaries = {
        market: _run_market(config, db, settings, purchase_order_lots, latest_archive_source)
        for market, settings in settings_by_market.items()
    }
    return {"product_catalog": product_catalog, "content_workflow": content_workflow, "document_master": document_master, "markets": summaries}
