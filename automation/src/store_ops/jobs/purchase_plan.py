from __future__ import annotations

import calendar
import json
import math
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from ..config import ProjectConfig
from ..db import StateDb
from ..sku import SkuNormalizer


def _atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _number(value: object) -> float:
    if value in (None, "", "-") or isinstance(value, bool):
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _compact(value: object) -> str:
    return re.sub(r"\s+", "", "" if value is None else str(value).strip())


def _read_order_allocations(config: ProjectConfig, normalizer: SkuNormalizer) -> tuple[dict[str, dict], list[dict]]:
    settings = config.inventory_dashboard
    purchase_settings = settings.get("purchase_plan", {})
    workbook_path = (config.data_root / str(settings["master_workbook"])).resolve()
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    allocations: dict[str, dict] = defaultdict(lambda: {"US": 0, "CA": 0, "productName": "", "factory": "", "cartonQty": None})
    sources = []
    try:
        configurations = [
            ("US", str(settings["master_sheet"]), str(purchase_settings.get("us_order_header", "US下单"))),
            ("CA", str(settings.get("markets", {}).get("CA", {}).get("master_sheet", "加拿大库存计划 ")), str(purchase_settings.get("ca_order_header", "CA下单"))),
        ]
        for market, sheet_name, order_header in configurations:
            sheet = workbook[sheet_name]
            header = [_compact(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            indexes = {value: index for index, value in enumerate(header) if value}
            sku_index = indexes.get("MSKU", indexes.get("SKU"))
            order_index = indexes.get(_compact(order_header))
            if sku_index is None or order_index is None:
                raise ValueError(f"{sheet_name} 缺少 SKU 或 {order_header} 列")
            name_index = indexes.get("品名")
            factory_index = indexes.get("工厂")
            carton_index = indexes.get("装箱数")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                match = normalizer.extract(row[sku_index] if sku_index < len(row) else None)
                sku = match.bases[0] if match and len(match.bases) == 1 else None
                quantity = int(round(_number(row[order_index] if order_index < len(row) else None)))
                if not sku:
                    continue
                item = allocations[sku]
                item[market] = max(0, quantity)
                if name_index is not None and name_index < len(row) and row[name_index]:
                    item["productName"] = str(row[name_index]).strip()
                if factory_index is not None and factory_index < len(row) and row[factory_index]:
                    item["factory"] = str(row[factory_index]).strip()
                carton = int(round(_number(row[carton_index] if carton_index is not None and carton_index < len(row) else None)))
                if carton > 0:
                    item["cartonQty"] = carton
            sources.append({
                "kind": f"{market.lower()}_purchase_allocation",
                "path": str(workbook_path.relative_to(config.data_root)).replace("\\", "/"),
                "sheet": sheet_name,
                "column": get_column_letter(order_index + 1),
                "header": order_header,
            })
    finally:
        workbook.close()
    return dict(allocations), sources


def _po_date(po_number: str) -> date | None:
    match = re.search(r"(\d{6})", po_number)
    if not match:
        return None
    digits = match.group(1)
    try:
        return date(2000 + int(digits[:2]), int(digits[2:4]), int(digits[4:6]))
    except ValueError:
        return None


def _next_cycle(latest_order_date: date) -> tuple[str, date, str]:
    if latest_order_date.day <= 20:
        last_day = calendar.monthrange(latest_order_date.year, latest_order_date.month)[1]
        return "mid_month", date(latest_order_date.year, latest_order_date.month, last_day), "month_end"
    if latest_order_date.month == 12:
        return "month_end", date(latest_order_date.year + 1, 1, 15), "mid_month"
    return "month_end", date(latest_order_date.year, latest_order_date.month + 1, 15), "mid_month"


def _reconciliation_status(planned: int, ordered: int) -> str:
    if planned > 0 and ordered > 0:
        if planned == ordered:
            return "MATCHED"
        return "ORDERED_MORE" if ordered > planned else "ORDERED_LESS"
    if planned > 0:
        return "PLAN_ONLY"
    if ordered > 0:
        return "ORDER_ONLY"
    return "NO_ORDER"


def _round_to_multiple(quantity: int, multiple: int) -> int:
    if quantity <= 0:
        return 0
    return int(math.ceil(quantity / max(1, multiple)) * max(1, multiple))


def run(config: ProjectConfig, db: StateDb) -> dict:
    run_id = db.start_run("build-purchase-plan")
    try:
        reports = config.runtime_root / "reports"
        output = reports / "purchase_plan.json"
        previous_source_document_count = 0
        if output.exists():
            try:
                previous_source_document_count = int(
                    json.loads(output.read_text(encoding="utf-8")).get("summary", {}).get("sourceDocumentCount", 0) or 0
                )
            except (OSError, ValueError, TypeError, json.JSONDecodeError):
                previous_source_document_count = 0
        normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
        allocations, allocation_sources = _read_order_allocations(config, normalizer)
        us = json.loads((reports / "inventory_dashboard.json").read_text(encoding="utf-8"))
        ca = json.loads((reports / "inventory_dashboard.ca.json").read_text(encoding="utf-8"))
        document_master = json.loads((reports / "document_master.json").read_text(encoding="utf-8"))
        products_path = reports / "product_catalog.json"
        products = json.loads(products_path.read_text(encoding="utf-8")) if products_path.exists() else {"items": []}

        archive_source = document_master.get("sources", {}).get("purchaseOrderArchive")
        latest_lots = [
            lot for lot in document_master.get("purchaseOrderLots", [])
            if archive_source and str(lot.get("sourcePath", "")).startswith(f"{archive_source}::")
        ]
        latest_dates = [item for item in (_po_date(str(lot.get("poNumber", ""))) for lot in latest_lots) if item]
        latest_order_date = max(latest_dates) if latest_dates else date.today()
        cycle_stage, next_review_date, next_stage = _next_cycle(latest_order_date)
        latest_lots = [lot for lot in latest_lots if _po_date(str(lot.get("poNumber", ""))) == latest_order_date]

        ordered_by_sku: dict[str, int] = defaultdict(int)
        order_details: dict[str, list[dict]] = defaultdict(list)
        for lot in latest_lots:
            sku = str(lot.get("sku", ""))
            quantity = int(lot.get("orderedQuantity", 0) or 0)
            ordered_by_sku[sku] += quantity
            order_details[sku].append({
                "poNumber": lot.get("poNumber", ""),
                "factory": lot.get("factory", ""),
                "quantity": quantity,
                "unitPrice": float(lot.get("unitPrice", 0) or 0),
                "sourcePath": lot.get("sourcePath", ""),
            })

        us_rows = {row["sku"]: row for row in us["rows"]}
        ca_rows = {row["sku"]: row for row in ca["rows"]}
        products_by_sku = {item["sku"]: item for item in products.get("items", [])}
        purchase_settings = config.inventory_dashboard.get("purchase_plan", {})
        production_lead_days = int(purchase_settings.get("production_lead_days", 45))
        ocean_lead_days = int(purchase_settings.get("ocean_lead_days", config.inventory_dashboard.get("parameters", {}).get("lead_time_days", 75)))
        review_cycle_days = int(purchase_settings.get("review_cycle_days", 15))
        safety_stock_days = int(purchase_settings.get("safety_stock_days", config.inventory_dashboard.get("parameters", {}).get("safety_stock_days", 21)))
        horizon_days = production_lead_days + ocean_lead_days + review_cycle_days + safety_stock_days

        rows = []
        all_skus = sorted(set(us_rows) | set(ca_rows) | set(allocations) | set(ordered_by_sku))
        for sku in all_skus:
            us_row = us_rows.get(sku, {})
            ca_row = ca_rows.get(sku, {})
            allocation = allocations.get(sku, {})
            product = products_by_sku.get(sku, {})
            us_daily = float(us_row.get("dailySales", 0) or 0)
            ca_daily = float(ca_row.get("dailySales", 0) or 0)
            combined_daily = us_daily + ca_daily
            us_network = int(us_row.get("eligibleInventoryPosition", 0) or 0)
            ca_network = int(ca_row.get("eligibleInventoryPosition", 0) or 0)
            shared_row = us_row or ca_row
            local_inventory = int(shared_row.get("localInventory", 0) or 0)
            ordered = ordered_by_sku.get(sku, 0)
            source_pending_quantity = int(shared_row.get("pendingOrderQty", 0) or 0)
            latest_po_numbers = {str(order.get("poNumber", "")) for order in order_details.get(sku, [])}
            reflected_latest_quantity = sum(
                int(order.get("remainingQuantity", 0) or 0)
                for order in shared_row.get("pendingOrders", [])
                if str(order.get("poNumber", "")) in latest_po_numbers
            )
            unreflected_latest_quantity = max(0, ordered - reflected_latest_quantity)
            pending_quantity = source_pending_quantity + unreflected_latest_quantity
            inventory_position = us_network + ca_network + local_inventory + pending_quantity
            projected_demand = int(math.ceil(combined_daily * horizon_days))
            gross_need = max(0, projected_demand - inventory_position)
            carton_qty = int(product.get("cartonQty") or allocation.get("cartonQty") or us_row.get("cartonQty") or ca_row.get("cartonQty") or 1)
            suggested_quantity = _round_to_multiple(gross_need, carton_qty)
            us_planned = int(allocation.get("US", 0) or 0)
            ca_planned = int(allocation.get("CA", 0) or 0)
            planned = us_planned + ca_planned
            cover_days = inventory_position / combined_daily if combined_daily > 0 else None
            if combined_daily <= 0:
                risk = "data"
            elif cover_days is not None and cover_days < production_lead_days + ocean_lead_days:
                risk = "critical"
            elif cover_days is not None and cover_days < horizon_days:
                risk = "watch"
            else:
                risk = "healthy"
            rows.append({
                "sku": sku,
                "productName": allocation.get("productName") or us_row.get("productName") or ca_row.get("productName") or product.get("chineseName") or "",
                "factory": (order_details.get(sku) or [{}])[0].get("factory") or allocation.get("factory") or us_row.get("factory") or ca_row.get("factory") or "",
                "cartonQty": carton_qty,
                "usDailySales": round(us_daily, 4),
                "caDailySales": round(ca_daily, 4),
                "combinedDailySales": round(combined_daily, 4),
                "usNetworkInventory": us_network,
                "caNetworkInventory": ca_network,
                "localInventory": local_inventory,
                "sourcePendingOrderQty": source_pending_quantity,
                "unreflectedLatestOrderQty": unreflected_latest_quantity,
                "pendingOrderQty": pending_quantity,
                "inventoryPosition": inventory_position,
                "projectedDemand": projected_demand,
                "coverageDays": round(cover_days, 1) if cover_days is not None else None,
                "suggestedPurchaseQty": suggested_quantity,
                "usPlannedQty": us_planned,
                "caPlannedQty": ca_planned,
                "manualPlannedQty": planned,
                "actualOrderedQty": ordered,
                "varianceQty": ordered - planned,
                "reconciliationStatus": _reconciliation_status(planned, ordered),
                "riskLevel": risk,
                "orders": order_details.get(sku, []),
            })

        manual_rows = [row for row in rows if row["manualPlannedQty"] > 0]
        actual_rows = [row for row in rows if row["actualOrderedQty"] > 0]
        next_rows = [row for row in rows if row["suggestedPurchaseQty"] > 0]
        discrepancy_rows = [row for row in rows if row["reconciliationStatus"] not in {"MATCHED", "NO_ORDER"}]
        source_document_count = previous_source_document_count
        if archive_source:
            archive_path = config.data_root / archive_source
            if archive_path.exists():
                with zipfile.ZipFile(archive_path) as archive:
                    source_document_count = sum(1 for item in archive.infolist() if not item.is_dir() and Path(item.filename).suffix.lower() in {".xls", ".xlsx"})

        payload = {
            "schemaVersion": 1,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "cycle": {
                "latestOrderDate": latest_order_date.isoformat(),
                "latestStage": cycle_stage,
                "nextReviewDate": next_review_date.isoformat(),
                "nextStage": next_stage,
            },
            "parameters": {
                "productionLeadDays": production_lead_days,
                "oceanLeadDays": ocean_lead_days,
                "reviewCycleDays": review_cycle_days,
                "safetyStockDays": safety_stock_days,
                "demandHorizonDays": horizon_days,
            },
            "summary": {
                "manualPlanSkuCount": len(manual_rows),
                "manualPlanQuantity": sum(row["manualPlannedQty"] for row in manual_rows),
                "actualOrderSkuCount": len(actual_rows),
                "actualOrderQuantity": sum(row["actualOrderedQty"] for row in actual_rows),
                "actualPurchaseOrderCount": len({order["poNumber"] for row in actual_rows for order in row["orders"]}),
                "sourceDocumentCount": source_document_count,
                "matchedSkuCount": sum(1 for row in rows if row["reconciliationStatus"] == "MATCHED"),
                "discrepancySkuCount": len(discrepancy_rows),
                "varianceQuantity": sum(row["varianceQty"] for row in rows),
                "nextCycleSkuCount": len(next_rows),
                "nextCycleQuantity": sum(row["suggestedPurchaseQty"] for row in next_rows),
                "criticalSkuCount": sum(1 for row in next_rows if row["riskLevel"] == "critical"),
            },
            "sources": [
                *allocation_sources,
                {"kind": "purchase_order_archive", "path": archive_source or "", "documentCount": source_document_count},
                {"kind": "inventory_dashboard", "path": "runtime/reports/inventory_dashboard.json"},
                {"kind": "inventory_dashboard_ca", "path": "runtime/reports/inventory_dashboard.ca.json"},
            ],
            "rows": rows,
        }
        _atomic_json(output, payload)
        summary = {"status": "completed", "output": str(output), **payload["summary"]}
        db.finish_run(run_id, "completed", summary)
        return summary
    except Exception as error:
        db.finish_run(run_id, "failed", error=str(error))
        raise
