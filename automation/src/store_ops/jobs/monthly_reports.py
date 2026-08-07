from __future__ import annotations

import hashlib
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from ..config import ProjectConfig
from ..db import StateDb


REPORT_NAME_RE = re.compile(
    r"(20\d{2})[._-]?(0?[1-9]|1[0-2])月.*销售和毛利报告-(US|CA|MX)\.xlsx$",
    re.IGNORECASE,
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _report_identity(path: Path) -> tuple[str, str]:
    match = REPORT_NAME_RE.search(path.name)
    if not match:
        raise ValueError(f"无法从文件名识别月份和站点: {path.name}")
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}", match.group(3).upper()


def _validate_report(path: Path) -> tuple[str, str, int]:
    if not path.is_file():
        raise FileNotFoundError(path)
    report_month, market = _report_identity(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "SKU销售汇总" not in workbook.sheetnames:
            raise ValueError(f"月报缺少 SKU销售汇总: {path}")
        worksheet = workbook["SKU销售汇总"]
        data_rows = sum(1 for row in worksheet.iter_rows(min_row=4, values_only=True) if row and row[0])
        if data_rows == 0:
            raise ValueError(f"月报 SKU销售汇总 无有效数据: {path}")
    finally:
        workbook.close()
    return report_month, market, data_rows


def run(config: ProjectConfig, db: StateDb, sources: list[Path]) -> dict:
    run_id = db.start_run("import-monthly-reports")
    try:
        if not sources:
            raise ValueError("至少需要一个 --source 月报文件")
        imported = []
        for source in sources:
            resolved = source.expanduser().resolve()
            report_month, market, data_rows = _validate_report(resolved)
            digest = _sha256(resolved)
            destination_dir = config.runtime_root / "incoming" / "monthly-sales-reports" / report_month / market
            destination_dir.mkdir(parents=True, exist_ok=True)
            destination = destination_dir / resolved.name
            temporary = destination.with_suffix(destination.suffix + ".tmp")
            shutil.copy2(resolved, temporary)
            temporary.replace(destination)
            imported.append({
                "market": market,
                "reportMonth": report_month,
                "sourcePath": str(resolved),
                "storedPath": str(destination),
                "sha256": digest,
                "dataRows": data_rows,
                "importedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            })
        summary = {"run_id": run_id, "imported": imported, "count": len(imported)}
        db.finish_run(run_id, "completed", summary=summary)
        return summary
    except Exception as exc:
        db.finish_run(run_id, "failed", error=repr(exc))
        raise
