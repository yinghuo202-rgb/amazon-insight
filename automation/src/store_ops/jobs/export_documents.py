from __future__ import annotations

import json
import os
import re
import subprocess
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

from ..config import ProjectConfig
from ..db import StateDb
from ..documents import PurchaseOrderLot, PurchaseOrderShortage, allocate_purchase_orders_fifo
from ..ooxml_layout import restore_layout_metadata


def _read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _atomic_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def _factory_matches(expected: str, actual: str) -> bool:
    expected = expected.strip().replace("有限公司", "")
    actual = actual.strip().replace("有限公司", "")
    aliases = {"宁水": ("宁波水表",)}
    return not expected or expected in actual or actual in expected or any(token in actual for token in aliases.get(expected, ()))


def _declaration_layout(path: Path, market: str) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        sheet = workbook["报关空表"]
        header_row = next((row for row in range(1, min(sheet.max_row, 40) + 1) if _cell_text(sheet.cell(row, 1).value) == "合同号"), None)
        if header_row is None:
            raise ValueError(f"无法识别报运单模板的表头：{path}")
        headers: dict[str, int] = {}
        for column in range(1, min(sheet.max_column, 32) + 1):
            label = re.sub(r"\s+", "", _cell_text(sheet.cell(header_row, column).value)).lower()
            if label:
                # Historical templates sometimes repeat helper headers after the
                # visible table. The first occurrence is the business column.
                headers.setdefault(label, column)
        required = {
            "poNumber": "合同号",
            "factory": "工厂",
            "productName": "品名",
            "quantity": "数量",
            "cartons": "箱数",
            "weightKg": "重量",
            "volumeM3": "体积",
            "purchaseAmountRmb": "采购金额",
            "note": "备注",
            "taxRate": "退税率",
        }
        columns = {key: headers.get(label.lower()) for key, label in required.items()}
        if any(value is None for value in columns.values()):
            missing = "、".join(required[key] for key, value in columns.items() if value is None)
            raise ValueError(f"报运单模板缺少列：{missing}（{path}）")
        total_row = next((row for row in range(header_row + 1, min(sheet.max_row, 300) + 1) if _cell_text(sheet.cell(row, 1).value) == "合计"), None)
        if total_row is None:
            raise ValueError(f"无法识别报运单模板的数据区：{path}")
        shipment_date_column = next((column for column in range(1, 16) if _cell_text(sheet.cell(2, column).value) == "出运日期"), None)
        if shipment_date_column is None:
            raise ValueError(f"无法识别报运单模板的出运日期列：{path}")
        declaration_sku_column = next((column for label, column in headers.items() if label in {"sku", "产品编号", "采购产品编号"}), None)
        return {
            "dataStart": header_row + 1,
            "dataEnd": total_row - 1,
            "totalRow": total_row,
            "shipmentDateCell": f"{sheet.cell(3, shipment_date_column).coordinate}",
            "columns": {
                **columns,
                "declarationSku": declaration_sku_column or 13,
                "unitPriceRmb": 14,
                "lineAmountRmb": 15,
            },
        }
    finally:
        workbook.close()


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _node_executable() -> Path:
    configured = os.environ.get("STORE_OPS_NODE", "").strip()
    if configured:
        return Path(configured)
    bundled_bin = Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "node" / "bin"
    candidates = [bundled_bin / "node.exe"] if os.name == "nt" else [bundled_bin / "node"]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return Path("node")


def _validate_request(request: dict) -> tuple[str, list[str], list[dict]]:
    market = str(request.get("market", "US")).upper()
    if market not in {"US", "CA"}:
        raise ValueError("market 必须是 US 或 CA")
    document_types = [str(item) for item in request.get("documentTypes", [])]
    if not document_types or any(item not in {"shipment", "declaration"} for item in document_types):
        raise ValueError("documentTypes 仅支持 shipment/declaration")
    entries = request.get("entries") or []
    if not entries:
        raise ValueError("至少选择一个 SKU")
    normalized = []
    seen = set()
    for item in entries:
        sku = str(item.get("sku", "")).strip().upper()
        quantity = int(item.get("quantity", 0))
        if not re.fullmatch(r"[A-Z]{2}\d{3}", sku) or quantity <= 0:
            raise ValueError(f"无效 SKU 或数量：{sku}")
        if sku in seen:
            raise ValueError(f"SKU 重复：{sku}")
        seen.add(sku)
        normalized.append({"sku": sku, "quantity": quantity})
    sku_sort = str(request.get("skuSort", "asc")).lower()
    if sku_sort not in {"asc", "desc"}:
        raise ValueError("skuSort 仅支持 asc/desc")
    normalized.sort(key=lambda item: item["sku"], reverse=sku_sort == "desc")
    return market, document_types, normalized


def _purchase_lots(master: dict) -> list[PurchaseOrderLot]:
    return [
        PurchaseOrderLot(
            po_number=item["poNumber"],
            po_date=date.fromisoformat(item["poDate"]),
            sku=item["sku"],
            factory=item.get("factory", ""),
            ordered_quantity=int(item["orderedQuantity"]),
            previously_shipped_quantity=int(item.get("previouslyShippedQuantity", 0)),
            unit_price=float(item.get("unitPrice", 0)),
            product_name=item.get("productName", ""),
        )
        for item in master.get("purchaseOrderLots", [])
    ]


def _shipment_rows(entries: list[dict], logistics: dict, template_source: str) -> list[dict]:
    rows = []
    problems = []
    for entry in entries:
        sku, quantity = entry["sku"], entry["quantity"]
        item = logistics.get(sku)
        if not item:
            problems.append(f"{sku} 缺少历史物流参数")
            continue
        required = ("cartonQty", "grossWeightKg", "lengthCm", "widthCm", "heightCm", "cartonVolumeM3")
        if any(item.get(key) in (None, 0) for key in required):
            problems.append(f"{sku} 箱规、重量或尺寸不完整")
            continue
        carton = int(item["cartonQty"])
        if quantity % carton != 0:
            problems.append(f"{sku} 数量 {quantity} 不是装箱量 {carton} 的整数倍")
            continue
        cartons = quantity // carton
        rows.append({
            **{key: item.get(key) for key in ("sku", "cartonQty", "netWeightKg", "grossWeightKg", "lengthCm", "widthCm", "heightCm", "cartonVolumeM3", "unitPriceRmb")},
            "quantity": quantity,
            "cartons": cartons,
            "totalWeightKg": round(float(item["grossWeightKg"]) * cartons, 4),
            "totalVolumeM3": round(float(item["cartonVolumeM3"]) * cartons, 6),
            "imageFormula": item.get("imageFormula", "") if item.get("sourcePath") == template_source else "",
        })
    if problems:
        raise ValueError("；".join(problems))
    return rows


def _chunk_rows(rows: list[dict], capacity: int) -> list[list[dict]]:
    if capacity <= 0:
        raise ValueError("模板容量必须大于 0")
    return [rows[index:index + capacity] for index in range(0, len(rows), capacity)]


def _split_total(total: float, quantities: list[int], precision: int) -> list[float]:
    quantity_total = sum(quantities)
    if not quantities or quantity_total <= 0:
        return []
    values = [round(total * quantity / quantity_total, precision) for quantity in quantities[:-1]]
    values.append(round(total - sum(values), precision))
    return values


def _declaration_rows(entries: list[dict], logistics: dict, profiles: dict, lots: list[PurchaseOrderLot], payments_by_po: dict | None = None) -> list[dict]:
    rows = []
    problems = []
    payments_by_po = payments_by_po or {}
    available_by_lot = {(lot.po_number, lot.sku, lot.factory): lot.available_quantity for lot in lots}
    for entry_index, entry in enumerate(entries):
        sku, parent_quantity = entry["sku"], entry["quantity"]
        logistics_item = logistics.get(sku)
        profile = profiles.get(sku)
        if not logistics_item:
            problems.append(f"{sku} 缺少物流参数")
            continue
        if not profile or not profile.get("components"):
            problems.append(f"{sku} 缺少历史报关组成")
            continue
        for component_index, component in enumerate(profile["components"]):
            exact_quantity = parent_quantity * float(component.get("quantityFactor", 1))
            component_quantity = int(round(exact_quantity))
            if abs(component_quantity - exact_quantity) > 1e-6:
                problems.append(f"{sku} 的报关组成数量无法取整")
                continue
            matching_lots = [lot for lot in lots if lot.sku == sku and _factory_matches(component.get("factory", ""), lot.factory)]
            try:
                allocations = allocate_purchase_orders_fifo(sku, component_quantity, matching_lots)
            except PurchaseOrderShortage as error:
                problems.append(f"{component.get('declarationSku', sku)}（{component.get('factory', '')}）需要 {error.requested}，可用采购批次仅 {error.available}")
                continue
            carry_logistics = component_index == 0
            if carry_logistics:
                carton = int(logistics_item["cartonQty"])
                total_cartons = parent_quantity / carton
                carton_values = _split_total(total_cartons, [item.quantity for item in allocations], 4)
                weight_values = _split_total(float(logistics_item["grossWeightKg"]) * total_cartons, [item.quantity for item in allocations], 4)
                volume_values = _split_total(float(logistics_item["cartonVolumeM3"]) * total_cartons, [item.quantity for item in allocations], 6)
            else:
                carton_values = [0.0] * len(allocations)
                weight_values = [0.0] * len(allocations)
                volume_values = [0.0] * len(allocations)
            for allocation_index, allocation in enumerate(allocations):
                if allocation.unit_price <= 0:
                    problems.append(f"{allocation.po_number} / {component.get('declarationSku', sku)} 缺少采购单价")
                    continue
                payment = payments_by_po.get(allocation.po_number, {})
                note_parts = []
                payers = [str(value) for value in payment.get("paymentPayers", []) if str(value).strip()]
                methods = [str(value) for value in payment.get("paymentMethods", []) if str(value).strip()]
                if payers:
                    note_parts.append(f"{'/'.join(payers)}已付款" if len(payers) == 1 else f"付款方：{'/'.join(payers)}")
                if methods:
                    note_parts.append(f"付款方式：{'/'.join(methods)}")
                available = available_by_lot.get((allocation.po_number, allocation.sku, allocation.factory), 0)
                if allocation.quantity >= available > 0:
                    note_parts.append("货出完了")
                rows.append({
                    "entryIndex": entry_index,
                    "componentIndex": component_index,
                    "parentSku": sku,
                    "declarationSku": component.get("declarationSku") or sku,
                    "poNumber": allocation.po_number,
                    "poDate": allocation.po_date.isoformat(),
                    "factory": component.get("factory") or allocation.factory,
                    "productName": component.get("productName") or allocation.product_name,
                    "quantity": allocation.quantity,
                    "cartons": carton_values[allocation_index],
                    "weightKg": weight_values[allocation_index],
                    "volumeM3": volume_values[allocation_index],
                    "unitPriceRmb": allocation.unit_price,
                    "purchaseAmountRmb": round(allocation.quantity * allocation.unit_price, 2),
                    "taxRate": float(component.get("taxRate", 0.13)),
                    "note": "\n".join(note_parts),
                })
    if problems:
        raise ValueError("采购批次预检未通过：" + "；".join(problems))
    rows.sort(key=lambda item: (item["entryIndex"], item["componentIndex"], item["poDate"], item["poNumber"]))
    return rows


def run(config: ProjectConfig, db: StateDb, request_path: Path) -> dict:
    run_id = db.start_run("export-documents")
    try:
        request = _read_json(request_path)
        market, document_types, entries = _validate_request(request)
        master_path = config.runtime_root / "reports" / "document_master.json"
        if not master_path.exists():
            raise FileNotFoundError("缺少 document_master.json，请先运行 build-document-master")
        master = _read_json(master_path)
        batch_number = str(request.get("batchNumber", "")).strip().upper()
        if not re.fullmatch(r"CM\d{3,4}", batch_number):
            raise ValueError("批次号格式应为 CM319")
        shipment_date = date.fromisoformat(str(request.get("shipmentDate"))).isoformat()
        export_id = f"{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:8]}"
        output_dir = config.runtime_root / "output" / "exports" / export_id
        output_dir.mkdir(parents=True, exist_ok=False)
        preview_dir = output_dir / "previews"
        if request.get("renderPreviews"):
            preview_dir.mkdir()
        logistics = master["logistics"].get(market, {})
        templates = master.get("templates", {})
        plan = {"resultPath": str(output_dir / "result.json"), "files": [], "shipments": [], "declarations": []}

        if "shipment" in document_types:
            retained = Path.home() / ".codex" / "skills" / "artifact-template-measureman" / "assets" / "reference.xlsx"
            source_template = templates["shipment"][market]
            template_path = retained if market == "US" and retained.exists() else config.data_root / source_template
            rows = _shipment_rows(entries, logistics, source_template)
            parts = _chunk_rows(rows, 43)
            for part_index, part_rows in enumerate(parts, start=1):
                suffix = "" if len(parts) == 1 else f"-分册{part_index:02d}of{len(parts):02d}"
                filename = f"{batch_number}-{market}-{shipment_date.replace('-', '')}发货清单{suffix}.xlsx"
                document = {
                    "market": market,
                    "templatePath": str(template_path),
                    "outputPath": str(output_dir / filename),
                    "previewPath": str(preview_dir / f"{batch_number}-{market}-发货清单{suffix}.png") if request.get("renderPreviews") else "",
                    "title": f"1店{'加拿大' if market == 'CA' else '美国'}发货清单{batch_number}" + (f"（分册 {part_index}/{len(parts)}）" if len(parts) > 1 else ""),
                    "rows": part_rows,
                }
                plan["shipments"].append(document)
                plan["files"].append({"type": "shipment", "filename": filename, "part": part_index, "partCount": len(parts)})

        if "declaration" in document_types:
            template_path = config.data_root / templates["declaration"][market]
            layout = _declaration_layout(template_path, market)
            rows = _declaration_rows(
                entries,
                logistics,
                master["declarationProfiles"].get(market, {}),
                _purchase_lots(master),
                {item["poNumber"]: {
                    "paymentMethods": item.get("paymentMethods", []),
                    "paymentPayers": item.get("paymentPayers", []),
                } for item in master.get("purchaseOrderLots", [])},
            )
            capacity = layout["dataEnd"] - layout["dataStart"] + 1
            date_token = shipment_date.replace("-", "")
            shipment_id = str(request.get("shipmentId", "")).strip()
            tracking_id = str(request.get("trackingId", "")).strip()
            marks = "\n".join(item for item in (f"货件编号：{shipment_id}" if shipment_id else "", f"货件追踪编号:{tracking_id}" if tracking_id else "") if item)
            parts = _chunk_rows(rows, capacity)
            for part_index, part_rows in enumerate(parts, start=1):
                suffix = "" if len(parts) == 1 else f"-分册{part_index:02d}of{len(parts):02d}"
                filename = f"{batch_number}-{market}亚马逊报运单{date_token}提货{suffix}.xlsx"
                invoice_number = str(request.get("invoiceNumber") or f"DRAFT-{batch_number}")
                document = {
                    "market": market,
                    "templatePath": str(template_path),
                    "outputPath": str(output_dir / filename),
                    "previewPath": str(preview_dir / f"{batch_number}-{market}-报运单{suffix}.png") if request.get("renderPreviews") else "",
                    "invoiceNumber": invoice_number + (f"-{part_index:02d}" if len(parts) > 1 else ""),
                    "shipmentDate": shipment_date,
                    "freightReference": str(request.get("freightReference", "")),
                    "shippingMarks": marks,
                    "consignee": str(request.get("consignee", "")),
                    "originPort": str(request.get("originPort", "宁波")),
                    "rows": part_rows,
                    **layout,
                }
                plan["declarations"].append(document)
                plan["files"].append({"type": "declaration", "filename": filename, "part": part_index, "partCount": len(parts)})

        plan_path = output_dir / "plan.json"
        _atomic_json(plan_path, plan)
        builder = config.project_root / "src" / "store_ops" / "export_builder.mjs"
        completed = subprocess.run(
            [str(_node_executable()), str(builder), str(plan_path)],
            cwd=config.project_root,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=180,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or completed.stdout.strip() or "表格生成失败")
        for document in plan["shipments"]:
            restore_layout_metadata(
                Path(document["templatePath"]),
                Path(document["outputPath"]),
                "Measureman",
            )
        for document in plan["declarations"]:
            restore_layout_metadata(
                Path(document["templatePath"]),
                Path(document["outputPath"]),
                "报关空表",
            )
        result = _read_json(output_dir / "result.json")
        payload = {"status": "completed", "exportId": export_id, "files": result["files"]}
        db.finish_run(run_id, "completed", payload)
        return payload
    except Exception as error:
        db.finish_run(run_id, "failed", error=str(error))
        raise
