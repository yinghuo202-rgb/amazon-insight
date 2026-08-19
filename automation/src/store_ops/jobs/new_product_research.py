from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import unquote, urlsplit, urlunsplit

import openpyxl

from ..config import ProjectConfig
from ..db import StateDb


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def _safe_url(value: object) -> str:
    raw = _text(value)
    if not raw.startswith(("http://", "https://")):
        return ""
    parts = urlsplit(raw)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))


RESEARCH_RMB_PER_USD = 7.2


def _compact(value: object) -> str:
    return _text(value).replace("\n", "").replace(" ", "")


def _asin_from_url(value: object) -> str:
    raw = _text(value)
    match = re.search(r"/(?:dp|gp/product)/([A-Z0-9]{10})(?:[/?]|$)", raw, re.IGNORECASE)
    return match.group(1).upper() if match else ""


def _name_from_url(value: object) -> str:
    raw = _safe_url(value)
    if not raw:
        return ""
    path = urlsplit(raw).path
    title_match = re.search(r"/([^/]+)/dp/[A-Z0-9]{10}(?:/|$)", path, re.IGNORECASE)
    slug = unquote(title_match.group(1) if title_match else path.rstrip("/").rsplit("/", 1)[-1])
    if slug.lower() in {"dp", "product"}:
        return ""
    return re.sub(r"[-_]+", " ", slug).strip()[:120]


def _unique_sku(value: str, sheet_name: str, row_number: int, seen: set[str]) -> str:
    base = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff_.-]+", "-", value).strip("-.") or f"RESEARCH-{row_number}"
    sku = base
    if sku in seen:
        sheet_key = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "-", sheet_name).strip("-") or "SHEET"
        sku = f"{base}-{sheet_key}-{row_number}"
    seen.add(sku)
    return sku


def _header_indexes(row: tuple) -> dict[str, int]:
    return {_compact(value): index for index, value in enumerate(row) if _compact(value)}


def _cost_header(indexes: dict[str, int]) -> bool:
    return "现价" in indexes and ("美金成本价" in indexes or "采购价" in indexes) and "利润率" in indexes


def _cost_candidate(sheet_name: str, row_number: int, header: dict[str, int], row: tuple, fallback_name: str = "") -> dict | None:
    price_i = header.get("现价")
    if price_i is None:
        return None
    price = _number(row[price_i] if len(row) > price_i else None)
    if price is None:
        return None
    link_i = next((header[key] for key in ("竞品链接", "链接", "ASIN") if key in header), None)
    link_value = row[link_i] if link_i is not None and len(row) > link_i else None
    url = _safe_url(link_value) or next((_safe_url(value) for value in row if _safe_url(value)), "")
    asin = _text(row[header["ASIN"]]) if "ASIN" in header and len(row) > header["ASIN"] else ""
    asin = asin or _asin_from_url(link_value)
    raw_sku = asin or _name_from_url(link_value) or f"{sheet_name}-{row_number}"
    sku = raw_sku
    purchase_i = header.get("美金成本价", header.get("采购价"))
    purchase_usd = _number(row[purchase_i] if purchase_i is not None and len(row) > purchase_i else None)
    purchase_rmb = purchase_usd * RESEARCH_RMB_PER_USD if purchase_usd is not None else None
    name = fallback_name or _name_from_url(link_value) or asin or f"{sheet_name} 第 {row_number} 行"
    first_mile_i = header.get("海运入FBA仓成本", header.get("头程"))
    untaxed_i = header.get("未税价格")
    untaxed_price = _number(row[untaxed_i]) if untaxed_i is not None and len(row) > untaxed_i else None
    if untaxed_price is not None and untaxed_price <= 0:
        untaxed_price = None
    candidate = {
        "sku": sku,
        "name": name,
        "amazonPrice": price,
        "firstMile": _number(row[first_mile_i]) if first_mile_i is not None and len(row) > first_mile_i else None,
        "storageFee": _number(row[header["FBA仓储费"]]) if "FBA仓储费" in header and len(row) > header["FBA仓储费"] else None,
        "commission": _number(row[header["佣金"]]) if "佣金" in header and len(row) > header["佣金"] else None,
        "orderFee": _number(row[header["订单处理费"]]) if "订单处理费" in header and len(row) > header["订单处理费"] else None,
        "importDutyRate": _number(row[header["进口税"]]) if "进口税" in header and len(row) > header["进口税"] else None,
        "purchaseCostRmb": purchase_rmb,
        "untaxedPriceUsd": untaxed_price,
        "grossProfit": _number(row[header["利润"]]) if "利润" in header and len(row) > header["利润"] else None,
        "grossMargin": _number(row[header["利润率"]]) if len(row) > header["利润率"] else None,
        "competitorUrl": url,
    }
    return _recalculate_candidate(candidate)


def _recalculate_candidate(candidate: dict) -> dict:
    purchase_usd = None if candidate.get("purchaseCostRmb") is None else candidate["purchaseCostRmb"] / RESEARCH_RMB_PER_USD
    untaxed_price_usd = candidate.get("untaxedPriceUsd")
    if untaxed_price_usd is None and purchase_usd is not None:
        untaxed_price_usd = round(purchase_usd * 0.885, 6)
        candidate["untaxedPriceUsd"] = untaxed_price_usd
    costs = [candidate.get("firstMile"), candidate.get("storageFee"), candidate.get("commission"), candidate.get("orderFee"), candidate.get("importDutyRate"), purchase_usd, untaxed_price_usd]
    if candidate.get("amazonPrice") is None or candidate["amazonPrice"] <= 0 or any(value is None for value in costs):
        candidate["totalCostUsd"] = None
        candidate["grossProfit"] = None
        candidate["grossMargin"] = None
    else:
        total_cost = sum(costs)
        gross_profit = candidate["amazonPrice"] - total_cost
        candidate["totalCostUsd"] = total_cost
        candidate["grossProfit"] = gross_profit
        candidate["grossMargin"] = gross_profit / candidate["amazonPrice"]
    return candidate


def _fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _candidate_rows(sheet, seen: set[str] | None = None) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    seen = seen if seen is not None else set()
    legacy_layout = sheet.title.strip() == "Sheet1"
    candidates: list[dict] = []
    for index, row in enumerate(rows):
        sku = _text(row[0] if row else None)
        if legacy_layout and sku and sku not in {"SKU", "序号"} and index + 1 < len(rows):
            detail = rows[index + 1]
            price = _number(detail[2] if len(detail) > 2 else None)
            purchase = _number(detail[9] if len(detail) > 9 else None)
            margin = _number(detail[13] if len(detail) > 13 else None)
            profit = _number(detail[12] if len(detail) > 12 else None)
            if price is not None or purchase is not None or margin is not None:
                unique = _unique_sku(sku, sheet.title, index + 1, seen)
                candidates.append(_recalculate_candidate({
                    "sku": unique,
                    "name": sku,
                    "amazonPrice": price,
                    "firstMile": _number(detail[3] if len(detail) > 3 else None),
                    "storageFee": _number(detail[4] if len(detail) > 4 else None),
                    "commission": _number(detail[5] if len(detail) > 5 else None),
                    "orderFee": _number(detail[6] if len(detail) > 6 else None),
                    "importDutyRate": _number(detail[8] if len(detail) > 8 else None),
                    "purchaseCostRmb": purchase,
                    "untaxedPriceUsd": None,
                    "grossProfit": profit,
                    "grossMargin": margin,
                    "competitorUrl": _safe_url(row[15] if len(row) > 15 else None) or _safe_url(detail[15] if len(detail) > 15 else None),
                }))
        header = _header_indexes(row)
        if not _cost_header(header) or index + 1 >= len(rows):
            continue
        candidate = _cost_candidate(sheet.title, index + 2, header, rows[index + 1])
        if candidate is None:
            continue
        candidate["sku"] = _unique_sku(candidate["sku"], sheet.title, index + 2, seen)
        candidates.append(candidate)
    return candidates


def _monthly_rows(sheet, month: str) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    result: list[dict] = []
    for row_number, row in enumerate(rows[2:], start=3):
        sku = _text(row[1] if len(row) > 1 else None)
        name = _text(row[3] if len(row) > 3 else None)
        if not sku and not name:
            continue
        if sku in {"SKU", "货号"}:
            continue
        sku = sku or f"{month}-ROW-{row_number}"
        result.append({
            "month": month,
            "sku": sku,
            "name": name or sku,
            "orderQuantity": _number(row[4] if len(row) > 4 else None),
            "costRmb": _number(row[5] if len(row) > 5 else None),
            "status": _text(row[6] if len(row) > 6 else None),
            "usStatus": _text(row[7] if len(row) > 7 else None),
            "caStatus": _text(row[8] if len(row) > 8 else None),
        })
    return result


def extract_research_rows(workbook) -> tuple[list[dict], list[dict]]:
    candidates: list[dict] = []
    seen: set[str] = set()
    for sheet in workbook.worksheets:
        candidates.extend(_candidate_rows(sheet, seen))
    monthly_orders: list[dict] = []
    for sheet_name in workbook.sheetnames:
        match = re.fullmatch(r"(\d{1,2})月新品", sheet_name.strip())
        if match:
            monthly_orders.extend(_monthly_rows(workbook[sheet_name], f"{datetime.now().year}-{int(match.group(1)):02d}"))
    return candidates, monthly_orders


def build_report(config: ProjectConfig, workbook_path: Path) -> dict:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        candidates, monthly_orders = extract_research_rows(workbook)
    finally:
        workbook.close()

    margins = [row["grossMargin"] for row in candidates if row["grossMargin"] is not None]
    ordered = [row for row in monthly_orders if (row["orderQuantity"] or 0) > 0]
    modified_at = datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds")
    return {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": {"path": str(workbook_path.relative_to(config.data_root)).replace("\\", "/"), "modifiedAt": modified_at, "sha256": _fingerprint(workbook_path), "sheet": "多工作表"},
        "summary": {
            "candidateCount": len(candidates),
            "viableCandidateCount": sum(1 for margin in margins if margin >= 0.3),
            "averageGrossMargin": sum(margins) / len(margins) if margins else 0,
            "latestOrderMonth": max((row["month"] for row in ordered), default=None),
            "orderedSkuCount": len({row["sku"] for row in ordered}),
            "plannedUnits": int(sum(row["orderQuantity"] or 0 for row in ordered)),
            "monthCount": len({row["month"] for row in monthly_orders}),
        },
        "candidates": candidates,
        "monthlyOrders": monthly_orders,
    }


def run(config: ProjectConfig, db: StateDb) -> dict:
    relative = str(config.inventory_dashboard.get("new_product_research_workbook", "新品调研表8.13.xlsx"))
    source = (config.data_root / relative).resolve()
    if not source.exists():
        raise FileNotFoundError(f"新品调研源文件不存在: {source}")
    report = build_report(config, source)
    output = config.runtime_root / "reports" / "new_product_research.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "completed", "report": str(output), "summary": report["summary"]}
