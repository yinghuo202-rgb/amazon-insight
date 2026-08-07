from __future__ import annotations

import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from ..config import ProjectConfig
from ..sku import SkuNormalizer


URL_RE = re.compile(r"https?://[^\s]+", re.IGNORECASE)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _skus(normalizer: SkuNormalizer, values: list[object]) -> set[str]:
    result: set[str] = set()
    for value in values:
        match = normalizer.extract(value)
        if match:
            result.update(match.bases)
    return result


def _brief_kind(path: Path) -> str:
    name = path.stem.upper()
    has_main = "主图" in path.stem
    has_aplus = "A+" in name
    if has_main and has_aplus:
        return "combined"
    return "a_plus" if has_aplus else "main_image"


def _row_text(row: tuple) -> list[str]:
    return [_text(value) for value in row if _text(value) and not _text(value).startswith("=_xlfn.DISPIMG")]


def _panel_sections(matrix: list[tuple], kind: str) -> list[dict]:
    sections: list[dict] = []
    for header_index, header in enumerate(matrix):
        labels = {_text(value): index for index, value in enumerate(header) if _text(value)}
        if "文案" not in labels or "要求" not in labels:
            continue
        copy_index = labels["文案"]
        requirement_index = labels["要求"]
        size_index = labels.get("尺寸")
        for row_index in range(header_index + 1, min(len(matrix), header_index + 25)):
            row = matrix[row_index]
            first = _text(row[0] if row else None)
            if first in {"参考链接", "模版名称"} or "文案" in {_text(value) for value in row}:
                break
            copy = _text(row[copy_index] if len(row) > copy_index else None)
            requirement = _text(row[requirement_index] if len(row) > requirement_index else None)
            size = _text(row[size_index] if size_index is not None and len(row) > size_index else None)
            if not any((copy, requirement, size)):
                continue
            if not first:
                first = f"模块 {len(sections) + 1}"
            elif first.isdigit():
                first = f"主图 {first}"
            sections.append({
                "section": first,
                "channel": "a_plus" if kind == "a_plus" else "main_image",
                "size": size,
                "copy": copy,
                "requirement": requirement,
                "sourceRow": row_index + 1,
            })
    return sections


def _template_sections(matrix: list[tuple]) -> list[dict]:
    sections: list[dict] = []
    marker_rows = [index for index, row in enumerate(matrix) if "模版名称" in {_text(value) for value in row}]
    for position, start in enumerate(marker_rows):
        stop = marker_rows[position + 1] if position + 1 < len(marker_rows) else min(len(matrix), start + 40)
        block = matrix[start:stop]
        template = size = description = requirement = ""
        mode = ""
        for row in block:
            texts = _row_text(row)
            if not texts:
                continue
            if "模版名称" in texts:
                template = next((value for value in texts if value != "模版名称"), template)
                mode = "template"
            elif "图片尺寸" in texts:
                size = next((value for value in texts if value != "图片尺寸"), size)
                mode = "size"
            elif "图片描述" in texts:
                description = next((value for value in texts if value != "图片描述"), description)
                mode = "description"
            elif "要求" in texts:
                requirement = next((value for value in texts if value != "要求"), requirement)
                mode = "requirement"
            else:
                value = "\n".join(texts)
                if mode == "description":
                    description = "\n".join(filter(None, (description, value)))
                elif mode == "requirement":
                    requirement = "\n".join(filter(None, (requirement, value)))
        if any((template, size, description, requirement)):
            sections.append({
                "section": template or f"A+ 模块 {len(sections) + 1}",
                "channel": "a_plus",
                "size": size,
                "copy": description,
                "requirement": requirement,
                "sourceRow": start + 1,
            })
    return sections


def read_creative_briefs(config: ProjectConfig, normalizer: SkuNormalizer) -> tuple[list[dict], list[dict]]:
    folder = (config.data_root / config.inventory_dashboard.get("creative_brief_folder", "归档")).resolve()
    if not folder.exists():
        return [], [{"category": "creative_brief_folder_missing", "path": str(folder)}]
    briefs: list[dict] = []
    exceptions: list[dict] = []
    for path in sorted(folder.glob("*.xls")):
        exceptions.append({
            "category": "creative_brief_legacy_xls_skipped",
            "path": str(path.relative_to(config.data_root)).replace("\\", "/"),
            "details": "Legacy .xls files need conversion to .xlsx before structured import.",
        })
    for path in sorted(folder.glob("*.xlsx")):
        if path.name.startswith(("~$", ".~")):
            continue
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
        except Exception as error:
            exceptions.append({
                "category": "creative_brief_unreadable",
                "path": str(path.relative_to(config.data_root)).replace("\\", "/"),
                "details": str(error),
            })
            continue
        try:
            kind = _brief_kind(path)
            source_skus = _skus(normalizer, [path.stem])
            references: set[str] = set()
            sections: list[dict] = []
            facts: list[str] = []
            for worksheet in workbook.worksheets:
                if worksheet.title == "WpsReserved_CellImgList":
                    continue
                matrix = [
                    tuple(row)
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=min(worksheet.max_row, 220),
                        max_col=min(worksheet.max_column, 30),
                        values_only=True,
                    )
                ]
                flattened = [value for row in matrix for value in row if value not in (None, "")]
                source_skus.update(_skus(normalizer, flattened))
                for value in flattened:
                    references.update(URL_RE.findall(_text(value)))
                sheet_sections = [*_panel_sections(matrix, kind), *_template_sections(matrix)]
                sections.extend(sheet_sections)
                if "产品信息" in worksheet.title or not sheet_sections:
                    for row in matrix:
                        texts = _row_text(row)
                        if texts:
                            facts.append(" | ".join(texts))
                        if len(facts) >= 60:
                            break
            relative_path = str(path.relative_to(config.data_root)).replace("\\", "/")
            digest = hashlib.sha256(relative_path.encode("utf-8")).hexdigest()[:16]
            briefs.append({
                "briefId": f"brief-{digest}",
                "name": path.stem,
                "kind": kind,
                "skus": sorted(source_skus),
                "references": sorted(references),
                "sections": sections,
                "productFacts": facts,
                "sourcePath": relative_path,
                "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds"),
            })
        finally:
            workbook.close()
    return briefs, exceptions
