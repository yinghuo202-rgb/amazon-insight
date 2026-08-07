from __future__ import annotations

import hashlib
import json
import mimetypes
import posixpath
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils import column_index_from_string

from ..config import ProjectConfig
from ..db import StateDb
from ..sku import SkuNormalizer


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _base_sku(normalizer: SkuNormalizer, value: object) -> str | None:
    match = normalizer.extract(value)
    return match.bases[0] if match and len(match.bases) == 1 else None


def _fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_meta(config: ProjectConfig, path: Path, kind: str) -> dict:
    return {
        "kind": kind,
        "path": str(path.relative_to(config.data_root)).replace("\\", "/"),
        "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds"),
        "sha256": _fingerprint(path),
    }


def _header_indexes(header: tuple) -> dict[str, int]:
    return {_text(value): index for index, value in enumerate(header) if _text(value)}


_DISPLAY_IMAGE_ID = re.compile(r'DISPIMG\(\s*"([^"]+)"', re.IGNORECASE)


def _read_sku_image_ids(
    path: Path,
    sheet_name: str,
    image_column: str,
    normalizer: SkuNormalizer,
) -> dict[str, str]:
    """Read WPS DISPIMG formulas without evaluating or modifying the workbook."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = _header_indexes(header)
        sku_index = indexes.get("MEASUREMAN", indexes.get("MSKU"))
        if sku_index is None:
            raise ValueError(f"Product details sheet is missing SKU column: {sheet_name}")
        image_index = column_index_from_string(image_column) - 1
        result: dict[str, str] = {}
        for row in rows:
            sku = _base_sku(normalizer, row[sku_index] if len(row) > sku_index else None)
            formula = _text(row[image_index] if len(row) > image_index else None)
            match = _DISPLAY_IMAGE_ID.search(formula)
            if sku and match and sku not in result:
                result[sku] = match.group(1)
        return result
    finally:
        workbook.close()


def _cell_image_targets(archive: zipfile.ZipFile) -> dict[str, str]:
    """Map WPS cell image IDs to safe OOXML archive members."""
    relationships = ElementTree.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
    relationship_targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships
        if relationship.attrib.get("Id") and relationship.attrib.get("Target")
    }
    cell_images = ElementTree.fromstring(archive.read("xl/cellimages.xml"))
    result: dict[str, str] = {}
    for cell_image in cell_images.findall(".//{*}cellImage"):
        properties = cell_image.find(".//{*}cNvPr")
        blip = cell_image.find(".//{*}blip")
        if properties is None or blip is None:
            continue
        image_id = properties.attrib.get("name", "")
        relationship_id = next((value for key, value in blip.attrib.items() if key.endswith("}embed")), "")
        target = relationship_targets.get(relationship_id, "")
        member = posixpath.normpath(posixpath.join("xl", target))
        if image_id and member.startswith("xl/media/") and member in archive.namelist():
            result[image_id] = member
    return result


def _extract_product_images(
    workbook_path: Path,
    sheet_name: str,
    image_column: str,
    output_root: Path,
    normalizer: SkuNormalizer,
) -> dict[str, dict[str, str]]:
    image_ids = _read_sku_image_ids(workbook_path, sheet_name, image_column, normalizer)
    output_root.mkdir(parents=True, exist_ok=True)
    result: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(workbook_path) as archive:
        targets = _cell_image_targets(archive)
        for sku, image_id in image_ids.items():
            member = targets.get(image_id)
            if not member:
                continue
            content = archive.read(member)
            digest = hashlib.sha256(content).hexdigest()
            suffix = Path(member).suffix.lower()
            if suffix == ".jpg":
                suffix = ".jpeg"
            mime_type = mimetypes.types_map.get(suffix, "application/octet-stream")
            if not mime_type.startswith("image/"):
                continue
            filename = f"{sku}-{digest[:16]}{suffix}"
            destination = output_root / filename
            if not destination.exists() or destination.read_bytes() != content:
                temporary = destination.with_suffix(f"{destination.suffix}.tmp")
                temporary.write_bytes(content)
                temporary.replace(destination)
            result[sku] = {
                "imageFile": filename,
                "imageMimeType": mime_type,
                "imageSha256": digest,
            }
    return result


def _read_product_details(path: Path, sheet_name: str, normalizer: SkuNormalizer) -> dict[str, dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = _header_indexes(header)
        is_current_layout = "MEASUREMAN" in indexes
        sku_index = indexes["MEASUREMAN"] if is_current_layout else indexes["MSKU"]
        fnsku_index = indexes.get("FUSKU", indexes.get("FNSKU"))
        result: dict[str, dict] = {}
        for row in rows:
            sku = _base_sku(normalizer, row[sku_index] if len(row) > sku_index else None)
            if not sku:
                continue
            if is_current_layout:
                carton_quantity = _number(row[indexes["装箱量"]] if len(row) > indexes["装箱量"] else None)
                dimensions = {
                    "length": _number(row[9] if len(row) > 9 else None),
                    "width": _number(row[10] if len(row) > 10 else None),
                    "height": _number(row[11] if len(row) > 11 else None),
                }
                item = {
                    "sku": sku,
                    "fnsku": "",
                    "productDescription": _text(row[indexes["中文品名"]] if len(row) > indexes["中文品名"] else None),
                    "category": _text(row[indexes["品类"]] if len(row) > indexes["品类"] else None),
                    "packaging": _text(row[indexes["包装方式"]] if len(row) > indexes["包装方式"] else None),
                    "hsCode": "",
                    "englishName": "",
                    "chineseName": _text(row[indexes["中文品名"]] if len(row) > indexes["中文品名"] else None),
                    "declarationElements": "",
                    "declarationElementsEnglish": "",
                    "cartonQty": int(round(carton_quantity)) if carton_quantity and carton_quantity > 0 else None,
                    "cartonNetWeightKg": _number(row[indexes["净重kg"]] if len(row) > indexes["净重kg"] else None),
                    "cartonGrossWeightKg": _number(row[indexes["毛重kg"]] if len(row) > indexes["毛重kg"] else None),
                    "cartonDimensionsCm": dimensions,
                    "cartonVolumeM3": _number(row[indexes["体积m³"]] if len(row) > indexes["体积m³"] else None),
                    "productWeightG": _number(row[indexes["产品重量(G)"]] if len(row) > indexes["产品重量(G)"] else None),
                    "shippingSizeCm": _text(row[indexes["产品尺寸\n(cm)"]] if len(row) > indexes["产品尺寸\n(cm)"] else None),
                    "purchaseCostRmbTaxIncluded": _number(row[indexes["采购成本\n含税价\n（人民币）"]] if len(row) > indexes["采购成本\n含税价\n（人民币）"] else None),
                    "purchaseCostRmbTaxExcluded": _number(row[indexes["采购成本\n不含税价\n（人民币）"]] if len(row) > indexes["采购成本\n不含税价\n（人民币）"] else None),
                    "purchaseCostUsd": _number(row[indexes["成本-一店\n（不含税、美金）"]] if len(row) > indexes["成本-一店\n（不含税、美金）"] else None),
                    "imageFile": "",
                    "imageMimeType": "",
                    "imageSha256": "",
                    "listing": None,
                }
            else:
                carton_quantity = _number(row[11] if len(row) > 11 else None)
                item = {
                "sku": sku,
                "fnsku": _text(row[fnsku_index] if fnsku_index is not None and len(row) > fnsku_index else None),
                "productDescription": _text(row[4] if len(row) > 4 else None),
                "category": "",
                "packaging": _text(row[5] if len(row) > 5 else None),
                "hsCode": _text(row[6] if len(row) > 6 else None),
                "englishName": _text(row[7] if len(row) > 7 else None),
                "chineseName": _text(row[8] if len(row) > 8 else None),
                "declarationElements": _text(row[9] if len(row) > 9 else None),
                "declarationElementsEnglish": _text(row[10] if len(row) > 10 else None),
                "cartonQty": int(round(carton_quantity)) if carton_quantity and carton_quantity > 0 else None,
                "cartonNetWeightKg": _number(row[12] if len(row) > 12 else None),
                "cartonGrossWeightKg": _number(row[13] if len(row) > 13 else None),
                "cartonDimensionsCm": {
                    "length": _number(row[14] if len(row) > 14 else None),
                    "width": _number(row[15] if len(row) > 15 else None),
                    "height": _number(row[16] if len(row) > 16 else None),
                },
                "cartonVolumeM3": _number(row[17] if len(row) > 17 else None),
                "productWeightG": _number(row[18] if len(row) > 18 else None),
                "shippingSizeCm": _text(row[19] if len(row) > 19 else None),
                "purchaseCostRmbTaxIncluded": None,
                "purchaseCostRmbTaxExcluded": None,
                "purchaseCostUsd": None,
                "imageFile": "",
                "imageMimeType": "",
                "imageSha256": "",
                "listing": None,
            }
            result[sku] = item
        return result
    finally:
        workbook.close()


def _merge_product_details(primary: dict[str, dict], legacy: dict[str, dict]) -> dict[str, dict]:
    result = {sku: dict(item) for sku, item in legacy.items()}
    for sku, current in primary.items():
        prior = result.get(sku, {})
        merged = dict(prior)
        for key, value in current.items():
            if key == "cartonDimensionsCm":
                old_dimensions = prior.get(key, {})
                merged[key] = {
                    dimension: current[key].get(dimension) if current[key].get(dimension) is not None else old_dimensions.get(dimension)
                    for dimension in ("length", "width", "height")
                }
            elif value not in (None, "", [], {}):
                merged[key] = value
            elif key not in merged:
                merged[key] = value
        result[sku] = merged
    return result


def _read_listing(path: Path, excluded_sheets: set[str], normalizer: SkuNormalizer) -> dict[str, dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    result: dict[str, dict] = {}
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title in excluded_sheets or worksheet.max_row < 2:
                continue
            matrix = list(worksheet.iter_rows(values_only=True))
            labels = {_text(row[0]): index for index, row in enumerate(matrix) if row and _text(row[0])}
            sku_row = labels.get("货号")
            if sku_row is None:
                continue
            field_rows = {
                "brand": labels.get("品牌"),
                "title": labels.get("标题"),
                "description": labels.get("描述"),
                "workingPressure": labels.get("工作压力"),
                "usageFeatures": labels.get("产品使用特点"),
                "mainMaterial": labels.get("主体材质"),
                "materialFeatures": labels.get("主体材质的特点"),
                "inflationHeadType": labels.get("充气头的类型（螺旋、卡式或者其他）"),
                "threadSpecification": labels.get("螺纹规格"),
                "applicableTo": labels.get("适用于"),
                "sellingPointSummary": labels.get("卖点总结"),
            }
            bullet_rows = [index for label, index in labels.items() if label.startswith("五点描述")]
            engineering_rows = [index for label, index in labels.items() if label.startswith("工程部补充")]
            width = max((len(row) for row in matrix), default=0)
            for column in range(1, width):
                raw_sku = matrix[sku_row][column] if len(matrix[sku_row]) > column else None
                sku = _base_sku(normalizer, raw_sku)
                if not sku:
                    continue

                def value_at(row_index: int | None) -> str:
                    return _text(matrix[row_index][column]) if row_index is not None and len(matrix[row_index]) > column else ""

                listing = {
                    "brand": value_at(field_rows["brand"]),
                    "title": value_at(field_rows["title"]),
                    "description": value_at(field_rows["description"]),
                    "bullets": [value_at(index) for index in sorted(bullet_rows) if value_at(index)],
                    "attributes": {
                        key: value_at(row_index)
                        for key, row_index in field_rows.items()
                        if key not in {"brand", "title", "description"} and value_at(row_index)
                    },
                    "engineeringNotes": [value_at(index) for index in sorted(engineering_rows) if value_at(index)],
                    "sourceSheet": worksheet.title,
                }
                prior = result.get(sku)
                if prior is None or len(json.dumps(listing, ensure_ascii=False)) > len(json.dumps(prior, ensure_ascii=False)):
                    result[sku] = listing
        return result
    finally:
        workbook.close()


def run(config: ProjectConfig, db: StateDb) -> dict:
    run_id = db.start_run("build-product-catalog")
    try:
        settings = config.inventory_dashboard
        details_path = (config.data_root / settings["product_details_workbook"]).resolve()
        legacy_details_path = (config.data_root / settings["product_details_legacy_workbook"]).resolve() if settings.get("product_details_legacy_workbook") else None
        listing_path = (config.data_root / settings["listing_workbook"]).resolve()
        for path in (details_path, listing_path):
            if not path.exists():
                raise FileNotFoundError(path)
        normalizer = SkuNormalizer(config.sku_pattern, config.ignore_values)
        current_items = _read_product_details(details_path, settings.get("product_details_sheet", "Measureman"), normalizer)
        legacy_items = _read_product_details(
            legacy_details_path,
            settings.get("product_details_legacy_sheet", "Measureman"),
            normalizer,
        ) if legacy_details_path and legacy_details_path.exists() else {}
        items = _merge_product_details(current_items, legacy_items)
        image_output = config.runtime_root / settings.get("product_image_output", "output/product-images")
        images = _extract_product_images(
            details_path,
            settings.get("product_details_sheet", "Measureman"),
            settings.get("product_image_column", "D"),
            image_output,
            normalizer,
        )
        for sku, image in images.items():
            if sku in items:
                items[sku].update(image)
        listings = _read_listing(listing_path, set(settings.get("listing_exclude_sheets", [])), normalizer)
        for sku, listing in listings.items():
            item = items.setdefault(sku, {
                "sku": sku,
                "fnsku": "",
                "productDescription": "",
                "category": "",
                "packaging": "",
                "hsCode": "",
                "englishName": "",
                "chineseName": "",
                "declarationElements": "",
                "declarationElementsEnglish": "",
                "cartonQty": None,
                "cartonNetWeightKg": None,
                "cartonGrossWeightKg": None,
                "cartonDimensionsCm": {"length": None, "width": None, "height": None},
                "cartonVolumeM3": None,
                "productWeightG": None,
                "shippingSizeCm": "",
                "purchaseCostRmbTaxIncluded": None,
                "purchaseCostRmbTaxExcluded": None,
                "purchaseCostUsd": None,
                "imageFile": "",
                "imageMimeType": "",
                "imageSha256": "",
                "listing": None,
            })
            item["listing"] = listing
        payload = {
            "schemaVersion": 1,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "sources": [
                _source_meta(config, details_path, "product_details"),
                *([_source_meta(config, legacy_details_path, "product_details_legacy")] if legacy_details_path and legacy_details_path.exists() else []),
                _source_meta(config, listing_path, "listing"),
            ],
            "items": [items[sku] for sku in sorted(items)],
        }
        report = config.runtime_root / "reports" / "product_catalog.json"
        report.parent.mkdir(parents=True, exist_ok=True)
        temporary = report.with_suffix(".json.tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(report)
        summary = {
            "run_id": run_id,
            "report_path": str(report),
            "sku_count": len(items),
            "listing_count": sum(1 for item in items.values() if item["listing"]),
            "specification_count": sum(1 for item in items.values() if item["productDescription"] or item["cartonQty"]),
            "image_count": sum(1 for item in items.values() if item["imageFile"]),
            "image_output": str(image_output),
        }
        db.finish_run(run_id, "completed", summary=summary)
        return summary
    except Exception as exc:
        db.finish_run(run_id, "failed", error=repr(exc))
        raise
