from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils.cell import column_index_from_string

from .sku import SkuNormalizer


@dataclass(frozen=True)
class SourceRecord:
    source: str
    raw: str
    cell: str
    fnsku: str | None = None
    status_marker: str | None = None


def _text(value: object) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def read_source(path: Path, source: dict, normalizer: SkuNormalizer) -> Iterator[SourceRecord]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        mode = source["mode"]
        if mode == "column":
            ws = wb[source["sheet"]]
            sku_col = column_index_from_string(source["column"])
            fnsku_col = column_index_from_string(source["fnsku_column"]) if source.get("fnsku_column") else None
            status_col = column_index_from_string(source["status_column"]) if source.get("status_column") else None
            max_col = max(sku_col, fnsku_col or 0, status_col or 0)
            for row in ws.iter_rows(min_row=source.get("start_row",1), max_col=max_col, values_only=False):
                cell = row[sku_col-1]
                raw = _text(cell.value)
                if raw:
                    yield SourceRecord(source["name"],raw,f"{ws.title}!{cell.coordinate}",_text(row[fnsku_col-1].value) if fnsku_col else None,_text(row[status_col-1].value) if status_col else None)
        elif mode == "transposed":
            excluded = set(source.get("exclude_sheets",[]))
            for ws in wb.worksheets:
                if ws.title in excluded:
                    continue
                for row_number in source.get("scan_rows",[1,2,3]):
                    for row in ws.iter_rows(min_row=row_number,max_row=row_number,min_col=2):
                        for cell in row:
                            if normalizer.extract(cell.value):
                                yield SourceRecord(f"{source['name']}-{ws.title}",str(cell.value).strip(),f"{ws.title}!{cell.coordinate}")
        elif mode == "scan":
            ws = wb[source["sheet"]]
            for row in ws.iter_rows():
                for cell in row:
                    if normalizer.extract(cell.value):
                        yield SourceRecord(source["name"],str(cell.value).strip(),f"{ws.title}!{cell.coordinate}")
        else:
            raise ValueError(f"Unsupported source mode: {mode}")
    finally:
        wb.close()
