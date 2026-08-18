from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl

from .advertising import AdvertisingParameters, recommend_campaign
from .jobs.new_product_research import extract_research_rows
from .jobs.product_catalog import _extract_product_images, _merge_product_details, _read_product_details
from .replenishment import ReplenishmentParameters, calculate_replenishment
from .sku import SkuNormalizer


TYPE_LABELS = {
    "inventory": "库存规划",
    "research": "新品调研",
    "product_details": "产品明细",
    "product_cost": "产品成本",
    "sales_us": "美国销售报告",
    "sales_ca": "加拿大销售报告",
    "advertising": "广告活动报告",
    "monthly_analysis": "月度分析",
    "unknown": "未识别文件",
}


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_url(value: object) -> str:
    raw = _text(value)
    match = re.match(r"(https?://[^?\s]+)", raw)
    return match.group(1) if match else ""


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {_text(value).replace("\n", "").replace(" ", ""): index for index, value in enumerate(row) if _text(value)}


def _field(indexes: dict[str, int], *names: str) -> int | None:
    for name in names:
        key = name.replace("\n", "").replace(" ", "")
        if key in indexes:
            return indexes[key]
    return None


def _inventory_snapshot_date(path: Path) -> date:
    match = re.search(r"(20\d{6})", path.name)
    if match:
        return datetime.strptime(match.group(1), "%Y%m%d").date()
    return datetime.fromtimestamp(path.stat().st_mtime).date()


def _completed_month(snapshot_date: date) -> str:
    return (snapshot_date.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")


def _base_sku(value: object) -> str | None:
    match = re.search(r"(?<![A-Z0-9])([A-Z]{2})\s*[-_]?\s*(\d{3})(?!\d)", _text(value).upper())
    return f"{match.group(1)}{match.group(2)}" if match else None


def _detect(path: Path, workbook) -> str:
    name = path.name.lower()
    sheets = {value.strip() for value in workbook.sheetnames}
    if "新品调研" in name or any(re.fullmatch(r"\d{1,2}月新品", value) for value in sheets):
        return "research"
    if "库存规划" in name or "库存规划" in sheets:
        return "inventory"
    if "产品总成本" in name or {"美国", "加拿大"}.issubset(sheets):
        return "product_cost"
    if "产品明细" in name or "一店" in sheets and path.stat().st_size > 20 * 1024 * 1024:
        return "product_details"
    if "广告活动" in name or "按月" in sheets:
        return "advertising"
    if "月度分析" in name or "仓租详情" in " ".join(sheets):
        return "monthly_analysis"
    if "report" in name and "Report" in workbook.sheetnames:
        return "sales_ca" if "-ca-" in name else "sales_us"
    return "unknown"


def _inventory_preview(workbook) -> dict[str, Any]:
    markets: dict[str, Any] = {}
    for market, sheet_name in (("US", "库存规划"), ("CA", "加拿大库存计划 ")):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        indexes = _header_map(header)
        sku_index = _field(indexes, "SKU")
        fba_index = _field(indexes, "FBA库存")
        local_index = _field(indexes, "工厂库存及已下订单")
        sales_index = _field(indexes, "最近月销售", "最近月销")
        valid = 0
        total_fba = 0
        total_local = 0
        for row in rows:
            if sku_index is None or not _text(row[sku_index] if len(row) > sku_index else None):
                continue
            valid += 1
            total_fba += int(_number(row[fba_index] if fba_index is not None and len(row) > fba_index else None) or 0)
            total_local += int(_number(row[local_index] if local_index is not None and len(row) > local_index else None) or 0)
        markets[market] = {"sheet": sheet_name, "skuCount": valid, "fbaUnits": total_fba, "domesticUnits": total_local, "hasRecentSales": sales_index is not None}
    return {"markets": markets, "impacts": ["运营总览", "库存视图", "采购计划"]}


def _research_preview(workbook) -> dict[str, Any]:
    candidate_rows, monthly_orders = extract_research_rows(workbook)
    candidates = len(candidate_rows)
    viable = sum(1 for row in candidate_rows if (row.get("grossMargin") or 0) >= 0.3)
    month_sheets = sorted([value for value in workbook.sheetnames if re.fullmatch(r"\d{1,2}月新品", value.strip())])
    return {"candidateCount": candidates, "viableCandidateCount": viable, "monthlyOrderCount": len(monthly_orders), "monthSheets": month_sheets, "impacts": ["新品调研"]}


def _sales_preview(workbook, market: str) -> dict[str, Any]:
    sheet = workbook["Report"]
    header = tuple(cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2)))
    indexes = _header_map(header)
    sku_index = _field(indexes, "SKU")
    units_index = _field(indexes, "售出的商品总数")
    revenue_index = _field(indexes, "净销售额", "总销售额")
    count = 0
    units = 0
    revenue = 0.0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        sku = _text(row[sku_index] if sku_index is not None and len(row) > sku_index else None)
        if not sku or sku.lower() in {"total", "总计"}:
            continue
        count += 1
        units += int(_number(row[units_index] if units_index is not None and len(row) > units_index else None) or 0)
        revenue += _number(row[revenue_index] if revenue_index is not None and len(row) > revenue_index else None) or 0
    return {"market": market, "skuCount": count, "units": units, "revenue": round(revenue, 2), "impacts": ["运营总览", "库存销量"]}


def _generic_preview(workbook, kind: str) -> dict[str, Any]:
    sheets = [{"name": name, "rows": workbook[name].max_row, "columns": workbook[name].max_column} for name in workbook.sheetnames if name != "WpsReserved_CellImgList"]
    impacts = {
        "product_cost": ["产品成本", "利润分析"],
        "product_details": ["产品目录", "产品待办"],
        "advertising": ["广告管理", "运营总览"],
        "monthly_analysis": ["运营总览", "仓储分析"],
    }.get(kind, [])
    return {"sheets": sheets[:12], "impacts": impacts}


def inspect_batch(batch_dir: Path) -> dict[str, Any]:
    files: list[dict[str, Any]] = []
    warnings: list[str] = []
    for path in sorted((batch_dir / "source").iterdir()):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        item: dict[str, Any] = {"name": path.name, "size": path.stat().st_size, "sha256": _sha256(path), "type": "unknown", "label": TYPE_LABELS["unknown"], "publishable": False}
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
            try:
                kind = _detect(path, workbook)
                item.update({"type": kind, "label": TYPE_LABELS[kind], "publishable": kind in {"inventory", "research", "advertising", "product_details"}, "sheets": workbook.sheetnames})
                if kind == "inventory":
                    item["preview"] = _inventory_preview(workbook)
                elif kind == "research":
                    item["preview"] = _research_preview(workbook)
                elif kind == "sales_us":
                    item["preview"] = _sales_preview(workbook, "US")
                elif kind == "sales_ca":
                    item["preview"] = _sales_preview(workbook, "CA")
                else:
                    item["preview"] = _generic_preview(workbook, kind)
            finally:
                workbook.close()
        except Exception as error:
            item["error"] = str(error)
            warnings.append(f"{path.name}: {error}")
        files.append(item)
    recognized = sum(1 for item in files if item["type"] != "unknown")
    manifest = {
        "schemaVersion": 1,
        "batchId": batch_dir.name,
        "createdAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "status": "ready" if recognized else "needs_review",
        "summary": {"fileCount": len(files), "recognizedCount": recognized, "unknownCount": len(files) - recognized, "publishableCount": sum(1 for item in files if item["publishable"])},
        "warnings": warnings,
        "files": files,
    }
    (batch_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def _atomic_json(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def _research_report(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        candidates, monthly = extract_research_rows(workbook)
    finally:
        workbook.close()
    margins = [item["grossMargin"] for item in candidates if item["grossMargin"] is not None]
    ordered = [item for item in monthly if (item["orderQuantity"] or 0) > 0]
    return {"schemaVersion": 1, "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"), "source": {"path": path.name, "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(timespec="seconds"), "sha256": _sha256(path), "sheet": "多工作表"}, "summary": {"candidateCount": len(candidates), "viableCandidateCount": sum(1 for value in margins if value >= 0.3), "averageGrossMargin": sum(margins) / len(margins) if margins else 0, "latestOrderMonth": max((item["month"] for item in ordered), default=None), "orderedSkuCount": len({item["sku"] for item in ordered}), "plannedUnits": int(sum(item["orderQuantity"] or 0 for item in ordered)), "monthCount": len({item["month"] for item in monthly})}, "candidates": candidates, "monthlyOrders": monthly}


def _patch_inventory(path: Path, reports_dir: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    updated: list[str] = []
    try:
        snapshot_date = _inventory_snapshot_date(path)
        sales_month = _completed_month(snapshot_date)
        for market, sheet_name, report_name in (("US", "库存规划", "inventory_dashboard.json"), ("CA", "加拿大库存计划 ", "inventory_dashboard.ca.json")):
            report_path = reports_dir / report_name
            if sheet_name not in workbook.sheetnames or not report_path.exists():
                continue
            report = json.loads(report_path.read_text(encoding="utf-8"))
            rows = workbook[sheet_name].iter_rows(values_only=True)
            indexes = _header_map(next(rows))
            sku_i = _field(indexes, "SKU")
            name_i = _field(indexes, "品名")
            fba_i = _field(indexes, "FBA库存")
            network_i = _field(indexes, "FBA+在途库存")
            local_i = _field(indexes, "工厂库存及已下订单")
            sales_i = _field(indexes, "最近月销售", "最近月销")
            incoming: dict[str, tuple[Any, ...]] = {}
            for row in rows:
                sku = _text(row[sku_i] if sku_i is not None and len(row) > sku_i else None).upper()
                if sku:
                    incoming[sku] = row
            raw_parameters = report.setdefault("parameters", {})
            raw_parameters["targetCoverDays"] = 90
            parameters = ReplenishmentParameters(
                lead_time_days=int(raw_parameters.get("leadTimeDays", 75)),
                review_cycle_days=int(raw_parameters.get("reviewCycleDays", 7)),
                target_cover_days=90,
                safety_stock_days=int(raw_parameters.get("safetyStockDays", 21)),
                excess_cover_days=int(raw_parameters.get("excessCoverDays", 240)),
                fba_transfer_trigger_days=int(raw_parameters.get("fbaTransferTriggerDays", 30)),
            )
            for item in report.get("rows", []):
                row = incoming.get(_text(item.get("sku")).upper())
                if not row:
                    continue
                if name_i is not None and _text(row[name_i]):
                    item["productName"] = _text(row[name_i])
                fba_value = _number(row[fba_i]) if fba_i is not None else None
                if fba_value is not None:
                    item["fbaSellable"] = max(0, int(fba_value))
                network_value = _number(row[network_i]) if network_i is not None else None
                if network_value is not None:
                    item["inTransitInventory"] = max(0, int(network_value) - int(item.get("fbaSellable", 0) or 0))
                else:
                    item["inTransitInventory"] = max(0, int(item.get("inTransitInventory", item.get("awdInbound", 0)) or 0))
                if local_i is not None:
                    item["localInventory"] = max(0, int(_number(row[local_i]) or 0))
                    item["domesticSupplyTotal"] = item["localInventory"] + int(item.get("pendingOrderQty", 0))
                if sales_i is not None and _number(row[sales_i]) is not None:
                    units = max(0, int(_number(row[sales_i]) or 0))
                    item["dailySales"] = units / 30
                    history = [entry for entry in item.get("salesHistoryByMonth", []) if entry.get("month") != sales_month]
                    item["salesHistoryByMonth"] = [*history, {"month": sales_month, "units": units}]
                    item["salesByMonth"] = [{"month": sales_month, "units": units}]
                decision = calculate_replenishment(
                    daily_sales=float(item.get("dailySales", 0) or 0),
                    fba_sellable=int(item.get("fbaSellable", 0) or 0),
                    awd_available=int(item.get("awdAvailable", 0) or 0),
                    awd_outbound_to_fba=int(item.get("awdOutboundToFba", 0) or 0),
                    carton_quantity=item.get("cartonQty"),
                    parameters=parameters,
                    in_transit_inventory=int(item.get("inTransitInventory", 0) or 0),
                )
                item.update(decision)
                item["readyToShipQty"] = min(int(item.get("localInventory", 0) or 0), int(item["suggestedShipmentQty"]))
                item["suggestedProductionQty"] = max(0, int(item["suggestedShipmentQty"]) - int(item.get("domesticSupplyTotal", 0) or 0))
            report["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            snapshots = report.setdefault("snapshots", {})
            snapshots["fbaDate"] = snapshot_date.isoformat()
            snapshots.setdefault("awdSourceAvailable", market == "US")
            snapshots.setdefault("awdDate", snapshot_date.isoformat())
            if not snapshots["awdSourceAvailable"]:
                snapshots["awdDate"] = snapshot_date.isoformat()
            awd_date = date.fromisoformat(str(snapshots.get("awdDate", snapshot_date.isoformat())))
            latest_snapshot = max(snapshot_date, awd_date)
            snapshots["alignmentGapDays"] = abs((snapshot_date - awd_date).days)
            snapshots["aligned"] = snapshots["alignmentGapDays"] <= 2
            snapshots["ageDays"] = max(0, (date.today() - latest_snapshot).days)
            snapshots.setdefault("staleAfterDays", 14)
            snapshots["isStale"] = snapshots["ageDays"] > int(snapshots["staleAfterDays"])
            report.setdefault("sales", {})["windowMonths"] = [sales_month]
            summary = report.setdefault("summary", {})
            summary["fbaSellable"] = sum(int(item.get("fbaSellable", 0)) for item in report["rows"])
            summary["inTransitInventory"] = sum(int(item.get("inTransitInventory", 0)) for item in report["rows"])
            summary["localInventory"] = sum(int(item.get("localInventory", 0)) for item in report["rows"])
            summary["readyToShipQty"] = sum(int(item.get("readyToShipQty", 0)) for item in report["rows"])
            summary["suggestedProductionQty"] = sum(int(item.get("suggestedProductionQty", 0)) for item in report["rows"])
            summary["suggestedShipmentQty"] = sum(int(item.get("suggestedShipmentQty", 0)) for item in report["rows"])
            report.setdefault("localRefresh", {})["uploadedBatch"] = path.name
            _atomic_json(report_path, report)
            updated.append(report_name)
    finally:
        workbook.close()
    return updated


def _patch_advertising(path: Path, reports_dir: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        indexes = _header_map(next(rows))
        required = {
            "start": _field(indexes, "数据开始时间"),
            "end": _field(indexes, "数据结束时间"),
            "market": _field(indexes, "国家"),
            "campaign": _field(indexes, "广告活动"),
            "status": _field(indexes, "状态"),
            "budget": _field(indexes, "预算"),
            "impressions": _field(indexes, "曝光量"),
            "clicks": _field(indexes, "点击量"),
            "spend": _field(indexes, "花费"),
            "sales": _field(indexes, "广告总销售额"),
            "orders": _field(indexes, "广告总订单量"),
        }
        missing = [name for name, index in required.items() if index is None]
        if missing:
            raise ValueError(f"广告报表缺少字段: {', '.join(missing)}")
        required_indexes = {name: int(index) for name, index in required.items()}
        for row in rows:
            market = _text(row[required_indexes["market"]]).upper()
            start = _text(row[required_indexes["start"]])[:10]
            campaign = _text(row[required_indexes["campaign"]])
            if market not in {"US", "CA"} or not re.fullmatch(r"20\d{2}-\d{2}-\d{2}", start) or not campaign:
                continue
            month = start[:7]
            key = (market, month)
            bucket = grouped.setdefault(key, {
                "month": month,
                "spend": 0.0,
                "advertisingSales": 0.0,
                "orders": 0,
                "clicks": 0,
                "impressions": 0,
                "campaigns": {},
                "endDate": start,
            })
            spend = float(_number(row[required_indexes["spend"]]) or 0)
            sales = float(_number(row[required_indexes["sales"]]) or 0)
            orders = int(_number(row[required_indexes["orders"]]) or 0)
            clicks = int(_number(row[required_indexes["clicks"]]) or 0)
            impressions = int(_number(row[required_indexes["impressions"]]) or 0)
            end = _text(row[required_indexes["end"]])[:10]
            if re.fullmatch(r"20\d{2}-\d{2}-\d{2}", end):
                bucket["endDate"] = max(bucket["endDate"], end)
            bucket["spend"] += spend
            bucket["advertisingSales"] += sales
            bucket["orders"] += orders
            bucket["clicks"] += clicks
            bucket["impressions"] += impressions
            item = bucket["campaigns"].setdefault(campaign, {
                "campaign": campaign,
                "sku": _base_sku(campaign),
                "status": _text(row[required_indexes["status"]]),
                "budget": float(_number(row[required_indexes["budget"]]) or 0),
                "spend": 0.0,
                "advertisingSales": 0.0,
                "orders": 0,
                "clicks": 0,
                "impressions": 0,
                "periodDays": 1,
            })
            item["spend"] += spend
            item["advertisingSales"] += sales
            item["orders"] += orders
            item["clicks"] += clicks
            item["impressions"] += impressions
            item["periodDays"] = max(item["periodDays"], (date.fromisoformat(bucket["endDate"]) - date.fromisoformat(f"{month}-01")).days + 1)
    finally:
        workbook.close()

    updated: list[str] = []
    for market in ("US", "CA"):
        market_months = sorted(month for report_market, month in grouped if report_market == market)
        if not market_months:
            continue
        report_name = "inventory_dashboard.json" if market == "US" else "inventory_dashboard.ca.json"
        report_path = reports_dir / report_name
        if not report_path.exists():
            continue
        report = json.loads(report_path.read_text(encoding="utf-8"))
        latest_month = market_months[-1]
        advertising = report.setdefault("advertising", {})
        raw_params = advertising.get("parameters", {})
        parameters = AdvertisingParameters(
            target_acos_percent=float(raw_params.get("targetAcosPercent", 30)),
            minimum_evidence_spend=float(raw_params.get("minimumEvidenceSpend", 20)),
            no_order_spend=float(raw_params.get("noOrderSpend", 20)),
            winner_min_orders=int(raw_params.get("winnerMinOrders", 5)),
            scale_min_orders=int(raw_params.get("scaleMinOrders", 1)),
            low_volume_max_clicks=int(raw_params.get("lowVolumeMaxClicks", 30)),
            budget_utilization_threshold_percent=float(raw_params.get("budgetUtilizationThresholdPercent", 80)),
            scale_max_acos_ratio=float(raw_params.get("scaleMaxAcosRatio", 0.9)),
        )
        replacement_months = set(market_months)
        monthly = [item for item in advertising.get("monthlySeries", []) if item.get("month") not in replacement_months]
        for month in market_months:
            bucket = grouped[(market, month)]
            spend = round(float(bucket["spend"]), 2)
            sales = round(float(bucket["advertisingSales"]), 2)
            monthly.append({
                "month": month,
                "spend": spend,
                "advertisingSales": sales,
                "orders": int(bucket["orders"]),
                "clicks": int(bucket["clicks"]),
                "impressions": int(bucket["impressions"]),
                "acos": round(spend / sales * 100, 2) if sales > 0 else None,
                "roas": round(sales / spend, 2) if spend > 0 else None,
            })
        rows_by_sku = {str(item.get("sku")): item for item in report.get("rows", [])}
        campaign_rows = []
        for item in grouped[(market, latest_month)]["campaigns"].values():
            item["spend"] = round(float(item["spend"]), 2)
            item["advertisingSales"] = round(float(item["advertisingSales"]), 2)
            inventory = rows_by_sku.get(str(item.get("sku")))
            inventory_risk = inventory.get("riskLevel") if inventory else None
            item["inventoryRisk"] = inventory_risk
            item["inventoryDaysCover"] = inventory.get("daysCoverNetwork") if inventory else None
            item.update(recommend_campaign(
                spend=item["spend"], advertising_sales=item["advertisingSales"], orders=int(item["orders"]),
                clicks=int(item["clicks"]), impressions=int(item["impressions"]), budget=float(item["budget"]),
                period_days=int(item["periodDays"]), inventory_risk=inventory_risk, parameters=parameters,
            ))
            campaign_rows.append(item)
        advertising["latestMonth"] = latest_month
        advertising["monthlySeries"] = sorted(monthly, key=lambda item: str(item.get("month", "")))
        advertising["campaigns"] = sorted(campaign_rows, key=lambda item: (-float(item["spend"]), str(item["campaign"])))
        latest_end = date.fromisoformat(grouped[(market, latest_month)]["endDate"])
        snapshot = date.fromisoformat(str(report.get("snapshots", {}).get("fbaDate", latest_end.isoformat())))
        advertising["ageDaysAtSnapshot"] = max(0, (snapshot - latest_end).days)
        report["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        report["sources"] = [item for item in report.get("sources", []) if item.get("kind") != "uploaded_advertising_campaign_month"]
        report["sources"].append({
            "kind": "uploaded_advertising_campaign_month",
            "path": path.name,
            "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(timespec="seconds"),
            "sha256": _sha256(path),
        })
        _atomic_json(report_path, report)
        updated.append(report_name)
    return updated


def _patch_product_details(path: Path, reports_dir: Path) -> list[str]:
    report_path = reports_dir / "product_catalog.json"
    if not report_path.exists():
        return []
    normalizer = SkuNormalizer(r"(?<![A-Z0-9])([A-Z]{2}\s*[-_]?\s*\d{3})(?!\d)", frozenset({"", "SKU", "MSKU"}))
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = "一店" if "一店" in workbook.sheetnames else workbook.sheetnames[0]
    finally:
        workbook.close()
    current = _read_product_details(path, sheet_name, normalizer)
    report = json.loads(report_path.read_text(encoding="utf-8"))
    existing = {str(item.get("sku")): item for item in report.get("items", [])}
    merged = _merge_product_details(current, existing)
    image_output = reports_dir.parent / "output" / "product-images"
    images = _extract_product_images(path, sheet_name, "B", image_output, normalizer)
    for sku, image in images.items():
        if sku in merged:
            merged[sku].update(image)
    report["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    report["sources"] = [item for item in report.get("sources", []) if item.get("kind") != "product_details"]
    report["sources"].insert(0, {
        "kind": "product_details",
        "path": path.name,
        "modifiedAt": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(timespec="seconds"),
        "sha256": _sha256(path),
    })
    report["items"] = [merged[sku] for sku in sorted(merged)]
    _atomic_json(report_path, report)
    return ["product_catalog.json"]


def publish_batch(batch_dir: Path, reports_dir: Path, snapshots_dir: Path) -> dict[str, Any]:
    manifest = json.loads((batch_dir / "manifest.json").read_text(encoding="utf-8"))
    version = f"data-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{batch_dir.name[-6:]}"
    snapshot = snapshots_dir / version / "reports"
    snapshot.mkdir(parents=True, exist_ok=False)
    if reports_dir.exists():
        for path in reports_dir.glob("*.json"):
            shutil.copy2(path, snapshot / path.name)
    reports_dir.mkdir(parents=True, exist_ok=True)
    updated: list[str] = []
    skipped: list[str] = []
    for item in manifest["files"]:
        source = batch_dir / "source" / item["name"]
        if item["type"] == "research":
            _atomic_json(reports_dir / "new_product_research.json", _research_report(source))
            updated.append("new_product_research.json")
        elif item["type"] == "inventory":
            updated.extend(_patch_inventory(source, reports_dir))
        elif item["type"] == "advertising":
            updated.extend(_patch_advertising(source, reports_dir))
        elif item["type"] == "product_details":
            updated.extend(_patch_product_details(source, reports_dir))
        else:
            skipped.append(item["name"])
    manifest.update({"status": "published", "publishedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"), "dataVersion": version, "updatedReports": sorted(set(updated)), "stagedFiles": skipped})
    (batch_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (snapshots_dir / version / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def restore_version(version: str, reports_dir: Path, snapshots_dir: Path) -> dict[str, Any]:
    if not re.fullmatch(r"data-\d{8}-\d{6}-[a-zA-Z0-9-]{1,20}", version):
        raise ValueError("数据版本编号不正确")
    source = snapshots_dir / version / "reports"
    if not source.is_dir():
        raise FileNotFoundError(f"数据版本不存在: {version}")
    backup = snapshots_dir / f"data-{datetime.now().strftime('%Y%m%d-%H%M%S')}-rollback" / "reports"
    backup.mkdir(parents=True, exist_ok=False)
    reports_dir.mkdir(parents=True, exist_ok=True)
    for path in reports_dir.glob("*.json"):
        shutil.copy2(path, backup / path.name)
    restored: list[str] = []
    for path in source.glob("*.json"):
        temporary = reports_dir / f"{path.name}.tmp"
        shutil.copy2(path, temporary)
        os.replace(temporary, reports_dir / path.name)
        restored.append(path.name)
    return {"status": "restored", "dataVersion": version, "restoredReports": sorted(restored), "restoredAt": datetime.now(timezone.utc).isoformat(timespec="seconds")}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["inspect", "publish", "restore"])
    parser.add_argument("--batch-dir")
    parser.add_argument("--reports-dir")
    parser.add_argument("--snapshots-dir")
    parser.add_argument("--version")
    args = parser.parse_args()
    if args.command == "inspect":
        if not args.batch_dir:
            parser.error("inspect requires --batch-dir")
        payload = inspect_batch(Path(args.batch_dir))
    elif args.command == "publish":
        if not args.batch_dir or not args.reports_dir or not args.snapshots_dir:
            parser.error("publish requires --batch-dir, --reports-dir and --snapshots-dir")
        payload = publish_batch(Path(args.batch_dir), Path(args.reports_dir), Path(args.snapshots_dir))
    else:
        if not args.version or not args.reports_dir or not args.snapshots_dir:
            parser.error("restore requires --version, --reports-dir and --snapshots-dir")
        payload = restore_version(str(args.version), Path(args.reports_dir), Path(args.snapshots_dir))
    print(json.dumps(payload, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
